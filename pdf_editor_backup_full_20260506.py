import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import PyPDF2
import os
import re
from collections import Counter
import openpyxl
import xlrd
import fitz  # PyMuPDF

class PDFEditor:
    def __init__(self, root):
        self.root = root
        self.root.title("PDF操作工具 made by xuxuquan")
        # 设置默认窗口大小
        self.root.geometry("1000x700")
        # 允许窗口调整大小
        self.root.resizable(True, True)
        
        # 添加最大化窗口的功能
        def maximize_window():
            self.root.state('zoomed')  # 最大化窗口，保留边框和任务栏
        
        # 添加最大化窗口的快捷键
        self.root.bind('<F11>', lambda e: maximize_window())
        # 添加恢复窗口大小的快捷键
        self.root.bind('<Escape>', lambda e: self.root.state('normal'))
        
        # 变量
        self.main_pdf = ""
        self.insert_pdfs = ["", "", "", "", ""]  # 5个PDF文件
        self.insert_interval = 1
        self.insert_pages = 1
        # 一分多变量
        self.split_pdf = ""
        self.split_pages = 1
        self.excel_file = ""
        self.naming_method = "auto"
        self.selected_header = None
        self.selected_id_header = None
        # PDF页面交换变量
        self.swap_pdfs = []
        self.page_a = 1
        self.page_b = 2
        # PDF批量重命名变量
        self.rename_pdfs = []
        self.rename_excel_file = ""
        self.rename_name_header = None
        self.rename_id_header = None
        self.filename_parts = []  # 文件名构建部分
        self.match_fields = []  # 匹配字段
        self.rename_mode_var = None  # 重命名模式变量
        self.class_rename_special_char_var = None  # 按班级重命名特殊符号变量
        # 指定导出页变量（支持批量处理）
        self.export_pdfs = []  # 支持多个PDF文件
        self.export_pages = ""
        self.export_excel_file = ""
        self.export_name_header = None
        self.export_id_header = None
        # PDF重排序变量
        self.reorder_pdfs = []
        self.reorder_excel_file = ""
        self.reorder_name_header = None
        self.reorder_id_header = None
        self.reorder_data = None
        # PDF页面变量
        self.page_pdfs = []
        self.page_rotation_scheme = []
        # 批量提取变量
        self.batch_extract_pdfs = []
        self.batch_extract_page = 1
        self.batch_extract_output_dir = ""
        # 批量提取Excel匹配变量
        self.batch_extract_excel_files = []
        self.batch_extract_name_header = None
        self.batch_extract_class_header = None
        # 批量提取姓名提取变量
        self.batch_extract_special_char_var = None
        self.batch_extract_extracted_names = {}
        # 批量追加页变量
        self.batch_append_pdfs = []
        self.batch_append_source_pdf = ""
        self.batch_append_output_dir = ""
        # 批量追加页Excel匹配变量
        self.batch_append_excel_files = []
        self.batch_append_name_header = None
        self.batch_append_class_header = None
        # 批量追加页姓名提取变量
        self.batch_append_special_char_var = None
        self.batch_append_extracted_names = {}
        # 学生班级PDF合并变量
        self.student_pdfs = []
        self.class_pdfs = []
        self.student_class_excel_file = ""
        self.student_class_name_header = None
        self.student_class_class_header = None
        self.student_class_output_dir = ""
        
        # 界面框架引用
        self.insert_frame = None
        self.split_frame = None
        self.swap_frame = None
        self.export_frame = None
        self.rename_frame = None
        self.reorder_frame = None
        self.page_frame = None
        self.batch_extract_frame = None
        self.batch_append_frame = None
        self.student_class_frame = None
        self.excel_to_pdf_frame = None
        self.word_to_pdf_frame = None
        # Excel匹配框架引用
        self.batch_extract_excel_frame = None
        self.batch_append_excel_frame = None
        
        # Excel批量导出PDF相关变量
        self.excel_files = []
        self.excel_output_dir = ""
        self.excel_files_var = tk.StringVar(value="未选择文件")
        self.excel_output_var = tk.StringVar(value="未选择目录")
        
        # Word批量导出PDF相关变量
        self.word_files = []
        self.word_output_dir = ""
        self.word_files_var = tk.StringVar(value="未选择文件")
        self.word_output_var = tk.StringVar(value="未选择目录")
        # Word转PDF页面选择变量
        self.word_export_pages_var = tk.StringVar(value="")
        self.word_export_all_pages_var = tk.BooleanVar(value=True)
        
        # Excel批量导出PDF增强变量
        self.excel_export_all_sheets_var = tk.BooleanVar(value=True)
        self.excel_selected_sheets = []
        
        # 创建界面
        self.create_widgets()
    
    def create_widgets(self):
        # 创建主滚动框架
        main_scroll_frame = ttk.Frame(self.root)
        main_scroll_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建滚动条
        scrollbar = ttk.Scrollbar(main_scroll_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建画布
        canvas = tk.Canvas(main_scroll_frame, yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 配置滚动条
        scrollbar.config(command=canvas.yview)
        
        # 主框架
        main_frame = ttk.Frame(canvas, padding="20")
        canvas.create_window((0, 0), window=main_frame, anchor=tk.NW)
        
        # 绑定鼠标滚轮事件
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # 保存画布和主框架引用
        self.canvas = canvas
        self.main_frame = main_frame
        
        # 更新滚动区域的函数
        def update_scroll_region():
            canvas.update_idletasks()
            canvas.config(scrollregion=canvas.bbox("all"))
        
        self.update_scroll_region = update_scroll_region
        
        # 标题
        title_label = ttk.Label(main_frame, text="PDF操作工具", font=('微软雅黑', 16, 'bold'))
        title_label.pack(pady=10)
        
        # 功能选择（放在最前面）
        function_frame = ttk.LabelFrame(main_frame, text="功能选择", padding="10")
        function_frame.pack(fill=tk.X, pady=5)
        
        self.function_var = tk.StringVar(value="insert")
        insert_radio = ttk.Radiobutton(function_frame, text="PDF插入", variable=self.function_var, value="insert", command=self.on_function_change)
        insert_radio.pack(side=tk.LEFT, padx=10)
        
        split_radio = ttk.Radiobutton(function_frame, text="PDF分割", variable=self.function_var, value="split", command=self.on_function_change)
        split_radio.pack(side=tk.LEFT, padx=10)
        
        swap_radio = ttk.Radiobutton(function_frame, text="PDF页面交换", variable=self.function_var, value="swap", command=self.on_function_change)
        swap_radio.pack(side=tk.LEFT, padx=10)
        
        export_radio = ttk.Radiobutton(function_frame, text="指定导出页", variable=self.function_var, value="export", command=self.on_function_change)
        export_radio.pack(side=tk.LEFT, padx=10)
        
        rename_radio = ttk.Radiobutton(function_frame, text="PDF批量重命名", variable=self.function_var, value="rename", command=self.on_function_change)
        rename_radio.pack(side=tk.LEFT, padx=10)
        
        reorder_radio = ttk.Radiobutton(function_frame, text="PDF重排序", variable=self.function_var, value="reorder", command=self.on_function_change)
        reorder_radio.pack(side=tk.LEFT, padx=10)
        
        page_radio = ttk.Radiobutton(function_frame, text="PDF页面", variable=self.function_var, value="page", command=self.on_function_change)
        page_radio.pack(side=tk.LEFT, padx=10)
        
        batch_extract_radio = ttk.Radiobutton(function_frame, text="批量提取", variable=self.function_var, value="batch_extract", command=self.on_function_change)
        batch_extract_radio.pack(side=tk.LEFT, padx=10)
        
        batch_append_radio = ttk.Radiobutton(function_frame, text="批量追加页", variable=self.function_var, value="batch_append", command=self.on_function_change)
        batch_append_radio.pack(side=tk.LEFT, padx=10)
        
        student_class_radio = ttk.Radiobutton(function_frame, text="学生班级PDF合并", variable=self.function_var, value="student_class", command=self.on_function_change)
        student_class_radio.pack(side=tk.LEFT, padx=10)
        
        excel_to_pdf_radio = ttk.Radiobutton(function_frame, text="Excel批量导出PDF", variable=self.function_var, value="excel_to_pdf", command=self.on_function_change)
        excel_to_pdf_radio.pack(side=tk.LEFT, padx=10)
        
        word_to_pdf_radio = ttk.Radiobutton(function_frame, text="Word批量导出PDF", variable=self.function_var, value="word_to_pdf", command=self.on_function_change)
        word_to_pdf_radio.pack(side=tk.LEFT, padx=10)
        
        # 使用说明按钮
        help_btn = ttk.Button(function_frame, text="使用说明", command=self.show_help)
        help_btn.pack(side=tk.RIGHT, padx=10)
        
        # PDF插入设置框架
        self.insert_frame = ttk.LabelFrame(main_frame, text="PDF插入设置", padding="10")
        self.insert_frame.pack(fill=tk.X, pady=5)
        
        # 主PDF选择
        main_pdf_frame = ttk.LabelFrame(self.insert_frame, text="主PDF文件", padding="5")
        main_pdf_frame.pack(fill=tk.X, pady=5)
        
        main_pdf_content = ttk.Frame(main_pdf_frame)
        main_pdf_content.pack(fill=tk.X, padx=5, pady=5)
        
        self.main_pdf_var = tk.StringVar()
        main_pdf_entry = ttk.Entry(main_pdf_content, textvariable=self.main_pdf_var, state="readonly", width=60)
        main_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        main_pdf_btn = ttk.Button(main_pdf_content, text="选择文件", command=self.select_main_pdf)
        main_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # 主PDF信息显示
        self.main_pdf_info_var = tk.StringVar(value="")
        main_pdf_info = ttk.Label(main_pdf_frame, textvariable=self.main_pdf_info_var, font=('Arial', 9), foreground='#666666')
        main_pdf_info.pack(side=tk.LEFT, padx=10, pady=2, anchor=tk.W)
        
        # 插入PDF选择（5个文件）
        insert_pdf_frame = ttk.LabelFrame(self.insert_frame, text="插入PDF文件", padding="5")
        insert_pdf_frame.pack(fill=tk.X, pady=5)
        
        # 5个PDF文件选择控件
        self.insert_pdf_vars = []
        self.insert_pdf_info_vars = []
        
        for i in range(5):
            pdf_frame = ttk.Frame(insert_pdf_frame)
            pdf_frame.pack(fill=tk.X, padx=5, pady=3)
            
            pdf_var = tk.StringVar()
            self.insert_pdf_vars.append(pdf_var)
            
            pdf_entry = ttk.Entry(pdf_frame, textvariable=pdf_var, state="readonly", width=60)
            pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
            
            pdf_btn = ttk.Button(pdf_frame, text=f"选择文件 {i+1}", command=lambda idx=i: self.select_single_insert_pdf(idx))
            pdf_btn.pack(side=tk.RIGHT, padx=5)
            
            # 每个PDF的信息显示
            pdf_info_var = tk.StringVar(value="")
            self.insert_pdf_info_vars.append(pdf_info_var)
            pdf_info = ttk.Label(pdf_frame, textvariable=pdf_info_var, font=('Arial', 9), foreground='#666666')
            pdf_info.pack(side=tk.LEFT, padx=10, pady=2, anchor=tk.W)
        
        # 插入设置
        settings_frame = ttk.Frame(self.insert_frame)
        settings_frame.pack(fill=tk.X, pady=5)
        
        # 插入间隔
        interval_frame = ttk.Frame(settings_frame)
        interval_frame.pack(fill=tk.X, pady=5)
        
        interval_label = ttk.Label(interval_frame, text="插入间隔（页）:")
        interval_label.pack(side=tk.LEFT, padx=5)
        
        self.interval_var = tk.StringVar(value="1")
        interval_entry = ttk.Entry(interval_frame, textvariable=self.interval_var, width=5)
        interval_entry.pack(side=tk.LEFT, padx=5)
        
        # 每次插入页数
        insert_pages_frame = ttk.Frame(settings_frame)
        insert_pages_frame.pack(fill=tk.X, pady=5)
        
        insert_pages_label = ttk.Label(insert_pages_frame, text="每次插入页数:")
        insert_pages_label.pack(side=tk.LEFT, padx=5)
        
        self.insert_pages_var = tk.StringVar(value="1")
        insert_pages_entry = ttk.Entry(insert_pages_frame, textvariable=self.insert_pages_var, width=5)
        insert_pages_entry.pack(side=tk.LEFT, padx=5)
        
        # 操作模式
        mode_frame = ttk.Frame(settings_frame)
        mode_frame.pack(fill=tk.X, pady=5)
        
        mode_label = ttk.Label(mode_frame, text="操作模式:")
        mode_label.pack(side=tk.LEFT, padx=5)
        
        self.mode_var = tk.StringVar(value="single")
        single_mode = ttk.Radiobutton(mode_frame, text="单个PDF插入（隔页）", variable=self.mode_var, value="single")
        single_mode.pack(side=tk.LEFT, padx=10)
        
        batch_mode = ttk.Radiobutton(mode_frame, text="批量PDF插入（每页依次）", variable=self.mode_var, value="batch")
        batch_mode.pack(side=tk.LEFT, padx=10)
        
        multi_mode = ttk.Radiobutton(mode_frame, text="多PDF顺序插入", variable=self.mode_var, value="multi")
        multi_mode.pack(side=tk.LEFT, padx=10)
        
        # 每次插入PDF数量
        multi_count_frame = ttk.Frame(settings_frame)
        multi_count_frame.pack(fill=tk.X, pady=5)
        
        multi_count_label = ttk.Label(multi_count_frame, text="每次插入PDF数量:")
        multi_count_label.pack(side=tk.LEFT, padx=5)
        
        self.multi_count_var = tk.StringVar(value="4")
        multi_count_entry = ttk.Entry(multi_count_frame, textvariable=self.multi_count_var, width=5)
        multi_count_entry.pack(side=tk.LEFT, padx=5)
        
        # 预览按钮
        insert_preview_btn = ttk.Button(self.insert_frame, text="预览插入结果", command=self.preview_insert)
        insert_preview_btn.pack(pady=5)
        
        # PDF分割设置框架
        self.split_frame = ttk.LabelFrame(main_frame, text="PDF分割设置", padding="10")
        self.split_frame.pack(fill=tk.X, pady=5)
        
        # 分割PDF选择
        split_pdf_frame = ttk.Frame(self.split_frame)
        split_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.split_pdf_var = tk.StringVar()
        split_pdf_entry = ttk.Entry(split_pdf_frame, textvariable=self.split_pdf_var, state="readonly", width=50)
        split_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        split_pdf_btn = ttk.Button(split_pdf_frame, text="选择文件", command=self.select_split_pdf)
        split_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # 分割页数
        split_pages_frame = ttk.Frame(self.split_frame)
        split_pages_frame.pack(fill=tk.X, pady=5)
        
        split_pages_label = ttk.Label(split_pages_frame, text="每X页分割:")
        split_pages_label.pack(side=tk.LEFT, padx=5)
        
        self.split_pages_var = tk.StringVar(value="1")
        split_pages_entry = ttk.Entry(split_pages_frame, textvariable=self.split_pages_var, width=5)
        split_pages_entry.pack(side=tk.LEFT, padx=5)
        
        # 命名方法
        naming_frame = ttk.Frame(self.split_frame)
        naming_frame.pack(fill=tk.X, pady=5)
        
        naming_label = ttk.Label(naming_frame, text="命名方法:")
        naming_label.pack(side=tk.LEFT, padx=5)
        
        self.naming_var = tk.StringVar(value="excel")
        excel_naming = ttk.Radiobutton(naming_frame, text="Excel表格命名", variable=self.naming_var, value="excel")
        excel_naming.pack(side=tk.LEFT, padx=10)
        
        # Excel文件选择
        excel_frame = ttk.Frame(self.split_frame)
        excel_frame.pack(fill=tk.X, pady=5)
        
        self.excel_var = tk.StringVar()
        excel_entry = ttk.Entry(excel_frame, textvariable=self.excel_var, state="readonly", width=50)
        excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        excel_btn = ttk.Button(excel_frame, text="选择Excel文件", command=self.select_excel_file)
        excel_btn.pack(side=tk.RIGHT, padx=5)
        
        # 表头选择按钮
        header_frame = ttk.Frame(self.split_frame)
        header_frame.pack(fill=tk.X, pady=5)
        
        self.header_var = tk.StringVar(value="未选择")
        header_entry = ttk.Entry(header_frame, textvariable=self.header_var, state="readonly", width=50)
        header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        header_btn = ttk.Button(header_frame, text="选择姓名表头", command=self.select_header)
        header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 序号表头选择按钮
        id_header_frame = ttk.Frame(self.split_frame)
        id_header_frame.pack(fill=tk.X, pady=5)
        
        self.id_header_var = tk.StringVar(value="未选择")
        id_header_entry = ttk.Entry(id_header_frame, textvariable=self.id_header_var, state="readonly", width=50)
        id_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        id_header_btn = ttk.Button(id_header_frame, text="选择序号表头", command=self.select_id_header)
        id_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 中转表按钮框架
        split_preview_frame = ttk.Frame(self.split_frame)
        split_preview_frame.pack(fill=tk.X, pady=5)
        
        # 打开中转表按钮
        split_open_preview_btn = ttk.Button(split_preview_frame, text="打开中转表预览", command=self.open_split_preview)
        split_open_preview_btn.pack(side=tk.LEFT, padx=5)
        
        # 预览对应关系按钮
        preview_btn = ttk.Button(split_preview_frame, text="预览对应关系", command=self.preview_mapping)
        preview_btn.pack(side=tk.LEFT, padx=5)
        
        # PDF页面交换设置框架
        self.swap_frame = ttk.LabelFrame(main_frame, text="PDF页面交换设置", padding="10")
        self.swap_frame.pack(fill=tk.X, pady=5)
        
        # 交换PDF选择
        swap_pdf_frame = ttk.Frame(self.swap_frame)
        swap_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.swap_pdf_var = tk.StringVar()
        swap_pdf_entry = ttk.Entry(swap_pdf_frame, textvariable=self.swap_pdf_var, state="readonly", width=50)
        swap_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        swap_pdf_btn = ttk.Button(swap_pdf_frame, text="选择文件", command=self.select_swap_pdfs)
        swap_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # 页面交换设置
        page_settings_frame = ttk.Frame(self.swap_frame)
        page_settings_frame.pack(fill=tk.X, pady=5)
        
        # 页面A
        page_a_frame = ttk.Frame(page_settings_frame)
        page_a_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        page_a_label = ttk.Label(page_a_frame, text="页面A:")
        page_a_label.pack(side=tk.LEFT, padx=5)
        
        self.page_a_var = tk.StringVar(value="1")
        page_a_entry = ttk.Entry(page_a_frame, textvariable=self.page_a_var, width=5)
        page_a_entry.pack(side=tk.LEFT, padx=5)
        
        # 页面B
        page_b_frame = ttk.Frame(page_settings_frame)
        page_b_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        page_b_label = ttk.Label(page_b_frame, text="页面B:")
        page_b_label.pack(side=tk.LEFT, padx=5)
        
        self.page_b_var = tk.StringVar(value="2")
        page_b_entry = ttk.Entry(page_b_frame, textvariable=self.page_b_var, width=5)
        page_b_entry.pack(side=tk.LEFT, padx=5)
        
        # 预览按钮
        swap_preview_btn = ttk.Button(self.swap_frame, text="预览交换结果", command=self.preview_swap)
        swap_preview_btn.pack(pady=5)
        
        # PDF批量重命名设置框架
        self.rename_frame = ttk.LabelFrame(main_frame, text="PDF批量重命名设置", padding="10")
        self.rename_frame.pack(fill=tk.X, pady=5)
        
        # 重命名PDF选择
        rename_pdf_frame = ttk.Frame(self.rename_frame)
        rename_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.rename_pdf_var = tk.StringVar()
        rename_pdf_entry = ttk.Entry(rename_pdf_frame, textvariable=self.rename_pdf_var, state="readonly", width=50)
        rename_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        rename_pdf_btn = ttk.Button(rename_pdf_frame, text="选择文件", command=self.select_rename_pdfs)
        rename_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # Excel文件选择
        rename_excel_frame = ttk.Frame(self.rename_frame)
        rename_excel_frame.pack(fill=tk.X, pady=5)
        
        self.rename_excel_var = tk.StringVar()
        rename_excel_entry = ttk.Entry(rename_excel_frame, textvariable=self.rename_excel_var, state="readonly", width=50)
        rename_excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        rename_excel_btn = ttk.Button(rename_excel_frame, text="选择Excel文件", command=self.select_rename_excel_file)
        rename_excel_btn.pack(side=tk.RIGHT, padx=5)
        
        # 匹配字段选择按钮
        rename_name_header_frame = ttk.Frame(self.rename_frame)
        rename_name_header_frame.pack(fill=tk.X, pady=5)
        
        self.rename_name_header_var = tk.StringVar(value="未选择")
        rename_name_header_entry = ttk.Entry(rename_name_header_frame, textvariable=self.rename_name_header_var, state="readonly", width=50)
        rename_name_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        rename_name_header_btn = ttk.Button(rename_name_header_frame, text="选择匹配字段", command=self.select_rename_name_header)
        rename_name_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 值字段选择按钮
        rename_id_header_frame = ttk.Frame(self.rename_frame)
        rename_id_header_frame.pack(fill=tk.X, pady=5)
        
        self.rename_id_header_var = tk.StringVar(value="未选择")
        rename_id_header_entry = ttk.Entry(rename_id_header_frame, textvariable=self.rename_id_header_var, state="readonly", width=50)
        rename_id_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        rename_id_header_btn = ttk.Button(rename_id_header_frame, text="选择值字段", command=self.select_rename_id_header)
        rename_id_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 匹配字段和值字段管理
        match_fields_frame = ttk.LabelFrame(self.rename_frame, text="匹配规则", padding="5")
        match_fields_frame.pack(fill=tk.X, pady=5)
        
        match_fields_content = ttk.Frame(match_fields_frame)
        match_fields_content.pack(fill=tk.X, padx=5, pady=5)
        
        # 添加匹配字段按钮
        add_match_btn = ttk.Button(match_fields_content, text="添加匹配字段", command=self.add_match_field)
        add_match_btn.pack(side=tk.LEFT, padx=5)
        
        # 匹配字段列表
        self.match_fields = []
        self.match_fields_frame = ttk.Frame(match_fields_frame)
        self.match_fields_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 文件名构建
        filename_frame = ttk.LabelFrame(self.rename_frame, text="文件名构建", padding="10")
        filename_frame.pack(fill=tk.X, pady=5)
        
        # 构建规则显示
        preview_frame = ttk.Frame(filename_frame)
        preview_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(preview_frame, text="文件名预览:", width=10).pack(side=tk.LEFT, padx=5)
        self.filename_preview_var = tk.StringVar(value="点击添加元素来构建文件名")
        filename_preview = ttk.Label(preview_frame, textvariable=self.filename_preview_var, font=('Arial', 10), foreground='#666666', relief=tk.SUNKEN, padding=5)
        filename_preview.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        # 操作按钮
        button_frame = ttk.Frame(filename_frame)
        button_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(button_frame, text="添加文本", command=lambda: self.add_filename_part("text")).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="添加Excel列", command=lambda: self.add_filename_part("excel")).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="获取班级名称", command=lambda: self.add_filename_part("class")).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="删除最后元素", command=self.remove_filename_part).pack(side=tk.LEFT, padx=5)
        
        # 构建规则列表
        ttk.Label(filename_frame, text="构建元素（点击上下箭头调整顺序）:").pack(anchor=tk.W, pady=5)
        
        # 创建带滚动条的框架
        parts_scroll_frame = ttk.Frame(filename_frame)
        parts_scroll_frame.pack(fill=tk.X, pady=5)
        
        # 创建滚动条
        parts_scrollbar = ttk.Scrollbar(parts_scroll_frame, orient=tk.VERTICAL)
        parts_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建Canvas
        parts_canvas = tk.Canvas(parts_scroll_frame, yscrollcommand=parts_scrollbar.set, height=200)
        parts_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 配置滚动条
        parts_scrollbar.config(command=parts_canvas.yview)
        
        # 创建内部框架
        self.filename_parts_frame = ttk.Frame(parts_canvas, relief=tk.GROOVE, borderwidth=1, padding=5)
        parts_canvas.create_window((0, 0), window=self.filename_parts_frame, anchor=tk.NW)
        
        # 绑定配置事件，更新滚动区域
        def on_frame_configure(event):
            parts_canvas.configure(scrollregion=parts_canvas.bbox("all"))
        
        self.filename_parts_frame.bind("<Configure>", on_frame_configure)
        
        # 中转表按钮框架
        rename_preview_frame = ttk.Frame(self.rename_frame)
        rename_preview_frame.pack(fill=tk.X, pady=5)
        
        # 打开中转表按钮
        rename_open_preview_btn = ttk.Button(rename_preview_frame, text="打开中转表预览", command=self.open_rename_preview)
        rename_open_preview_btn.pack(side=tk.LEFT, padx=5)
        
        # 重命名模式选择
        rename_mode_frame = ttk.LabelFrame(self.rename_frame, text="重命名模式", padding="10")
        rename_mode_frame.pack(fill=tk.X, pady=5)
        
        self.rename_mode_var = tk.StringVar(value="excel")
        excel_mode = ttk.Radiobutton(rename_mode_frame, text="Excel匹配重命名", variable=self.rename_mode_var, value="excel")
        excel_mode.pack(side=tk.LEFT, padx=10)
        
        class_mode = ttk.Radiobutton(rename_mode_frame, text="按班级重命名", variable=self.rename_mode_var, value="class")
        class_mode.pack(side=tk.LEFT, padx=10)
        
        # 预览重命名结果按钮
        rename_preview_btn = ttk.Button(rename_preview_frame, text="预览重命名结果", command=self.preview_rename)
        rename_preview_btn.pack(side=tk.LEFT, padx=5)
        
        # 按班级重命名设置
        class_rename_frame = ttk.LabelFrame(self.rename_frame, text="按班级重命名设置", padding="10")
        class_rename_frame.pack(fill=tk.X, pady=5)
        
        # 特殊符号设置
        special_char_frame = ttk.Frame(class_rename_frame)
        special_char_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(special_char_frame, text="特殊符号：").pack(side=tk.LEFT, padx=5)
        self.class_rename_special_char_var = tk.StringVar(value="_")
        class_rename_special_char_entry = ttk.Entry(special_char_frame, textvariable=self.class_rename_special_char_var, width=10)
        class_rename_special_char_entry.pack(side=tk.LEFT, padx=5)
        
        # 预览按班级重命名结果按钮
        class_rename_preview_btn = ttk.Button(class_rename_frame, text="预览按班级重命名结果", command=self.preview_class_rename)
        class_rename_preview_btn.pack(pady=5)
        
        # 指定导出页设置框架
        self.export_frame = ttk.LabelFrame(main_frame, text="指定导出页设置", padding="10")
        self.export_frame.pack(fill=tk.X, pady=5)
        
        # PDF文件选择
        export_pdf_frame = ttk.LabelFrame(self.export_frame, text="PDF文件（支持批量选择）", padding="5")
        export_pdf_frame.pack(fill=tk.X, pady=5)
        
        export_pdf_content = ttk.Frame(export_pdf_frame)
        export_pdf_content.pack(fill=tk.X, padx=5, pady=5)
        
        self.export_pdf_var = tk.StringVar(value="未选择文件")
        export_pdf_entry = ttk.Entry(export_pdf_content, textvariable=self.export_pdf_var, state="readonly", width=60)
        export_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        export_pdf_btn = ttk.Button(export_pdf_content, text="选择文件（可多选）", command=self.select_export_pdf)
        export_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # PDF文件数量显示
        self.export_pdf_count_var = tk.StringVar(value="")
        export_pdf_count = ttk.Label(export_pdf_frame, textvariable=self.export_pdf_count_var, font=('Arial', 9), foreground='#666666')
        export_pdf_count.pack(side=tk.LEFT, padx=10, pady=2, anchor=tk.W)
        
        # 导出页面设置
        export_pages_frame = ttk.Frame(self.export_frame)
        export_pages_frame.pack(fill=tk.X, pady=5)
        
        export_pages_label = ttk.Label(export_pages_frame, text="导出页面（如：1-6,1,2,3,4）:")
        export_pages_label.pack(side=tk.LEFT, padx=5)
        
        self.export_pages_var = tk.StringVar(value="1-6,1,2,3,4")
        export_pages_entry = ttk.Entry(export_pages_frame, textvariable=self.export_pages_var, width=30)
        export_pages_entry.pack(side=tk.LEFT, padx=5)
        
        # 子PDF分割功能
        sub_pdf_frame = ttk.LabelFrame(self.export_frame, text="子PDF分割设置", padding="5")
        sub_pdf_frame.pack(fill=tk.X, pady=5)
        
        # 子PDF数量控制
        sub_pdf_count_frame = ttk.Frame(sub_pdf_frame)
        sub_pdf_count_frame.pack(fill=tk.X, pady=5)
        
        sub_pdf_count_label = ttk.Label(sub_pdf_count_frame, text="子PDF数量:")
        sub_pdf_count_label.pack(side=tk.LEFT, padx=5)
        
        self.sub_pdf_count_var = tk.StringVar(value="2")
        sub_pdf_count_entry = ttk.Entry(sub_pdf_count_frame, textvariable=self.sub_pdf_count_var, width=5, state="readonly")
        sub_pdf_count_entry.pack(side=tk.LEFT, padx=5)
        
        sub_pdf_add_btn = ttk.Button(sub_pdf_count_frame, text="+", command=self.add_sub_pdf)
        sub_pdf_add_btn.pack(side=tk.LEFT, padx=5)
        
        sub_pdf_remove_btn = ttk.Button(sub_pdf_count_frame, text="-", command=self.remove_sub_pdf)
        sub_pdf_remove_btn.pack(side=tk.LEFT, padx=5)
        
        # 子PDF配置区域
        self.sub_pdf_frames = []
        self.sub_pdf_range_vars = []
        self.sub_pdf_name_vars = []
        self.sub_pdf_path_vars = []
        
        # 初始创建2个子PDF配置
        for i in range(2):
            self.create_sub_pdf_frame(sub_pdf_frame, i)
        
        # 输出选项
        export_options_frame = ttk.Frame(self.export_frame)
        export_options_frame.pack(fill=tk.X, pady=5)
        
        self.export_use_original_name_var = tk.BooleanVar(value=False)
        export_use_original_name_check = ttk.Checkbutton(export_options_frame, text="以原PDF文件名输出（勾选后每个PDF按原文件名加页码范围命名）", variable=self.export_use_original_name_var)
        export_use_original_name_check.pack(side=tk.LEFT, padx=5)
        
        # Excel文件选择
        export_excel_frame = ttk.Frame(self.export_frame)
        export_excel_frame.pack(fill=tk.X, pady=5)
        
        self.export_excel_var = tk.StringVar()
        export_excel_entry = ttk.Entry(export_excel_frame, textvariable=self.export_excel_var, state="readonly", width=50)
        export_excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        export_excel_btn = ttk.Button(export_excel_frame, text="选择Excel文件", command=self.select_export_excel_file)
        export_excel_btn.pack(side=tk.RIGHT, padx=5)
        
        # 表头选择（姓名和序号放在同一行）
        export_headers_frame = ttk.Frame(self.export_frame)
        export_headers_frame.pack(fill=tk.X, pady=5)
        
        # 姓名表头选择
        export_name_header_frame = ttk.Frame(export_headers_frame)
        export_name_header_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.export_name_header_var = tk.StringVar(value="未选择")
        export_name_header_entry = ttk.Entry(export_name_header_frame, textvariable=self.export_name_header_var, state="readonly")
        export_name_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        export_name_header_btn = ttk.Button(export_name_header_frame, text="选择姓名表头", command=self.select_export_name_header)
        export_name_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 序号表头选择
        export_id_header_frame = ttk.Frame(export_headers_frame)
        export_id_header_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.export_id_header_var = tk.StringVar(value="未选择")
        export_id_header_entry = ttk.Entry(export_id_header_frame, textvariable=self.export_id_header_var, state="readonly")
        export_id_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        export_id_header_btn = ttk.Button(export_id_header_frame, text="选择序号表头", command=self.select_export_id_header)
        export_id_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 中转表按钮框架
        export_preview_frame = ttk.Frame(self.export_frame)
        export_preview_frame.pack(fill=tk.X, pady=5)
        
        # 打开中转表按钮
        export_open_preview_btn = ttk.Button(export_preview_frame, text="打开中转表预览", command=self.open_export_preview)
        export_open_preview_btn.pack(side=tk.LEFT, padx=5)
        
        # 预览导出结果按钮
        export_preview_btn = ttk.Button(export_preview_frame, text="预览导出结果", command=self.preview_export)
        export_preview_btn.pack(side=tk.LEFT, padx=5)
        
        # PDF重排序设置框架
        self.reorder_frame = ttk.LabelFrame(main_frame, text="PDF重排序设置", padding="10")
        self.reorder_frame.pack(fill=tk.X, pady=5)
        
        # 重排序PDF选择
        reorder_pdf_frame = ttk.Frame(self.reorder_frame)
        reorder_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.reorder_pdf_var = tk.StringVar()
        reorder_pdf_entry = ttk.Entry(reorder_pdf_frame, textvariable=self.reorder_pdf_var, state="readonly", width=50)
        reorder_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        reorder_pdf_btn = ttk.Button(reorder_pdf_frame, text="选择文件", command=self.select_reorder_pdfs)
        reorder_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # Excel文件选择
        reorder_excel_frame = ttk.Frame(self.reorder_frame)
        reorder_excel_frame.pack(fill=tk.X, pady=5)
        
        self.reorder_excel_var = tk.StringVar()
        reorder_excel_entry = ttk.Entry(reorder_excel_frame, textvariable=self.reorder_excel_var, state="readonly", width=50)
        reorder_excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        reorder_excel_btn = ttk.Button(reorder_excel_frame, text="选择Excel文件", command=self.select_reorder_excel_file)
        reorder_excel_btn.pack(side=tk.RIGHT, padx=5)
        
        # 姓名表头选择按钮
        reorder_name_header_frame = ttk.Frame(self.reorder_frame)
        reorder_name_header_frame.pack(fill=tk.X, pady=5)
        
        self.reorder_name_header_var = tk.StringVar(value="未选择")
        reorder_name_header_entry = ttk.Entry(reorder_name_header_frame, textvariable=self.reorder_name_header_var, state="readonly", width=50)
        reorder_name_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        reorder_name_header_btn = ttk.Button(reorder_name_header_frame, text="选择姓名表头", command=self.select_reorder_name_header)
        reorder_name_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 序号表头选择按钮
        reorder_id_header_frame = ttk.Frame(self.reorder_frame)
        reorder_id_header_frame.pack(fill=tk.X, pady=5)
        
        self.reorder_id_header_var = tk.StringVar(value="未选择")
        reorder_id_header_entry = ttk.Entry(reorder_id_header_frame, textvariable=self.reorder_id_header_var, state="readonly", width=50)
        reorder_id_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        reorder_id_header_btn = ttk.Button(reorder_id_header_frame, text="选择序号表头", command=self.select_reorder_id_header)
        reorder_id_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 中转表按钮框架
        reorder_preview_frame = ttk.Frame(self.reorder_frame)
        reorder_preview_frame.pack(fill=tk.X, pady=5)
        
        # 打开中转表按钮
        reorder_open_preview_btn = ttk.Button(reorder_preview_frame, text="打开中转表预览", command=self.open_reorder_preview)
        reorder_open_preview_btn.pack(side=tk.LEFT, padx=5)
        
        # 预览重排序结果按钮
        reorder_preview_btn = ttk.Button(reorder_preview_frame, text="预览重排序结果", command=self.preview_reorder)
        reorder_preview_btn.pack(side=tk.LEFT, padx=5)
        
        # PDF页面设置框架
        self.page_frame = ttk.LabelFrame(main_frame, text="PDF页面设置", padding="10")
        self.page_frame.pack(fill=tk.X, pady=5)
        
        # PDF文件选择
        page_pdf_frame = ttk.Frame(self.page_frame)
        page_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.page_pdf_var = tk.StringVar()
        page_pdf_entry = ttk.Entry(page_pdf_frame, textvariable=self.page_pdf_var, state="readonly", width=50)
        page_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        page_pdf_btn = ttk.Button(page_pdf_frame, text="选择文件（最多100个）", command=self.select_page_pdfs)
        page_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # 文件数量显示
        self.page_pdf_count_var = tk.StringVar(value="已选择 0 个文件")
        page_pdf_count_label = ttk.Label(self.page_frame, textvariable=self.page_pdf_count_var, font=('Arial', 9), foreground='#666666')
        page_pdf_count_label.pack(side=tk.LEFT, padx=10, pady=2, anchor=tk.W)
        
        # 功能按钮框架
        page_function_frame = ttk.Frame(self.page_frame)
        page_function_frame.pack(fill=tk.X, pady=10)
        
        # 颠倒页面顺序按钮
        reverse_btn = ttk.Button(page_function_frame, text="一键颠倒页面顺序", command=self.reverse_pages)
        reverse_btn.pack(side=tk.LEFT, padx=5)
        
        # 旋转页面按钮
        rotate_btn = ttk.Button(page_function_frame, text="设置页面旋转方案", command=self.open_rotation_preview)
        rotate_btn.pack(side=tk.LEFT, padx=5)
        
        # 批量提取设置框架
        self.batch_extract_frame = ttk.LabelFrame(main_frame, text="批量提取设置", padding="10")
        self.batch_extract_frame.pack(fill=tk.X, pady=5)
        
        # 批量PDF选择
        batch_extract_pdf_frame = ttk.Frame(self.batch_extract_frame)
        batch_extract_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.batch_extract_pdf_var = tk.StringVar()
        batch_extract_pdf_entry = ttk.Entry(batch_extract_pdf_frame, textvariable=self.batch_extract_pdf_var, state="readonly", width=50)
        batch_extract_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        batch_extract_pdf_btn = ttk.Button(batch_extract_pdf_frame, text="选择文件（批量）", command=self.select_batch_extract_pdfs)
        batch_extract_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # 提取页面设置
        batch_extract_page_frame = ttk.Frame(self.batch_extract_frame)
        batch_extract_page_frame.pack(fill=tk.X, pady=5)
        
        batch_extract_page_label = ttk.Label(batch_extract_page_frame, text="提取页码:")
        batch_extract_page_label.pack(side=tk.LEFT, padx=5)
        
        self.batch_extract_page_var = tk.StringVar(value="1")
        batch_extract_page_entry = ttk.Entry(batch_extract_page_frame, textvariable=self.batch_extract_page_var, width=5)
        batch_extract_page_entry.pack(side=tk.LEFT, padx=5)
        
        # 输出目录选择
        batch_extract_output_frame = ttk.Frame(self.batch_extract_frame)
        batch_extract_output_frame.pack(fill=tk.X, pady=5)
        
        self.batch_extract_output_var = tk.StringVar()
        batch_extract_output_entry = ttk.Entry(batch_extract_output_frame, textvariable=self.batch_extract_output_var, state="readonly", width=50)
        batch_extract_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        batch_extract_output_btn = ttk.Button(batch_extract_output_frame, text="选择输出目录", command=self.select_batch_extract_output_dir)
        batch_extract_output_btn.pack(side=tk.RIGHT, padx=5)
        
        # Excel匹配设置
        self.batch_extract_excel_frame = ttk.LabelFrame(self.batch_extract_frame, text="Excel匹配设置", padding="10")
        self.batch_extract_excel_frame.pack(fill=tk.X, pady=5)
        
        # Excel文件选择
        batch_extract_excel_frame = ttk.Frame(self.batch_extract_excel_frame)
        batch_extract_excel_frame.pack(fill=tk.X, pady=5)
        
        self.batch_extract_excel_var = tk.StringVar()
        batch_extract_excel_entry = ttk.Entry(batch_extract_excel_frame, textvariable=self.batch_extract_excel_var, state="readonly", width=50)
        batch_extract_excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        batch_extract_excel_btn = ttk.Button(batch_extract_excel_frame, text="选择Excel文件（最多5个）", command=self.select_batch_extract_excel_files)
        batch_extract_excel_btn.pack(side=tk.RIGHT, padx=5)
        
        batch_extract_excel_preview_btn = ttk.Button(batch_extract_excel_frame, text="Excel预览", command=self.preview_excel_content_batch_extract)
        batch_extract_excel_preview_btn.pack(side=tk.RIGHT, padx=5)
        
        # 表头选择（姓名和班级）
        batch_extract_headers_frame = ttk.Frame(self.batch_extract_excel_frame)
        batch_extract_headers_frame.pack(fill=tk.X, pady=5)
        
        # 姓名表头选择
        batch_extract_name_header_frame = ttk.Frame(batch_extract_headers_frame)
        batch_extract_name_header_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.batch_extract_name_header_var = tk.StringVar(value="未选择")
        batch_extract_name_header_entry = ttk.Entry(batch_extract_name_header_frame, textvariable=self.batch_extract_name_header_var, state="readonly")
        batch_extract_name_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        batch_extract_name_header_btn = ttk.Button(batch_extract_name_header_frame, text="选择姓名表头", command=self.select_batch_extract_name_header)
        batch_extract_name_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 班级表头选择
        batch_extract_class_header_frame = ttk.Frame(batch_extract_headers_frame)
        batch_extract_class_header_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.batch_extract_class_header_var = tk.StringVar(value="未选择")
        batch_extract_class_header_entry = ttk.Entry(batch_extract_class_header_frame, textvariable=self.batch_extract_class_header_var, state="readonly")
        batch_extract_class_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        batch_extract_class_header_btn = ttk.Button(batch_extract_class_header_frame, text="选择班级表头", command=self.select_batch_extract_class_header)
        batch_extract_class_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 姓名提取设置
        batch_extract_name_extract_frame = ttk.LabelFrame(self.batch_extract_frame, text="姓名提取设置", padding="10")
        batch_extract_name_extract_frame.pack(fill=tk.X, pady=5)
        
        # 自动提取按钮
        batch_extract_auto_extract_frame = ttk.Frame(batch_extract_name_extract_frame)
        batch_extract_auto_extract_frame.pack(fill=tk.X, pady=5)
        
        batch_extract_auto_extract_btn = ttk.Button(batch_extract_auto_extract_frame, text="自动提取姓名", command=self.auto_extract_names_batch_extract)
        batch_extract_auto_extract_btn.pack(side=tk.LEFT, padx=5)
        
        # 自定义特殊符号提取
        batch_extract_custom_extract_frame = ttk.Frame(batch_extract_name_extract_frame)
        batch_extract_custom_extract_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(batch_extract_custom_extract_frame, text="特殊符号：").pack(side=tk.LEFT, padx=5)
        self.batch_extract_special_char_var = tk.StringVar(value="_")
        batch_extract_special_char_entry = ttk.Entry(batch_extract_custom_extract_frame, textvariable=self.batch_extract_special_char_var, width=10)
        batch_extract_special_char_entry.pack(side=tk.LEFT, padx=5)
        
        batch_extract_custom_extract_btn = ttk.Button(batch_extract_custom_extract_frame, text="提取姓名", command=self.custom_extract_names_batch_extract)
        batch_extract_custom_extract_btn.pack(side=tk.LEFT, padx=5)
        
        # 预览按钮
        batch_extract_preview_frame = ttk.Frame(self.batch_extract_frame)
        batch_extract_preview_frame.pack(fill=tk.X, pady=5)
        
        batch_extract_preview_btn = ttk.Button(batch_extract_preview_frame, text="预览", command=self.preview_batch_extract)
        batch_extract_preview_btn.pack(side=tk.LEFT, padx=5)
        
        # 姓名识别预览按钮
        batch_extract_name_preview_btn = ttk.Button(batch_extract_preview_frame, text="姓名识别预览", command=self.preview_name_extraction_batch_extract)
        batch_extract_name_preview_btn.pack(side=tk.LEFT, padx=5)
        
        # 批量追加页设置框架
        self.batch_append_frame = ttk.LabelFrame(main_frame, text="批量追加页设置", padding="10")
        self.batch_append_frame.pack(fill=tk.X, pady=5)
        
        # 源PDF文件选择
        batch_append_source_frame = ttk.Frame(self.batch_append_frame)
        batch_append_source_frame.pack(fill=tk.X, pady=5)
        
        self.batch_append_source_var = tk.StringVar()
        batch_append_source_entry = ttk.Entry(batch_append_source_frame, textvariable=self.batch_append_source_var, state="readonly", width=50)
        batch_append_source_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        batch_append_source_btn = ttk.Button(batch_append_source_frame, text="选择源PDF文件", command=self.select_batch_append_source_pdf)
        batch_append_source_btn.pack(side=tk.RIGHT, padx=5)
        
        # 批量PDF选择
        batch_append_pdf_frame = ttk.Frame(self.batch_append_frame)
        batch_append_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.batch_append_pdf_var = tk.StringVar()
        batch_append_pdf_entry = ttk.Entry(batch_append_pdf_frame, textvariable=self.batch_append_pdf_var, state="readonly", width=50)
        batch_append_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        batch_append_pdf_btn = ttk.Button(batch_append_pdf_frame, text="选择文件（批量）", command=self.select_batch_append_pdfs)
        batch_append_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # 输出目录选择
        batch_append_output_frame = ttk.Frame(self.batch_append_frame)
        batch_append_output_frame.pack(fill=tk.X, pady=5)
        
        self.batch_append_output_var = tk.StringVar()
        batch_append_output_entry = ttk.Entry(batch_append_output_frame, textvariable=self.batch_append_output_var, state="readonly", width=50)
        batch_append_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        batch_append_output_btn = ttk.Button(batch_append_output_frame, text="选择输出目录", command=self.select_batch_append_output_dir)
        batch_append_output_btn.pack(side=tk.RIGHT, padx=5)
        
        # Excel匹配设置
        self.batch_append_excel_frame = ttk.LabelFrame(self.batch_append_frame, text="Excel匹配设置", padding="10")
        self.batch_append_excel_frame.pack(fill=tk.X, pady=5)
        
        # Excel文件选择
        batch_append_excel_frame = ttk.Frame(self.batch_append_excel_frame)
        batch_append_excel_frame.pack(fill=tk.X, pady=5)
        
        self.batch_append_excel_var = tk.StringVar()
        batch_append_excel_entry = ttk.Entry(batch_append_excel_frame, textvariable=self.batch_append_excel_var, state="readonly", width=50)
        batch_append_excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        batch_append_excel_btn = ttk.Button(batch_append_excel_frame, text="选择Excel文件（最多5个）", command=self.select_batch_append_excel_files)
        batch_append_excel_btn.pack(side=tk.RIGHT, padx=5)
        
        batch_append_excel_preview_btn = ttk.Button(batch_append_excel_frame, text="Excel预览", command=self.preview_excel_content_batch_append)
        batch_append_excel_preview_btn.pack(side=tk.RIGHT, padx=5)
        
        # 表头选择（姓名和班级）
        batch_append_headers_frame = ttk.Frame(self.batch_append_excel_frame)
        batch_append_headers_frame.pack(fill=tk.X, pady=5)
        
        # 姓名表头选择
        batch_append_name_header_frame = ttk.Frame(batch_append_headers_frame)
        batch_append_name_header_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.batch_append_name_header_var = tk.StringVar(value="未选择")
        batch_append_name_header_entry = ttk.Entry(batch_append_name_header_frame, textvariable=self.batch_append_name_header_var, state="readonly")
        batch_append_name_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        batch_append_name_header_btn = ttk.Button(batch_append_name_header_frame, text="选择姓名表头", command=self.select_batch_append_name_header)
        batch_append_name_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 班级表头选择
        batch_append_class_header_frame = ttk.Frame(batch_append_headers_frame)
        batch_append_class_header_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.batch_append_class_header_var = tk.StringVar(value="未选择")
        batch_append_class_header_entry = ttk.Entry(batch_append_class_header_frame, textvariable=self.batch_append_class_header_var, state="readonly")
        batch_append_class_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        batch_append_class_header_btn = ttk.Button(batch_append_class_header_frame, text="选择班级表头", command=self.select_batch_append_class_header)
        batch_append_class_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 姓名提取设置
        batch_append_name_extract_frame = ttk.LabelFrame(self.batch_append_frame, text="姓名提取设置", padding="10")
        batch_append_name_extract_frame.pack(fill=tk.X, pady=5)
        
        # 自动提取按钮
        batch_append_auto_extract_frame = ttk.Frame(batch_append_name_extract_frame)
        batch_append_auto_extract_frame.pack(fill=tk.X, pady=5)
        
        batch_append_auto_extract_btn = ttk.Button(batch_append_auto_extract_frame, text="自动提取姓名", command=self.auto_extract_names_batch_append)
        batch_append_auto_extract_btn.pack(side=tk.LEFT, padx=5)
        
        # 自定义特殊符号提取
        batch_append_custom_extract_frame = ttk.Frame(batch_append_name_extract_frame)
        batch_append_custom_extract_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(batch_append_custom_extract_frame, text="特殊符号：").pack(side=tk.LEFT, padx=5)
        self.batch_append_special_char_var = tk.StringVar(value="_")
        batch_append_special_char_entry = ttk.Entry(batch_append_custom_extract_frame, textvariable=self.batch_append_special_char_var, width=10)
        batch_append_special_char_entry.pack(side=tk.LEFT, padx=5)
        
        batch_append_custom_extract_btn = ttk.Button(batch_append_custom_extract_frame, text="提取姓名", command=self.custom_extract_names_batch_append)
        batch_append_custom_extract_btn.pack(side=tk.LEFT, padx=5)
        
        # 预览按钮
        batch_append_preview_frame = ttk.Frame(self.batch_append_frame)
        batch_append_preview_frame.pack(fill=tk.X, pady=5)
        
        batch_append_preview_btn = ttk.Button(batch_append_preview_frame, text="预览", command=self.preview_batch_append)
        batch_append_preview_btn.pack(side=tk.LEFT, padx=5)
        
        # 姓名识别预览按钮
        batch_append_name_preview_btn = ttk.Button(batch_append_preview_frame, text="姓名识别预览", command=self.preview_name_extraction_batch_append)
        batch_append_name_preview_btn.pack(side=tk.LEFT, padx=5)
        
        # 学生班级PDF合并设置框架
        self.student_class_frame = ttk.LabelFrame(main_frame, text="学生班级PDF合并设置", padding="10")
        self.student_class_frame.pack(fill=tk.X, pady=5)
        
        # 学生PDF文件选择
        student_pdf_frame = ttk.Frame(self.student_class_frame)
        student_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.student_pdf_var = tk.StringVar()
        student_pdf_entry = ttk.Entry(student_pdf_frame, textvariable=self.student_pdf_var, state="readonly", width=50)
        student_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        student_pdf_btn = ttk.Button(student_pdf_frame, text="选择学生PDF文件（批量）", command=self.select_student_pdfs)
        student_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # 班级PDF文件选择
        class_pdf_frame = ttk.Frame(self.student_class_frame)
        class_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.class_pdf_var = tk.StringVar()
        class_pdf_entry = ttk.Entry(class_pdf_frame, textvariable=self.class_pdf_var, state="readonly", width=50)
        class_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        class_pdf_btn = ttk.Button(class_pdf_frame, text="选择班级PDF文件（批量）", command=self.select_class_pdfs)
        class_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # Excel文件选择
        student_class_excel_frame = ttk.Frame(self.student_class_frame)
        student_class_excel_frame.pack(fill=tk.X, pady=5)
        
        self.student_class_excel_var = tk.StringVar()
        student_class_excel_entry = ttk.Entry(student_class_excel_frame, textvariable=self.student_class_excel_var, state="readonly", width=50)
        student_class_excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        student_class_excel_btn = ttk.Button(student_class_excel_frame, text="选择Excel文件", command=self.select_student_class_excel)
        student_class_excel_btn.pack(side=tk.RIGHT, padx=5)
        
        # 表头选择（姓名和班级）
        student_class_headers_frame = ttk.Frame(self.student_class_frame)
        student_class_headers_frame.pack(fill=tk.X, pady=5)
        
        # 姓名表头选择
        student_class_name_header_frame = ttk.Frame(student_class_headers_frame)
        student_class_name_header_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.student_class_name_header_var = tk.StringVar(value="未选择")
        student_class_name_header_entry = ttk.Entry(student_class_name_header_frame, textvariable=self.student_class_name_header_var, state="readonly")
        student_class_name_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        student_class_name_header_btn = ttk.Button(student_class_name_header_frame, text="选择姓名表头", command=self.select_student_class_name_header)
        student_class_name_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 班级表头选择
        student_class_class_header_frame = ttk.Frame(student_class_headers_frame)
        student_class_class_header_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.student_class_class_header_var = tk.StringVar(value="未选择")
        student_class_class_header_entry = ttk.Entry(student_class_class_header_frame, textvariable=self.student_class_class_header_var, state="readonly")
        student_class_class_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        student_class_class_header_btn = ttk.Button(student_class_class_header_frame, text="选择班级表头", command=self.select_student_class_class_header)
        student_class_class_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 输出目录选择
        student_class_output_frame = ttk.Frame(self.student_class_frame)
        student_class_output_frame.pack(fill=tk.X, pady=5)
        
        self.student_class_output_var = tk.StringVar()
        student_class_output_entry = ttk.Entry(student_class_output_frame, textvariable=self.student_class_output_var, state="readonly", width=50)
        student_class_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        student_class_output_btn = ttk.Button(student_class_output_frame, text="选择输出目录", command=self.select_student_class_output_dir)
        student_class_output_btn.pack(side=tk.RIGHT, padx=5)
        
        # 预览按钮
        student_class_preview_frame = ttk.Frame(self.student_class_frame)
        student_class_preview_frame.pack(fill=tk.X, pady=5)
        
        student_class_preview_btn = ttk.Button(student_class_preview_frame, text="预览合并结果", command=self.preview_student_class_merge)
        student_class_preview_btn.pack(side=tk.LEFT, padx=5)
        
        # Excel批量导出PDF设置框架
        self.excel_to_pdf_frame = ttk.LabelFrame(main_frame, text="Excel批量导出PDF设置", padding="10")
        self.excel_to_pdf_frame.pack(fill=tk.X, pady=5)
        
        # Excel文件选择
        excel_files_frame = ttk.Frame(self.excel_to_pdf_frame)
        excel_files_frame.pack(fill=tk.X, pady=5)
        
        excel_files_entry = ttk.Entry(excel_files_frame, textvariable=self.excel_files_var, state="readonly", width=50)
        excel_files_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        excel_files_btn = ttk.Button(excel_files_frame, text="选择Excel文件（批量）", command=self.select_excel_files)
        excel_files_btn.pack(side=tk.RIGHT, padx=5)
        
        # 输出目录选择
        excel_output_frame = ttk.Frame(self.excel_to_pdf_frame)
        excel_output_frame.pack(fill=tk.X, pady=5)
        
        excel_output_entry = ttk.Entry(excel_output_frame, textvariable=self.excel_output_var, state="readonly", width=50)
        excel_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        excel_output_btn = ttk.Button(excel_output_frame, text="选择输出目录", command=self.select_excel_output_dir)
        excel_output_btn.pack(side=tk.RIGHT, padx=5)
        
        # Excel Sheet选择
        excel_sheet_frame = ttk.Frame(self.excel_to_pdf_frame)
        excel_sheet_frame.pack(fill=tk.X, pady=5)
        
        self.excel_export_all_sheets_var = tk.BooleanVar(value=True)
        excel_all_sheets_check = ttk.Checkbutton(excel_sheet_frame, text="导出所有Sheet", variable=self.excel_export_all_sheets_var)
        excel_all_sheets_check.pack(side=tk.LEFT, padx=5)
        
        excel_select_sheets_btn = ttk.Button(excel_sheet_frame, text="选择Sheet", command=self.select_excel_sheets)
        excel_select_sheets_btn.pack(side=tk.RIGHT, padx=5)
        
        # Word批量导出PDF设置框架
        self.word_to_pdf_frame = ttk.LabelFrame(main_frame, text="Word批量导出PDF设置", padding="10")
        self.word_to_pdf_frame.pack(fill=tk.X, pady=5)
        
        # Word文件选择
        word_files_frame = ttk.Frame(self.word_to_pdf_frame)
        word_files_frame.pack(fill=tk.X, pady=5)
        
        word_files_entry = ttk.Entry(word_files_frame, textvariable=self.word_files_var, state="readonly", width=50)
        word_files_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        word_files_btn = ttk.Button(word_files_frame, text="选择Word文件（批量）", command=self.select_word_files)
        word_files_btn.pack(side=tk.RIGHT, padx=5)
        
        # 输出目录选择
        word_output_frame = ttk.Frame(self.word_to_pdf_frame)
        word_output_frame.pack(fill=tk.X, pady=5)
        
        word_output_entry = ttk.Entry(word_output_frame, textvariable=self.word_output_var, state="readonly", width=50)
        word_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        word_output_btn = ttk.Button(word_output_frame, text="选择输出目录", command=self.select_word_output_dir)
        word_output_btn.pack(side=tk.RIGHT, padx=5)
        
        # Word页面选择
        word_pages_frame = ttk.Frame(self.word_to_pdf_frame)
        word_pages_frame.pack(fill=tk.X, pady=5)
        
        self.word_export_all_pages_var = tk.BooleanVar(value=True)
        word_all_pages_check = ttk.Checkbutton(word_pages_frame, text="导出全部页面", variable=self.word_export_all_pages_var)
        word_all_pages_check.pack(side=tk.LEFT, padx=5)
        
        word_pages_label = ttk.Label(word_pages_frame, text="页码范围（如：2-5,7）：")
        word_pages_label.pack(side=tk.LEFT, padx=5)
        
        word_pages_entry = ttk.Entry(word_pages_frame, textvariable=self.word_export_pages_var, width=20)
        word_pages_entry.pack(side=tk.LEFT, padx=5)
        
        # 操作按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        self.process_btn = ttk.Button(button_frame, text="执行操作", command=self.process_pdf, state=tk.DISABLED)
        self.process_btn.pack(side=tk.LEFT, padx=5)
        
        self.clear_btn = ttk.Button(button_frame, text="清空", command=self.clear_all)
        self.clear_btn.pack(side=tk.LEFT, padx=5)
        
        # 设置保存和加载功能
        settings_frame = ttk.Frame(main_frame)
        settings_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(settings_frame, text="保存设置", command=self.save_settings).pack(side=tk.RIGHT, padx=5)
        ttk.Button(settings_frame, text="加载设置", command=self.load_settings).pack(side=tk.RIGHT, padx=5)
        
        # 状态信息
        self.status_var = tk.StringVar(value="就绪")
        status_label = ttk.Label(main_frame, textvariable=self.status_var, font= ("微软雅黑", 10))
        status_label.pack(pady=10)
        
        # 初始化显示状态
        self.on_function_change()
        
        # 配置画布滚动区域
        main_frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))
    
    def on_function_change(self):
        # 根据选择的功能显示相应的设置区域
        # 首先隐藏所有框架
        all_frames = [
            self.insert_frame,
            self.split_frame,
            self.swap_frame,
            self.export_frame,
            self.rename_frame,
            self.reorder_frame,
            self.page_frame,
            self.batch_extract_frame,
            self.batch_append_frame,
            self.student_class_frame,
            self.excel_to_pdf_frame,
            self.word_to_pdf_frame
        ]
        
        for frame in all_frames:
            frame.pack_forget()
        
        # 然后显示选中的功能对应的框架
        selected_func = self.function_var.get()
        if selected_func == "insert":
            self.insert_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "split":
            self.split_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "swap":
            self.swap_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "export":
            self.export_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "reorder":
            self.reorder_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "page":
            self.page_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "batch_extract":
            self.batch_extract_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "batch_append":
            self.batch_append_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "student_class":
            self.student_class_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "excel_to_pdf":
            self.excel_to_pdf_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "word_to_pdf":
            self.word_to_pdf_frame.pack(fill=tk.X, pady=5)
        else:  # rename (默认)
            self.rename_frame.pack(fill=tk.X, pady=5)
        self.update_button_state()
    
    def select_main_pdf(self):
        file_path = filedialog.askopenfilename(
            title="选择主PDF文件",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_path:
            if self.is_valid_pdf(file_path):
                self.main_pdf = file_path
                self.main_pdf_var.set(file_path)
                
                # 更新主PDF信息
                try:
                    # 获取文件大小
                    file_size = os.path.getsize(file_path) / 1024  # KB
                    
                    # 获取页数
                    with open(file_path, 'rb') as f:
                        reader = PyPDF2.PdfReader(f)
                        page_count = len(reader.pages)
                    
                    # 更新信息显示
                    self.main_pdf_info_var.set(f"文件大小: {file_size:.2f} KB, 页数: {page_count}页")
                except Exception as e:
                    self.main_pdf_info_var.set("无法读取文件信息")
                
                self.update_button_state()
            else:
                messagebox.showerror("错误", "选择的文件不是有效的PDF文件")
    
    def select_single_insert_pdf(self, index):
        file_path = filedialog.askopenfilename(
            title=f"选择插入PDF文件 {index+1}",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_path:
            if self.is_valid_pdf(file_path):
                # 更新对应位置的PDF文件
                self.insert_pdfs[index] = file_path
                self.insert_pdf_vars[index].set(file_path)
                
                # 更新该PDF的信息
                try:
                    # 获取文件大小
                    file_size = os.path.getsize(file_path) / 1024  # KB
                    
                    # 获取页数
                    with open(file_path, 'rb') as f:
                        reader = PyPDF2.PdfReader(f)
                        page_count = len(reader.pages)
                    
                    # 更新信息显示
                    self.insert_pdf_info_vars[index].set(f"文件大小: {file_size:.2f} KB, 页数: {page_count}页")
                except Exception as e:
                    self.insert_pdf_info_vars[index].set("无法读取文件信息")
                
                self.update_button_state()
            else:
                messagebox.showerror("错误", "选择的文件不是有效的PDF文件")
    
    def select_split_pdf(self):
        file_path = filedialog.askopenfilename(
            title="选择要分割的PDF文件",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_path:
            if self.is_valid_pdf(file_path):
                self.split_pdf = file_path
                self.split_pdf_var.set(os.path.basename(file_path))
                self.update_button_state()
            else:
                messagebox.showerror("错误", "选择的文件不是有效的PDF文件")
    
    def select_swap_pdfs(self):
        file_paths = filedialog.askopenfilenames(
            title="选择要交换页面的PDF文件",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_paths:
            valid_pdfs = []
            for path in file_paths:
                if self.is_valid_pdf(path):
                    valid_pdfs.append(path)
                else:
                    messagebox.showerror("错误", f"文件 {os.path.basename(path)} 不是有效的PDF文件")
            
            if valid_pdfs:
                self.swap_pdfs = valid_pdfs
                self.swap_pdf_var.set(f"已选择 {len(valid_pdfs)} 个文件")
                self.update_button_state()
    
    def select_rename_pdfs(self):
        file_paths = filedialog.askopenfilenames(
            title="选择要重命名的PDF文件",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_paths:
            valid_pdfs = []
            for path in file_paths:
                if self.is_valid_pdf(path):
                    valid_pdfs.append(path)
                else:
                    messagebox.showerror("错误", f"文件 {os.path.basename(path)} 不是有效的PDF文件")
            
            if valid_pdfs:
                self.rename_pdfs = valid_pdfs
                self.rename_pdf_var.set(f"已选择 {len(valid_pdfs)} 个文件")
                self.update_button_state()
    
    def select_reorder_pdfs(self):
        file_paths = filedialog.askopenfilenames(
            title="选择要重排序的PDF文件",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_paths:
            valid_pdfs = []
            for path in file_paths:
                if self.is_valid_pdf(path):
                    valid_pdfs.append(path)
                else:
                    messagebox.showerror("错误", f"文件 {os.path.basename(path)} 不是有效的PDF文件")
            
            if valid_pdfs:
                self.reorder_pdfs = valid_pdfs
                self.reorder_pdf_var.set(f"已选择 {len(valid_pdfs)} 个文件")
                self.update_button_state()
    
    def select_rename_excel_file(self):
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("Excel文件", "*.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.rename_excel_file = file_path
            self.rename_excel_var.set(os.path.basename(file_path))
            self.update_button_state()
    
    def select_reorder_excel_file(self):
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("Excel文件", "*.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.reorder_excel_file = file_path
            self.reorder_excel_var.set(os.path.basename(file_path))
            self.update_button_state()
    
    def select_rename_name_header(self):
        # 选择姓名表头
        if not self.rename_excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        try:
            # 获取文件扩展名
            ext = os.path.splitext(self.rename_excel_file)[1].lower()
            headers = []
            
            if ext == '.xlsx':
                # 使用openpyxl读取xlsx文件
                wb = openpyxl.load_workbook(self.rename_excel_file)
                ws = wb.active
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 1  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(1, min(max_check_rows, ws.max_row) + 1):
                    non_empty_cells = 0
                    for cell in ws[row]:
                        if cell.value:
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for cell in ws[header_row]:
                    if cell.value:
                        headers.append(cell.value)
            elif ext == '.xls':
                # 使用xlrd读取xls文件
                wb = xlrd.open_workbook(self.rename_excel_file)
                ws = wb.sheet_by_index(0)
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 0  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(min(max_check_rows, ws.nrows)):
                    non_empty_cells = 0
                    for col in range(ws.ncols):
                        if ws.cell_value(row, col):
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for col in range(ws.ncols):
                    cell_value = ws.cell_value(header_row, col)
                    if cell_value:
                        headers.append(cell_value)
            else:
                messagebox.showerror("错误", "不支持的Excel文件格式")
                return
            
            # 弹出窗口让用户选择表头
            if not headers:
                messagebox.showerror("错误", "Excel文件中没有找到表头")
                return
            
            # 创建选择窗口
            select_window = tk.Toplevel(self.root)
            select_window.title("选择姓名表头")
            select_window.geometry("400x300")
            
            # 变量
            selected_header = tk.StringVar(value=headers[0])
            
            # 创建列表框
            listbox = tk.Listbox(select_window, selectmode=tk.SINGLE, height=10)
            for header in headers:
                listbox.insert(tk.END, header)
            listbox.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)
            
            # 选择按钮
            def on_select():
                selected_index = listbox.curselection()
                if selected_index:
                    selected_header.set(listbox.get(selected_index[0]))
                select_window.destroy()
            
            select_btn = ttk.Button(select_window, text="确定", command=on_select)
            select_btn.pack(pady=10)
            
            # 等待用户选择
            self.root.wait_window(select_window)
            
            # 获取选中的表头
            self.rename_name_header = selected_header.get()
            self.rename_name_header_var.set(self.rename_name_header)
            messagebox.showinfo("信息", f"已选择姓名表头: {self.rename_name_header}")
            self.update_button_state()
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")
    
    def select_rename_id_header(self):
        # 选择序号表头
        if not self.rename_excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        if not self.rename_name_header:
            messagebox.showerror("错误", "请先选择姓名表头")
            return
        
        try:
            # 获取文件扩展名
            ext = os.path.splitext(self.rename_excel_file)[1].lower()
            headers = []
            
            if ext == '.xlsx':
                # 使用openpyxl读取xlsx文件
                wb = openpyxl.load_workbook(self.rename_excel_file)
                ws = wb.active
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 1  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(1, min(max_check_rows, ws.max_row) + 1):
                    non_empty_cells = 0
                    for cell in ws[row]:
                        if cell.value:
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for cell in ws[header_row]:
                    if cell.value:
                        headers.append(cell.value)
            elif ext == '.xls':
                # 使用xlrd读取xls文件
                wb = xlrd.open_workbook(self.rename_excel_file)
                ws = wb.sheet_by_index(0)
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 0  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(min(max_check_rows, ws.nrows)):
                    non_empty_cells = 0
                    for col in range(ws.ncols):
                        if ws.cell_value(row, col):
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for col in range(ws.ncols):
                    cell_value = ws.cell_value(header_row, col)
                    if cell_value:
                        headers.append(cell_value)
            else:
                messagebox.showerror("错误", "不支持的Excel文件格式")
                return
            
            # 弹出窗口让用户选择表头
            if not headers:
                messagebox.showerror("错误", "Excel文件中没有找到表头")
                return
            
            # 创建选择窗口
            select_window = tk.Toplevel(self.root)
            select_window.title("选择序号表头")
            select_window.geometry("400x300")
            
            # 变量
            selected_header = tk.StringVar(value=headers[0])
            
            # 创建列表框
            listbox = tk.Listbox(select_window, selectmode=tk.SINGLE, height=10)
            for header in headers:
                listbox.insert(tk.END, header)
            listbox.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)
            
            # 选择按钮
            def on_select():
                selected_index = listbox.curselection()
                if selected_index:
                    selected_header.set(listbox.get(selected_index[0]))
                select_window.destroy()
            
            select_btn = ttk.Button(select_window, text="确定", command=on_select)
            select_btn.pack(pady=10)
            
            # 等待用户选择
            self.root.wait_window(select_window)
            
            # 获取选中的表头
            self.rename_id_header = selected_header.get()
            self.rename_id_header_var.set(self.rename_id_header)
            
            # 读取Excel数据并显示预览窗口
            data = self.read_excel_data_for_reorder(self.rename_excel_file, self.rename_name_header, self.rename_id_header)
            if not data:
                messagebox.showerror("错误", "无法读取Excel数据")
                return
            
            # 显示中转表预览窗口
            preview_window = ExcelDataPreviewWindow(self.root, data, "批量重命名 - Excel数据预览")
            
            # 检查用户是否确认
            if preview_window.result is not None:
                self.rename_data = preview_window.result
                messagebox.showinfo("信息", f"已选择序号表头: {self.rename_id_header}\n数据已确认")
            else:
                self.rename_id_header = None
                self.rename_id_header_var.set("未选择")
                messagebox.showinfo("信息", "已取消选择")
            
            self.update_button_state()
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")
    
    def select_reorder_name_header(self):
        # 选择姓名表头
        if not self.reorder_excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        try:
            # 获取文件扩展名
            ext = os.path.splitext(self.reorder_excel_file)[1].lower()
            headers = []
            
            if ext == '.xlsx':
                # 使用openpyxl读取xlsx文件
                wb = openpyxl.load_workbook(self.reorder_excel_file)
                ws = wb.active
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 1  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(1, min(max_check_rows, ws.max_row) + 1):
                    non_empty_cells = 0
                    for cell in ws[row]:
                        if cell.value:
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for cell in ws[header_row]:
                    if cell.value:
                        headers.append(cell.value)
            elif ext == '.xls':
                # 使用xlrd读取xls文件
                wb = xlrd.open_workbook(self.reorder_excel_file)
                ws = wb.sheet_by_index(0)
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 0  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(min(max_check_rows, ws.nrows)):
                    non_empty_cells = 0
                    for col in range(ws.ncols):
                        if ws.cell_value(row, col):
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for col in range(ws.ncols):
                    cell_value = ws.cell_value(header_row, col)
                    if cell_value:
                        headers.append(cell_value)
            else:
                messagebox.showerror("错误", "不支持的Excel文件格式")
                return
            
            # 弹出窗口让用户选择表头
            if not headers:
                messagebox.showerror("错误", "Excel文件中没有找到表头")
                return
            
            # 创建选择窗口
            select_window = tk.Toplevel(self.root)
            select_window.title("选择姓名表头")
            select_window.geometry("400x300")
            
            # 变量
            selected_header = tk.StringVar(value=headers[0])
            
            # 创建列表框
            listbox = tk.Listbox(select_window, selectmode=tk.SINGLE, height=10)
            for header in headers:
                listbox.insert(tk.END, header)
            listbox.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)
            
            # 选择按钮
            def on_select():
                selected_index = listbox.curselection()
                if selected_index:
                    selected_header.set(listbox.get(selected_index[0]))
                select_window.destroy()
            
            select_btn = ttk.Button(select_window, text="确定", command=on_select)
            select_btn.pack(pady=10)
            
            # 等待用户选择
            self.root.wait_window(select_window)
            
            # 获取选中的表头
            self.reorder_name_header = selected_header.get()
            self.reorder_name_header_var.set(self.reorder_name_header)
            messagebox.showinfo("信息", f"已选择姓名表头: {self.reorder_name_header}")
            self.update_button_state()
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")
    
    def select_reorder_id_header(self):
        # 选择序号表头
        if not self.reorder_excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        if not self.reorder_name_header:
            messagebox.showerror("错误", "请先选择姓名表头")
            return
        
        try:
            # 获取文件扩展名
            ext = os.path.splitext(self.reorder_excel_file)[1].lower()
            headers = []
            
            if ext == '.xlsx':
                # 使用openpyxl读取xlsx文件
                wb = openpyxl.load_workbook(self.reorder_excel_file)
                ws = wb.active
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 1  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(1, min(max_check_rows, ws.max_row) + 1):
                    non_empty_cells = 0
                    for cell in ws[row]:
                        if cell.value:
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for cell in ws[header_row]:
                    if cell.value:
                        headers.append(cell.value)
            elif ext == '.xls':
                # 使用xlrd读取xls文件
                wb = xlrd.open_workbook(self.reorder_excel_file)
                ws = wb.sheet_by_index(0)
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 0  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(min(max_check_rows, ws.nrows)):
                    non_empty_cells = 0
                    for col in range(ws.ncols):
                        if ws.cell_value(row, col):
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for col in range(ws.ncols):
                    cell_value = ws.cell_value(header_row, col)
                    if cell_value:
                        headers.append(cell_value)
            else:
                messagebox.showerror("错误", "不支持的Excel文件格式")
                return
            
            # 弹出窗口让用户选择表头
            if not headers:
                messagebox.showerror("错误", "Excel文件中没有找到表头")
                return
            
            # 创建选择窗口
            select_window = tk.Toplevel(self.root)
            select_window.title("选择序号表头")
            select_window.geometry("400x300")
            
            # 变量
            selected_header = tk.StringVar(value=headers[0])
            
            # 创建列表框
            listbox = tk.Listbox(select_window, selectmode=tk.SINGLE, height=10)
            for header in headers:
                listbox.insert(tk.END, header)
            listbox.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)
            
            # 选择按钮
            def on_select():
                selected_index = listbox.curselection()
                if selected_index:
                    selected_header.set(listbox.get(selected_index[0]))
                select_window.destroy()
            
            select_btn = ttk.Button(select_window, text="确定", command=on_select)
            select_btn.pack(pady=10)
            
            # 等待用户选择
            self.root.wait_window(select_window)
            
            # 获取选中的表头
            self.reorder_id_header = selected_header.get()
            self.reorder_id_header_var.set(self.reorder_id_header)
            
            # 读取Excel数据并显示预览窗口
            data = self.read_excel_data_for_reorder(self.reorder_excel_file, self.reorder_name_header, self.reorder_id_header)
            if not data:
                messagebox.showerror("错误", "无法读取Excel数据")
                return
            
            # 显示中转表预览窗口
            preview_window = ExcelDataPreviewWindow(self.root, data, "PDF重排序 - Excel数据预览")
            
            # 检查用户是否确认
            if preview_window.result is not None:
                self.reorder_data = preview_window.result
                messagebox.showinfo("信息", f"已选择序号表头: {self.reorder_id_header}\n数据已确认")
            else:
                self.reorder_id_header = None
                self.reorder_id_header_var.set("未选择")
                messagebox.showinfo("信息", "已取消选择")
            
            self.update_button_state()
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")
    
    def open_reorder_preview(self):
        # 打开中转表预览窗口（可随时查看和修改）
        if not self.reorder_excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        if not self.reorder_name_header or not self.reorder_id_header:
            messagebox.showerror("错误", "请先选择姓名表头和序号表头")
            return
        
        try:
            # 如果已经有中转表数据，使用现有数据
            if hasattr(self, 'reorder_data') and self.reorder_data:
                data = self.reorder_data
            else:
                # 否则从Excel读取数据
                data = self.read_excel_data_for_reorder(self.reorder_excel_file, self.reorder_name_header, self.reorder_id_header)
            
            if not data:
                messagebox.showerror("错误", "无法读取Excel数据")
                return
            
            # 显示中转表预览窗口
            preview_window = ExcelDataPreviewWindow(self.root, data, "PDF重排序 - Excel数据预览")
            
            # 检查用户是否确认
            if preview_window.result is not None:
                self.reorder_data = preview_window.result
                messagebox.showinfo("信息", "中转表数据已更新")
            else:
                messagebox.showinfo("信息", "已取消修改")
            
        except Exception as e:
            messagebox.showerror("错误", f"打开中转表预览失败: {str(e)}")
    
    def open_rename_preview(self):
        # 打开中转表预览窗口（可随时查看和修改）
        if not self.rename_excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        if not self.rename_name_header or not self.rename_id_header:
            messagebox.showerror("错误", "请先选择姓名表头和序号表头")
            return
        
        try:
            # 如果已经有中转表数据，使用现有数据
            if hasattr(self, 'rename_data') and self.rename_data:
                data = self.rename_data
            else:
                # 否则从Excel读取数据
                data = self.read_excel_data_for_reorder(self.rename_excel_file, self.rename_name_header, self.rename_id_header)
            
            if not data:
                messagebox.showerror("错误", "无法读取Excel数据")
                return
            
            # 显示中转表预览窗口
            preview_window = ExcelDataPreviewWindow(self.root, data, "批量重命名 - Excel数据预览")
            
            # 检查用户是否确认
            if preview_window.result is not None:
                self.rename_data = preview_window.result
                messagebox.showinfo("信息", "中转表数据已更新")
            else:
                messagebox.showinfo("信息", "已取消修改")
            
        except Exception as e:
            messagebox.showerror("错误", f"打开中转表预览失败: {str(e)}")
    
    def select_excel_file(self):
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("Excel文件", "*.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.excel_file = file_path
            self.excel_var.set(os.path.basename(file_path))
            self.update_button_state()
    
    def is_valid_pdf(self, file_path):
        try:
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                return True
        except Exception:
            return False
    
    def process_batch_extract(self):
        # 批量提取PDF页面逻辑
        total_files = len(self.batch_extract_pdfs)
        processed_files = 0
        
        # 如果还没有提取姓名，先自动提取
        if not self.batch_extract_extracted_names:
            self.auto_extract_names_batch_extract()
        
        # 读取Excel数据（如果有）
        use_excel = bool(self.batch_extract_excel_files and self.batch_extract_name_header and self.batch_extract_class_header)
        data_index = {} if not use_excel else self.read_excel_data(
            self.batch_extract_excel_files, 
            self.batch_extract_name_header, 
            self.batch_extract_class_header
        )
        
        for pdf_path in self.batch_extract_pdfs:
            try:
                # 读取PDF文件
                with open(pdf_path, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    total_pages = len(reader.pages)
                
                # 检查页码是否有效
                if self.batch_extract_page > total_pages:
                    messagebox.showwarning("警告", f"文件 {os.path.basename(pdf_path)} 只有 {total_pages} 页，无法提取第 {self.batch_extract_page} 页")
                    processed_files += 1
                    continue
                
                # 创建输出PDF
                writer = PyPDF2.PdfWriter()
                
                # 提取指定页码（注意：PyPDF2的页码从0开始）
                with open(pdf_path, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    writer.add_page(reader.pages[self.batch_extract_page - 1])
                
                # 生成输出文件名
                base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                
                # 获取提取的姓名
                extracted_name = self.batch_extract_extracted_names.get(pdf_path, base_name)
                
                # 如果使用Excel匹配
                if use_excel:
                    # 使用提取的姓名进行匹配
                    matched_name, matched_class = self.match_filename_with_excel(extracted_name, data_index)
                    if matched_name and matched_class:
                        output_filename = f"{matched_name}_{matched_class}.pdf"
                    else:
                        output_filename = f"{base_name}-{self.batch_extract_page}页.pdf"
                else:
                    output_filename = f"{base_name}-{self.batch_extract_page}页.pdf"
                
                output_path = os.path.join(self.batch_extract_output_dir, output_filename)
                
                # 保存提取的页面
                with open(output_path, 'wb') as f:
                    writer.write(f)
                
                processed_files += 1
                
                # 显示进度
                progress = processed_files / total_files * 100
                self.status_var.set(f"处理中... {progress:.1f}%")
                self.root.update()
                
            except Exception as e:
                messagebox.showwarning("警告", f"处理文件 {os.path.basename(pdf_path)} 时出错：{str(e)}")
                processed_files += 1
                continue
        
        messagebox.showinfo("成功", f"批量提取完成！共处理 {processed_files} 个文件，保存到 {self.batch_extract_output_dir}")
    
    def read_excel_data(self, excel_files, name_header, class_header):
        """读取Excel数据并创建索引，支持前三行表头和合并单元格"""
        data_index = {}
        
        for excel_file in excel_files:
            try:
                if excel_file.endswith('.xlsx'):
                    import openpyxl
                    workbook = openpyxl.load_workbook(excel_file, data_only=False)  # 修改为False，读取公式和值
                    sheet = workbook.active
                    
                    # 获取合并单元格信息
                    merged_cells = sheet.merged_cells
                    
                    # 扫描前三行找到表头行
                    header_row = -1
                    headers = []
                    
                    for row_idx in range(1, 4):
                        row_headers = []
                        for col_idx in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            
                            # 检查是否是合并单元格
                            is_merged = False
                            for merged_range in merged_cells:
                                if cell.coordinate in merged_range:
                                    is_merged = True
                                    if cell.coordinate == merged_range.start_cell.coordinate:
                                        if cell.value is not None:
                                            row_headers.append(str(cell.value))
                                        else:
                                            row_headers.append("")
                                    break
                            
                            if not is_merged:
                                if cell.value is not None:
                                    row_headers.append(str(cell.value))
                                else:
                                    row_headers.append("")
                        
                        # 检查这一行是否包含所需的表头
                        if name_header in row_headers and class_header in row_headers:
                            header_row = row_idx
                            headers = row_headers
                            break
                    
                    if header_row == -1:
                        messagebox.showwarning("警告", f"Excel文件 {os.path.basename(excel_file)} 中未找到指定的表头")
                        continue
                    
                    # 查找表头列索引
                    name_col = headers.index(name_header) if name_header in headers else -1
                    class_col = headers.index(class_header) if class_header in headers else -1
                    
                    if name_col == -1 or class_col == -1:
                        messagebox.showwarning("警告", f"Excel文件 {os.path.basename(excel_file)} 中未找到指定的表头")
                        continue
                    
                    # 从表头行的下一行开始读取数据
                    for row in sheet.iter_rows(min_row=header_row + 1):  # 移除values_only=True
                        cell_value = row[name_col].value
                        if cell_value:
                            name = str(cell_value)
                            class_cell_value = row[class_col].value
                            class_name = str(class_cell_value) if class_cell_value else ''
                            data_index[name] = class_name
                else:
                    import xlrd
                    workbook = xlrd.open_workbook(excel_file)
                    sheet = workbook.sheet_by_index(0)
                    
                    # 扫描前三行找到表头行
                    header_row = -1
                    headers = []
                    
                    for row_idx in range(3):
                        row_headers = [str(sheet.cell(row_idx, col).value) if sheet.cell(row_idx, col).value else '' for col in range(sheet.ncols)]
                        if name_header in row_headers and class_header in row_headers:
                            header_row = row_idx
                            headers = row_headers
                            break
                    
                    if header_row == -1:
                        messagebox.showwarning("警告", f"Excel文件 {os.path.basename(excel_file)} 中未找到指定的表头")
                        continue
                    
                    # 查找表头列索引
                    name_col = headers.index(name_header) if name_header in headers else -1
                    class_col = headers.index(class_header) if class_header in headers else -1
                    
                    if name_col == -1 or class_col == -1:
                        messagebox.showwarning("警告", f"Excel文件 {os.path.basename(excel_file)} 中未找到指定的表头")
                        continue
                    
                    # 从表头行的下一行开始读取数据
                    for row_num in range(header_row + 1, sheet.nrows):
                        name = str(sheet.cell(row_num, name_col).value)
                        if name:
                            class_name = str(sheet.cell(row_num, class_col).value) if sheet.cell(row_num, class_col).value else ''
                            data_index[name] = class_name
            except Exception as e:
                messagebox.showwarning("警告", f"读取Excel文件 {os.path.basename(excel_file)} 时出错：{str(e)}")
                continue
        
        return data_index
    
    def match_filename_with_excel(self, filename, data_index):
        """根据文件名匹配Excel数据"""
        # 移除文件扩展名（如果有）
        base_name = os.path.splitext(filename)[0]
        
        # 尝试直接匹配
        if base_name in data_index:
            return base_name, data_index[base_name]
        
        # 尝试部分匹配（例如文件名包含姓名）
        for name in data_index:
            if name in base_name:
                return name, data_index[name]
        
        # 尝试反向匹配（Excel姓名包含在提取的姓名中）
        for name in data_index:
            if base_name in name:
                return name, data_index[name]
        
        return None, None
    
    def extract_name_from_filename(self, filename, special_char=None):
        """从文件名中提取姓名"""
        import re
        
        # 移除文件扩展名
        base_name = os.path.splitext(filename)[0]
        
        if special_char:
            # 使用特殊符号提取
            if special_char in base_name:
                return base_name.split(special_char)[0].strip()
        else:
            # 自动提取：提取所有汉字部分
            chinese_chars = re.findall(r'[\u4e00-\u9fa5]+', base_name)
            if chinese_chars:
                return ''.join(chinese_chars)
        
        # 如果无法提取，返回原文件名
        return base_name
    
    def auto_extract_names_batch_extract(self):
        """自动提取批量提取功能中的姓名"""
        if not self.batch_extract_pdfs:
            messagebox.showwarning("警告", "请先选择PDF文件")
            return
        
        # 提取姓名
        self.batch_extract_extracted_names = {}
        for pdf_path in self.batch_extract_pdfs:
            filename = os.path.basename(pdf_path)
            extracted_name = self.extract_name_from_filename(filename)
            self.batch_extract_extracted_names[pdf_path] = extracted_name
        
        messagebox.showinfo("成功", f"已自动提取 {len(self.batch_extract_extracted_names)} 个文件的姓名")
    
    def custom_extract_names_batch_extract(self):
        """使用自定义特殊符号提取批量提取功能中的姓名"""
        if not self.batch_extract_pdfs:
            messagebox.showwarning("警告", "请先选择PDF文件")
            return
        
        special_char = self.batch_extract_special_char_var.get()
        if not special_char:
            messagebox.showwarning("警告", "请输入特殊符号")
            return
        
        # 提取姓名
        self.batch_extract_extracted_names = {}
        for pdf_path in self.batch_extract_pdfs:
            filename = os.path.basename(pdf_path)
            extracted_name = self.extract_name_from_filename(filename, special_char)
            self.batch_extract_extracted_names[pdf_path] = extracted_name
        
        messagebox.showinfo("成功", f"已使用特殊符号 '{special_char}' 提取 {len(self.batch_extract_extracted_names)} 个文件的姓名")
    
    def preview_name_extraction_batch_extract(self):
        """预览批量提取功能中的姓名提取结果"""
        if not self.batch_extract_pdfs:
            messagebox.showwarning("警告", "请先选择PDF文件")
            return
        
        # 如果还没有提取姓名，先自动提取
        if not self.batch_extract_extracted_names:
            self.auto_extract_names_batch_extract()
        
        # 创建预览窗口
        preview_window = tk.Toplevel(self.root)
        preview_window.title("姓名提取预览 - 批量提取")
        preview_window.geometry("800x600")
        
        # 创建滚动条
        scrollbar = ttk.Scrollbar(preview_window)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建文本框
        text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('微软雅黑', 10))
        text.pack(fill=tk.BOTH, expand=True)
        
        # 配置滚动条
        scrollbar.config(command=text.yview)
        
        # 显示标题
        text.insert(tk.END, "姓名提取预览 - 批量提取\n")
        text.insert(tk.END, "=" * 80 + "\n")
        text.insert(tk.END, f"文件数量：{len(self.batch_extract_pdfs)}\n")
        text.insert(tk.END, "\n")
        
        # 显示提取结果
        text.insert(tk.END, "提取结果：\n")
        text.insert(tk.END, "-" * 80 + "\n")
        
        for pdf_path, extracted_name in list(self.batch_extract_extracted_names.items())[:10]:  # 只显示前10个
            filename = os.path.basename(pdf_path)
            text.insert(tk.END, f"原文件名：{filename}\n")
            text.insert(tk.END, f"提取姓名：{extracted_name}\n")
            text.insert(tk.END, "-" * 80 + "\n")
        
        if len(self.batch_extract_extracted_names) > 10:
            text.insert(tk.END, f"... 还有 {len(self.batch_extract_extracted_names) - 10} 个文件未显示\n")
        
        # 禁用文本编辑
        text.config(state=tk.DISABLED)

    def select_header(self):
        # 选择表头
        if not self.excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        try:
            # 获取文件扩展名
            ext = os.path.splitext(self.excel_file)[1].lower()
            headers = []
            
            if ext == '.xlsx':
                # 使用openpyxl读取xlsx文件
                wb = openpyxl.load_workbook(self.excel_file)
                ws = wb.active
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 1  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(1, min(max_check_rows, ws.max_row) + 1):
                    non_empty_cells = 0
                    for cell in ws[row]:
                        if cell.value:
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for cell in ws[header_row]:
                    if cell.value:
                        headers.append(cell.value)
            elif ext == '.xls':
                # 使用xlrd读取xls文件
                wb = xlrd.open_workbook(self.excel_file)
                ws = wb.sheet_by_index(0)
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 0  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(min(max_check_rows, ws.nrows)):
                    non_empty_cells = 0
                    for col in range(ws.ncols):
                        if ws.cell_value(row, col):
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for col in range(ws.ncols):
                    cell_value = ws.cell_value(header_row, col)
                    if cell_value:
                        headers.append(cell_value)
            else:
                messagebox.showerror("错误", "不支持的Excel文件格式")
                return
            
            # 弹出窗口让用户选择表头
            if not headers:
                messagebox.showerror("错误", "Excel文件中没有找到表头")
                return
            
            # 创建选择窗口
            select_window = tk.Toplevel(self.root)
            select_window.title("选择表头")
            select_window.geometry("400x300")
            
            # 变量
            selected_header = tk.StringVar(value=headers[0])
            
            # 创建列表框
            listbox = tk.Listbox(select_window, selectmode=tk.SINGLE, height=10)
            for header in headers:
                listbox.insert(tk.END, header)
            listbox.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)
            
            # 选择按钮
            def on_select():
                selected_index = listbox.curselection()
                if selected_index:
                    selected_header.set(listbox.get(selected_index[0]))
                select_window.destroy()
            
            select_btn = ttk.Button(select_window, text="确定", command=on_select)
            select_btn.pack(pady=10)
            
            # 等待用户选择
            self.root.wait_window(select_window)
            
            # 获取选中的表头
            self.selected_header = selected_header.get()
            self.header_var.set(self.selected_header)
            messagebox.showinfo("信息", f"已选择姓名表头: {self.selected_header}")
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")

    def select_id_header(self):
        # 选择序号表头
        if not self.excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        if not self.selected_header:
            messagebox.showerror("错误", "请先选择姓名表头")
            return
        
        try:
            # 获取文件扩展名
            ext = os.path.splitext(self.excel_file)[1].lower()
            headers = []
            
            if ext == '.xlsx':
                # 使用openpyxl读取xlsx文件
                wb = openpyxl.load_workbook(self.excel_file)
                ws = wb.active
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 1  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(1, min(max_check_rows, ws.max_row) + 1):
                    non_empty_cells = 0
                    for cell in ws[row]:
                        if cell.value:
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for cell in ws[header_row]:
                    if cell.value:
                        headers.append(cell.value)
            elif ext == '.xls':
                # 使用xlrd读取xls文件
                wb = xlrd.open_workbook(self.excel_file)
                ws = wb.sheet_by_index(0)
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 0  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(min(max_check_rows, ws.nrows)):
                    non_empty_cells = 0
                    for col in range(ws.ncols):
                        if ws.cell_value(row, col):
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for col in range(ws.ncols):
                    cell_value = ws.cell_value(header_row, col)
                    if cell_value:
                        headers.append(cell_value)
            else:
                messagebox.showerror("错误", "不支持的Excel文件格式")
                return
            
            # 弹出窗口让用户选择表头
            if not headers:
                messagebox.showerror("错误", "Excel文件中没有找到表头")
                return
            
            # 创建选择窗口
            select_window = tk.Toplevel(self.root)
            select_window.title("选择序号表头")
            select_window.geometry("400x300")
            
            # 变量
            selected_header = tk.StringVar(value=headers[0])
            
            # 创建列表框
            listbox = tk.Listbox(select_window, selectmode=tk.SINGLE, height=10)
            for header in headers:
                listbox.insert(tk.END, header)
            listbox.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)
            
            # 选择按钮
            def on_select():
                selected_index = listbox.curselection()
                if selected_index:
                    selected_header.set(listbox.get(selected_index[0]))
                select_window.destroy()
            
            select_btn = ttk.Button(select_window, text="确定", command=on_select)
            select_btn.pack(pady=10)
            
            # 等待用户选择
            self.root.wait_window(select_window)
            
            # 获取选中的表头
            self.selected_id_header = selected_header.get()
            self.id_header_var.set(self.selected_id_header)
            messagebox.showinfo("信息", f"已选择序号表头: {self.selected_id_header}")
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")

    def open_split_preview(self):
        # 打开中转表预览窗口（可随时查看和修改）
        if not self.excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        if not self.selected_header or not self.selected_id_header:
            messagebox.showerror("错误", "请先选择姓名表头和序号表头")
            return
        
        try:
            # 如果已经有中转表数据，使用现有数据
            if hasattr(self, 'split_data') and self.split_data:
                data = self.split_data
            else:
                # 否则从Excel读取数据
                data = self.read_excel_data(self.excel_file, self.selected_header, self.selected_id_header)
            
            if not data:
                messagebox.showerror("错误", "无法读取Excel数据")
                return
            
            # 显示中转表预览窗口
            preview_window = ExcelDataPreviewWindow(self.root, data, "一分多 - Excel数据预览")
            
            # 检查用户是否确认
            if preview_window.result is not None:
                self.split_data = preview_window.result
                messagebox.showinfo("信息", "中转表数据已更新")
            else:
                messagebox.showinfo("信息", "已取消修改")
        except Exception as e:
            messagebox.showerror("错误", f"打开预览窗口失败: {str(e)}")

    def preview_mapping(self):
        # 预览序号、姓名和子PDF的对应关系
        try:
            # 检查是否已选择Excel文件
            if not self.excel_file:
                messagebox.showerror("错误", "请先选择Excel文件")
                return
            
            # 检查是否已选择姓名表头
            if not self.selected_header:
                messagebox.showerror("错误", "请先选择姓名表头")
                return
            
            # 检查是否已选择序号表头
            if not self.selected_id_header:
                messagebox.showerror("错误", "请先选择序号表头")
                return
            
            # 如果有预览窗口确认的数据，直接使用
            if hasattr(self, 'split_data') and self.split_data and len(self.split_data) > 0:
                mapping = [(int(item[0]), item[1]) for item in self.split_data]
            else:
                # 如果没有预览数据，从Excel文件中读取
                # 获取文件扩展名
                ext = os.path.splitext(self.excel_file)[1].lower()
                mapping = []
                
                if ext == '.xlsx':
                    # 使用openpyxl读取xlsx文件
                    wb = openpyxl.load_workbook(self.excel_file)
                    ws = wb.active
                    
                    # 查找实际的表头行（跳过合并单元格行）
                    header_row = 1  # 默认第一行为表头
                    max_check_rows = 5  # 最多检查前5行
                    
                    # 检查前几行，找到第一个包含多个非空值的行作为表头
                    for row in range(1, min(max_check_rows, ws.max_row) + 1):
                        non_empty_cells = 0
                        for cell in ws[row]:
                            if cell.value:
                                non_empty_cells += 1
                        if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                            header_row = row
                            break
                    
                    # 查找姓名表头所在列
                    name_col = None
                    for col in range(1, ws.max_column + 1):
                        if ws.cell(row=header_row, column=col).value == self.selected_header:
                            name_col = col
                            break
                    
                    if not name_col:
                        messagebox.showerror("错误", "未找到选中的姓名表头")
                        return
                    
                    # 查找序号表头所在列
                    id_col = None
                    for col in range(1, ws.max_column + 1):
                        if ws.cell(row=header_row, column=col).value == self.selected_id_header:
                            id_col = col
                            break
                    
                    if not id_col:
                        messagebox.showerror("错误", "未找到选中的序号表头")
                        return
                    
                    # 提取数据
                    for row in range(header_row + 1, ws.max_row + 1):
                        name_value = ws.cell(row=row, column=name_col).value
                        id_value = ws.cell(row=row, column=id_col).value
                        if name_value and id_value:
                            try:
                                id_num = int(id_value)
                                mapping.append((id_num, str(name_value)))
                            except (ValueError, TypeError):
                                pass
                elif ext == '.xls':
                    # 使用xlrd读取xls文件
                    wb = xlrd.open_workbook(self.excel_file)
                    ws = wb.sheet_by_index(0)
                    
                    # 查找实际的表头行（跳过合并单元格行）
                    header_row = 0  # 默认第一行为表头
                    max_check_rows = 5  # 最多检查前5行
                    
                    # 检查前几行，找到第一个包含多个非空值的行作为表头
                    for row in range(min(max_check_rows, ws.nrows)):
                        non_empty_cells = 0
                        for col in range(ws.ncols):
                            if ws.cell_value(row, col):
                                non_empty_cells += 1
                        if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                            header_row = row
                            break
                    
                    # 查找姓名表头所在列
                    name_col = None
                    for col in range(ws.ncols):
                        if ws.cell_value(header_row, col) == self.selected_header:
                            name_col = col
                            break
                    
                    if not name_col:
                        messagebox.showerror("错误", "未找到选中的姓名表头")
                        return
                    
                    # 查找序号表头所在列
                    id_col = None
                    for col in range(ws.ncols):
                        if ws.cell_value(header_row, col) == self.selected_id_header:
                            id_col = col
                            break
                    
                    if not id_col:
                        messagebox.showerror("错误", "未找到选中的序号表头")
                        return
                    
                    # 提取数据
                    for row in range(header_row + 1, ws.nrows):
                        name_value = ws.cell_value(row, name_col)
                        id_value = ws.cell_value(row, id_col)
                        if name_value and id_value:
                            try:
                                id_num = int(id_value)
                                mapping.append((id_num, str(name_value)))
                            except (ValueError, TypeError):
                                pass
                else:
                    messagebox.showerror("错误", "不支持的Excel文件格式")
                    return
            
            # 按序号排序
            mapping.sort(key=lambda x: x[0])
            
            # 计算预计输出的子PDF数量
            expected_count = len(mapping)
            
            # 创建预览窗口
            preview_window = tk.Toplevel(self.root)
            preview_window.title("序号-姓名-子PDF对应关系")
            preview_window.geometry("600x400")
            
            # 创建滚动条
            scrollbar = ttk.Scrollbar(preview_window)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 创建文本框
            text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('Courier New', 10))
            text.pack(fill=tk.BOTH, expand=True)
            
            # 配置滚动条
            scrollbar.config(command=text.yview)
            
            # 显示标题
            text.insert(tk.END, "序号 → 姓名 → 子PDF\n")
            text.insert(tk.END, "=" * 50 + "\n")
            
            # 显示对应关系
            for i, (id_num, name) in enumerate(mapping, 1):
                pdf_name = f"{name}.pdf"
                text.insert(tk.END, f"{id_num:4d} → {name:20s} → {pdf_name}\n")
            
            # 显示统计信息
            text.insert(tk.END, "=" * 50 + "\n")
            text.insert(tk.END, f"序号数量: {expected_count}\n")
            text.insert(tk.END, f"预计输出子PDF数量: {expected_count}\n")
            
            # 禁用文本编辑
            text.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("错误", f"预览对应关系失败: {str(e)}")

    def preview_swap(self):
        # 预览页面交换结果
        if not self.swap_pdfs:
            messagebox.showerror("错误", "请选择要交换页面的PDF文件")
            return
        
        try:
            # 获取页面A和页面B
            page_a = int(self.page_a_var.get())
            page_b = int(self.page_b_var.get())
            
            if page_a < 1 or page_b < 1:
                messagebox.showerror("错误", "页面编号必须大于0")
                return
            
            # 创建预览窗口
            preview_window = tk.Toplevel(self.root)
            preview_window.title("页面交换预览")
            preview_window.geometry("600x400")
            
            # 创建滚动条
            scrollbar = ttk.Scrollbar(preview_window)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 创建文本框
            text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('Courier New', 10))
            text.pack(fill=tk.BOTH, expand=True)
            
            # 配置滚动条
            scrollbar.config(command=text.yview)
            
            # 显示标题
            text.insert(tk.END, "PDF文件页面交换预览\n")
            text.insert(tk.END, "=" * 60 + "\n")
            text.insert(tk.END, f"交换页面：{page_a} ↔ {page_b}\n\n")
            
            # 显示每个文件的预览
            for pdf_path in self.swap_pdfs:
                pdf_name = os.path.basename(pdf_path)
                text.insert(tk.END, f"文件：{pdf_name}\n")
                
                try:
                    with open(pdf_path, 'rb') as f:
                        reader = PyPDF2.PdfReader(f)
                        total_pages = len(reader.pages)
                    
                    if page_a > total_pages or page_b > total_pages:
                        text.insert(tk.END, f"  错误：文件页数不足（共{total_pages}页）\n\n")
                    else:
                        text.insert(tk.END, f"  总页数：{total_pages}\n")
                        text.insert(tk.END, f"  交换后：页面{page_a} ↔ 页面{page_b}\n\n")
                except Exception as e:
                    text.insert(tk.END, f"  错误：{str(e)}\n\n")
            
            # 显示统计信息
            text.insert(tk.END, "=" * 60 + "\n")
            text.insert(tk.END, f"总文件数：{len(self.swap_pdfs)}\n")
            
            # 禁用文本编辑
            text.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("错误", f"预览失败: {str(e)}")

    def add_match_field(self):
        """添加匹配字段"""
        if not hasattr(self, 'rename_excel_file') or not self.rename_excel_file:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return
        
        # 读取Excel表头
        try:
            headers = self.get_excel_headers(self.rename_excel_file)
            if not headers:
                messagebox.showerror("错误", "无法读取Excel表头")
                return
            
            # 创建选择对话框
            dialog = tk.Toplevel(self.root)
            dialog.title("选择匹配字段")
            dialog.geometry("400x300")
            dialog.transient(self.root)
            dialog.grab_set()
            
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            # 匹配字段选择
            ttk.Label(frame, text="匹配字段:").pack(anchor=tk.W, pady=5)
            match_var = tk.StringVar()
            match_combo = ttk.Combobox(frame, textvariable=match_var, values=headers, width=30)
            match_combo.pack(anchor=tk.W, pady=5)
            
            # 值字段选择
            ttk.Label(frame, text="值字段:").pack(anchor=tk.W, pady=5)
            value_var = tk.StringVar()
            value_combo = ttk.Combobox(frame, textvariable=value_var, values=headers, width=30)
            value_combo.pack(anchor=tk.W, pady=5)
            
            # 按钮
            btn_frame = ttk.Frame(frame)
            btn_frame.pack(fill=tk.X, pady=10)
            
            def add_field():
                match_field = match_var.get()
                value_field = value_var.get()
                if match_field and value_field:
                    self.match_fields.append((match_field, value_field))
                    self.update_match_fields_display()
                    dialog.destroy()
                else:
                    messagebox.showwarning("警告", "请选择匹配字段和值字段")
            
            ttk.Button(btn_frame, text="添加", command=add_field).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
            
        except Exception as e:
            messagebox.showerror("错误", f"添加匹配字段时出错：{str(e)}")

    def update_match_fields_display(self):
        """更新匹配字段显示"""
        # 清空现有显示
        for widget in self.match_fields_frame.winfo_children():
            widget.destroy()
        
        # 显示匹配字段
        for i, (match_field, value_field) in enumerate(self.match_fields):
            frame = ttk.Frame(self.match_fields_frame)
            frame.pack(fill=tk.X, pady=2)
            
            ttk.Label(frame, text=f"匹配字段 {i+1}:", width=10).pack(side=tk.LEFT, padx=5)
            ttk.Label(frame, text=match_field, width=20).pack(side=tk.LEFT, padx=5)
            ttk.Label(frame, text="→", width=3).pack(side=tk.LEFT, padx=5)
            ttk.Label(frame, text=value_field, width=20).pack(side=tk.LEFT, padx=5)
            ttk.Button(frame, text="删除", command=lambda idx=i: self.remove_match_field(idx)).pack(side=tk.RIGHT, padx=5)

    def remove_match_field(self, index):
        """删除匹配字段"""
        if 0 <= index < len(self.match_fields):
            self.match_fields.pop(index)
            self.update_match_fields_display()

    def remove_filename_part(self):
        """删除最后一个文件名部分"""
        if self.filename_parts:
            self.filename_parts.pop()
            self.update_filename_preview()
            self.update_filename_parts_display()

    def update_filename_preview(self):
        """更新文件名预览"""
        if not self.filename_parts:
            self.filename_preview_var.set("点击添加元素来构建文件名")
            return
        
        preview = []
        for part_type, part_value in self.filename_parts:
            if part_type == "text":
                preview.append(part_value)
            elif part_type == "excel":
                preview.append(f"[{part_value}]")
            elif part_type == "class":
                preview.append("[班级]")
        
        self.filename_preview_var.set("" .join(preview) + ".pdf")

    def update_filename_parts_display(self):
        """更新文件名部分显示"""
        # 清空现有显示
        for widget in self.filename_parts_frame.winfo_children():
            widget.destroy()
        
        # 显示文件名部分
        for i, (part_type, part_value) in enumerate(self.filename_parts):
            frame = ttk.Frame(self.filename_parts_frame)
            frame.pack(fill=tk.X, pady=2)
            
            if part_type == "text":
                ttk.Label(frame, text="文本:", width=10).pack(side=tk.LEFT, padx=5)
                ttk.Label(frame, text=part_value, width=30).pack(side=tk.LEFT, padx=5)
            elif part_type == "excel":
                ttk.Label(frame, text="Excel列:", width=10).pack(side=tk.LEFT, padx=5)
                ttk.Label(frame, text=part_value, width=30).pack(side=tk.LEFT, padx=5)
            elif part_type == "class":
                ttk.Label(frame, text="班级:", width=10).pack(side=tk.LEFT, padx=5)
                ttk.Label(frame, text="[班级]", width=30).pack(side=tk.LEFT, padx=5)
            
            ttk.Button(frame, text="删除", command=lambda idx=i: self.remove_filename_part_at(idx)).pack(side=tk.RIGHT, padx=5)

    def remove_filename_part_at(self, index):
        """删除指定位置的文件名部分"""
        if 0 <= index < len(self.filename_parts):
            self.filename_parts.pop(index)
            self.update_filename_preview()
            self.update_filename_parts_display()

    def get_excel_headers(self, excel_file):
        """获取Excel文件的表头"""
        try:
            if excel_file.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(excel_file, data_only=True)
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1] if cell.value]
            elif excel_file.endswith('.xls'):
                workbook = xlrd.open_workbook(excel_file)
                sheet = workbook.sheet_by_index(0)
                headers = [sheet.cell_value(0, col) for col in range(sheet.ncols) if sheet.cell_value(0, col)]
            else:
                return []
            return headers
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel表头时出错：{str(e)}")
            return []

    def preview_rename(self):
        # 预览批量重命名结果
        if not self.rename_pdfs:
            messagebox.showerror("错误", "请选择要重命名的PDF文件")
            return
        
        if not self.filename_parts:
            messagebox.showerror("错误", "请先构建文件名规则")
            return
        
        # 检查是否使用了Excel列部分
        uses_excel = any(part_type == "excel" for part_type, _ in self.filename_parts)
        
        # 获取Excel数据（如果需要）
        excel_data = []
        if uses_excel:
            if not self.rename_excel_file:
                messagebox.showerror("错误", "请选择Excel文件")
                return
            
            excel_data = self.get_rename_mapping()
            if not excel_data:
                messagebox.showerror("错误", "无法获取Excel数据")
                return
        
        try:
            
            # 创建预览窗口
            preview_window = tk.Toplevel(self.root)
            preview_window.title("批量重命名预览")
            preview_window.geometry("600x400")
            
            # 创建滚动条
            scrollbar = ttk.Scrollbar(preview_window)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 创建文本框
            text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('Courier New', 10))
            text.pack(fill=tk.BOTH, expand=True)
            
            # 配置滚动条
            scrollbar.config(command=text.yview)
            
            # 显示标题
            text.insert(tk.END, "PDF批量重命名预览\n")
            text.insert(tk.END, "=" * 60 + "\n")
            text.insert(tk.END, "文件名构建规则：\n")
            
            # 显示文件名构建规则
            for i, (part_type, part_value) in enumerate(self.filename_parts):
                if part_type == "text":
                    text.insert(tk.END, f"  {i+1}. 文本: {part_value}\n")
                elif part_type == "excel":
                    text.insert(tk.END, f"  {i+1}. Excel列: {part_value}\n")
                elif part_type == "class":
                    text.insert(tk.END, f"  {i+1}. 班级名称 (特殊符号: {part_value})\n")
            text.insert(tk.END, "\n")
            
            # 按顺序显示每个文件的预览
            processed_count = 0
            for pdf_path in self.rename_pdfs:
                original_name = os.path.basename(pdf_path)
                original_filename = os.path.splitext(original_name)[0]
                text.insert(tk.END, f"原文件名：{original_name}\n")
                
                # 匹配Excel数据（如果需要）
                matched_data = {}
                matched = True
                
                if uses_excel:
                    matched = False
                    for row_data in excel_data:
                        # 检查是否有匹配的字段
                        if isinstance(row_data, dict):
                            # 处理字典格式数据
                            for key, value in row_data.items():
                                if str(value) == original_filename:
                                    matched_data = row_data
                                    matched = True
                                    break
                        else:
                            # 处理元组格式数据（旧格式）
                            for value in row_data:
                                if str(value) == original_filename:
                                    # 将元组转换为字典
                                    if hasattr(self, 'rename_name_header') and self.rename_name_header != "未选择":
                                        matched_data[self.rename_name_header] = row_data[0] if len(row_data) > 0 else ""
                                    if hasattr(self, 'rename_id_header') and self.rename_id_header != "未选择":
                                        matched_data[self.rename_id_header] = row_data[1] if len(row_data) > 1 else ""
                                    matched = True
                                    break
                        if matched:
                            break
                
                if matched:
                    # 生成新文件名
                    new_name = self.build_filename(matched_data, original_filename) + ".pdf"
                    text.insert(tk.END, f"新文件名：{new_name}\n\n")
                    processed_count += 1
                else:
                    text.insert(tk.END, "新文件名：未找到匹配数据\n\n")
            
            # 显示统计信息
            text.insert(tk.END, "=" * 60 + "\n")
            text.insert(tk.END, f"总文件数：{len(self.rename_pdfs)}\n")
            text.insert(tk.END, f"处理成功：{processed_count}\n")
            if uses_excel:
                text.insert(tk.END, f"Excel数据条数：{len(excel_data)}\n")
        except Exception as e:
            messagebox.showerror("错误", f"预览失败: {str(e)}")

    def preview_class_rename(self):
        # 预览按班级重命名结果
        if not self.rename_pdfs:
            messagebox.showerror("错误", "请先选择要重命名的PDF文件")
            return
        
        try:
            # 获取特殊符号
            special_char = self.class_rename_special_char_var.get()
            if not special_char:
                messagebox.showerror("错误", "请输入特殊符号")
                return
            
            # 准备预览数据
            class_set = set()
            preview_data = []
            
            for pdf_path in self.rename_pdfs:
                # 提取文件名（不含扩展名）
                base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                
                # 提取班级信息
                if special_char in base_name:
                    class_name = base_name.split(special_char)[-1]
                    # 去除可能的空格
                    class_name = class_name.strip()
                    
                    # 检查班级是否已存在
                    if class_name not in class_set:
                        class_set.add(class_name)
                        new_name = f"{class_name}.pdf"
                        preview_data.append((os.path.basename(pdf_path), new_name))
                else:
                    preview_data.append((os.path.basename(pdf_path), "未找到班级信息"))
            
            # 创建预览窗口
            preview_window = tk.Toplevel(self.root)
            preview_window.title("按班级重命名预览")
            preview_window.geometry("600x400")
            
            # 创建滚动条
            scrollbar = ttk.Scrollbar(preview_window)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 创建文本框
            text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('Courier New', 10))
            text.pack(fill=tk.BOTH, expand=True)
            
            # 配置滚动条
            scrollbar.config(command=text.yview)
            
            # 显示标题
            text.insert(tk.END, "按班级重命名预览\n")
            text.insert(tk.END, "=" * 60 + "\n")
            text.insert(tk.END, f"特殊符号：{special_char}\n\n")
            
            # 显示预览结果
            for original_name, new_name in preview_data:
                text.insert(tk.END, f"原文件名：{original_name}\n")
                text.insert(tk.END, f"新文件名：{new_name}\n\n")
            
            # 显示统计信息
            text.insert(tk.END, "=" * 60 + "\n")
            text.insert(tk.END, f"总文件数：{len(self.rename_pdfs)}\n")
            text.insert(tk.END, f"生成班级文件数：{len(class_set)}\n")
        except Exception as e:
            messagebox.showerror("错误", f"预览失败: {str(e)}")

    def preview_reorder(self):
        # 预览PDF重排序结果
        if not self.reorder_pdfs:
            messagebox.showerror("错误", "请选择要重排序的PDF文件")
            return
        
        if not self.reorder_excel_file:
            messagebox.showerror("错误", "请选择Excel文件")
            return
        
        if not self.reorder_name_header:
            messagebox.showerror("错误", "请选择姓名表头")
            return
        
        if not self.reorder_id_header:
            messagebox.showerror("错误", "请选择序号表头")
            return
        
        try:
            # 获取Excel数据
            if hasattr(self, 'reorder_data') and self.reorder_data:
                name_id_mapping = self.reorder_data
            else:
                name_id_mapping = self.read_excel_data(self.reorder_excel_file, self.reorder_name_header, self.reorder_id_header)
            
            if not name_id_mapping:
                messagebox.showerror("错误", "无法获取Excel数据")
                return
            
            # 创建姓名到序号的映射
            name_to_id = {name: id_value for name, id_value in name_id_mapping}
            
            # 处理PDF文件
            pdf_info = []
            for pdf_path in self.reorder_pdfs:
                original_name = os.path.basename(pdf_path)
                # 从文件名中提取姓名（假设文件名格式为"XXX_02.pdf"或"XXX-27.pdf"）
                name_match = re.match(r'^([^_\-]+)[_\-].*\.pdf$', original_name)
                if name_match:
                    extracted_name = name_match.group(1)
                else:
                    extracted_name = original_name.split('.')[0]
                
                # 获取Excel中的序号
                if extracted_name in name_to_id:
                    id_value = name_to_id[extracted_name]
                else:
                    id_value = "未找到"
                
                pdf_info.append((extracted_name, id_value, original_name, pdf_path))
            
            # 按序号排序
            def get_sort_key(item):
                id_value = item[1]
                if id_value == "未找到":
                    return (1, item[0])  # 未找到的放在最后
                try:
                    return (0, int(id_value))
                except (ValueError, TypeError):
                    return (0, id_value)
            
            pdf_info.sort(key=get_sort_key)
            
            # 创建预览窗口
            preview_window = tk.Toplevel(self.root)
            preview_window.title("PDF重排序预览")
            preview_window.geometry("700x500")
            
            # 创建滚动条
            scrollbar = ttk.Scrollbar(preview_window)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 创建文本框
            text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('Courier New', 10))
            text.pack(fill=tk.BOTH, expand=True)
            
            # 配置滚动条
            scrollbar.config(command=text.yview)
            
            # 显示标题
            text.insert(tk.END, "PDF重排序预览\n")
            text.insert(tk.END, "=" * 70 + "\n")
            text.insert(tk.END, f"姓名表头：{self.reorder_name_header}\n")
            text.insert(tk.END, f"序号表头：{self.reorder_id_header}\n\n")
            
            # 显示排序结果
            text.insert(tk.END, "排序结果：\n")
            text.insert(tk.END, "-" * 70 + "\n")
            
            for i, (name, id_value, original_name, pdf_path) in enumerate(pdf_info, 1):
                # 生成新文件名
                if id_value != "未找到":
                    try:
                        id_num = int(id_value)
                        new_name = f"{name}_{id_num:02d}.pdf"
                    except (ValueError, TypeError):
                        new_name = f"{name}_{id_value}.pdf"
                else:
                    new_name = original_name
                
                text.insert(tk.END, f"{i:3d}. 原文件名：{original_name}\n")
                text.insert(tk.END, f"    提取姓名：{name}\n")
                text.insert(tk.END, f"    Excel序号：{id_value}\n")
                text.insert(tk.END, f"    新文件名：{new_name}\n\n")
            
            # 显示统计信息
            text.insert(tk.END, "=" * 70 + "\n")
            text.insert(tk.END, f"总文件数：{len(self.reorder_pdfs)}\n")
            text.insert(tk.END, f"Excel数据条数：{len(name_id_mapping)}\n")
            
            # 计算匹配率
            matched_count = sum(1 for item in pdf_info if item[1] != "未找到")
            match_rate = (matched_count / len(pdf_info)) * 100 if pdf_info else 0
            text.insert(tk.END, f"匹配率：{match_rate:.1f}% ({matched_count}/{len(pdf_info)})\n")
            
            # 禁用文本编辑
            text.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("错误", f"预览失败: {str(e)}")

    def select_export_pdf(self):
        file_paths = filedialog.askopenfilenames(
            title="选择要导出的PDF文件（支持多选）",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_paths:
            # 验证所有PDF文件
            valid_pdfs = []
            for file_path in file_paths:
                if self.is_valid_pdf(file_path):
                    valid_pdfs.append(file_path)
            
            if valid_pdfs:
                self.export_pdfs = valid_pdfs
                if len(valid_pdfs) == 1:
                    self.export_pdf_var.set(valid_pdfs[0])
                else:
                    self.export_pdf_var.set(f"已选择 {len(valid_pdfs)} 个PDF文件")
                
                # 更新PDF数量显示
                total_pages = 0
                for pdf_path in valid_pdfs:
                    try:
                        with open(pdf_path, 'rb') as f:
                            reader = PyPDF2.PdfReader(f)
                            total_pages += len(reader.pages)
                    except:
                        pass
                
                self.export_pdf_count_var.set(f"共 {len(valid_pdfs)} 个文件，总页数: {total_pages}页")
                
                self.update_button_state()
            else:
                messagebox.showerror("错误", "选择的文件不是有效的PDF文件")

    def select_export_excel_file(self):
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("Excel文件", "*.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.export_excel_file = file_path
            self.export_excel_var.set(os.path.basename(file_path))
            self.update_button_state()

    def select_export_name_header(self):
        if not self.export_excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        try:
            # 获取文件扩展名
            ext = os.path.splitext(self.export_excel_file)[1].lower()
            headers = []
            
            if ext == '.xlsx':
                # 使用openpyxl读取xlsx文件
                wb = openpyxl.load_workbook(self.export_excel_file)
                ws = wb.active
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 1  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(1, min(max_check_rows, ws.max_row) + 1):
                    non_empty_cells = 0
                    for cell in ws[row]:
                        if cell.value:
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for cell in ws[header_row]:
                    if cell.value:
                        headers.append(cell.value)
            elif ext == '.xls':
                # 使用xlrd读取xls文件
                wb = xlrd.open_workbook(self.export_excel_file)
                ws = wb.sheet_by_index(0)
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 0  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(min(max_check_rows, ws.nrows)):
                    non_empty_cells = 0
                    for col in range(ws.ncols):
                        if ws.cell_value(row, col):
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for col in range(ws.ncols):
                    cell_value = ws.cell_value(header_row, col)
                    if cell_value:
                        headers.append(cell_value)
            else:
                messagebox.showerror("错误", "不支持的Excel文件格式")
                return
            
            # 创建选择窗口
            select_window = tk.Toplevel(self.root)
            select_window.title("选择姓名表头")
            select_window.geometry("400x300")
            
            # 创建列表框
            listbox = tk.Listbox(select_window, selectmode=tk.SINGLE, height=10)
            for header in headers:
                listbox.insert(tk.END, header)
            listbox.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)
            
            # 选择按钮
            def on_select():
                selected_index = listbox.curselection()
                if selected_index:
                    self.export_name_header = listbox.get(selected_index[0])
                    self.export_name_header_var.set(self.export_name_header)
                select_window.destroy()
            
            select_btn = ttk.Button(select_window, text="确定", command=on_select)
            select_btn.pack(pady=10)
            
            # 等待用户选择
            self.root.wait_window(select_window)
            
            # 获取选中的表头
            if self.export_name_header:
                messagebox.showinfo("信息", f"已选择姓名表头: {self.export_name_header}")
            self.update_button_state()
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")

    def select_export_id_header(self):
        if not self.export_excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        if not self.export_name_header:
            messagebox.showerror("错误", "请先选择姓名表头")
            return
        
        try:
            # 获取文件扩展名
            ext = os.path.splitext(self.export_excel_file)[1].lower()
            headers = []
            
            if ext == '.xlsx':
                # 使用openpyxl读取xlsx文件
                wb = openpyxl.load_workbook(self.export_excel_file)
                ws = wb.active
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 1  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(1, min(max_check_rows, ws.max_row) + 1):
                    non_empty_cells = 0
                    for cell in ws[row]:
                        if cell.value:
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for cell in ws[header_row]:
                    if cell.value:
                        headers.append(cell.value)
            elif ext == '.xls':
                # 使用xlrd读取xls文件
                wb = xlrd.open_workbook(self.export_excel_file)
                ws = wb.sheet_by_index(0)
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 0  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(min(max_check_rows, ws.nrows)):
                    non_empty_cells = 0
                    for col in range(ws.ncols):
                        if ws.cell_value(row, col):
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 从找到的表头行提取表头
                for col in range(ws.ncols):
                    cell_value = ws.cell_value(header_row, col)
                    if cell_value:
                        headers.append(cell_value)
            else:
                messagebox.showerror("错误", "不支持的Excel文件格式")
                return
            
            # 创建选择窗口
            select_window = tk.Toplevel(self.root)
            select_window.title("选择序号表头")
            select_window.geometry("400x300")
            
            # 创建列表框
            listbox = tk.Listbox(select_window, selectmode=tk.SINGLE, height=10)
            for header in headers:
                listbox.insert(tk.END, header)
            listbox.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)
            
            # 选择按钮
            def on_select():
                selected_index = listbox.curselection()
                if selected_index:
                    self.export_id_header = listbox.get(selected_index[0])
                    self.export_id_header_var.set(self.export_id_header)
                select_window.destroy()
            
            select_btn = ttk.Button(select_window, text="确定", command=on_select)
            select_btn.pack(pady=10)
            
            # 等待用户选择
            self.root.wait_window(select_window)
            
            # 获取选中的表头
            if self.export_id_header:
                # 读取Excel数据并显示预览窗口
                data = self.read_excel_data_for_reorder(self.export_excel_file, self.export_name_header, self.export_id_header)
                if not data:
                    messagebox.showerror("错误", "无法读取Excel数据")
                    return
                
                # 显示中转表预览窗口
                preview_window = ExcelDataPreviewWindow(self.root, data, "指定导出页 - Excel数据预览")
                
                # 检查用户是否确认
                if preview_window.result is not None:
                    self.export_data = preview_window.result
                    messagebox.showinfo("信息", f"已选择序号表头: {self.export_id_header}\n数据已确认")
                else:
                    self.export_id_header = None
                    self.export_id_header_var.set("未选择")
                    messagebox.showinfo("信息", "已取消选择")
            
            self.update_button_state()
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")

    def open_export_preview(self):
        # 打开中转表预览窗口（可随时查看和修改）
        if not self.export_excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        if not self.export_name_header or not self.export_id_header:
            messagebox.showerror("错误", "请先选择姓名表头和序号表头")
            return
        
        try:
            # 如果已经有中转表数据，使用现有数据
            if hasattr(self, 'export_data') and self.export_data:
                data = self.export_data
            else:
                # 否则从Excel读取数据
                data = self.read_excel_data_for_reorder(self.export_excel_file, self.export_name_header, self.export_id_header)
            
            if not data:
                messagebox.showerror("错误", "无法读取Excel数据")
                return
            
            # 显示中转表预览窗口
            preview_window = ExcelDataPreviewWindow(self.root, data, "指定导出页 - Excel数据预览")
            
            # 检查用户是否确认
            if preview_window.result is not None:
                self.export_data = preview_window.result
                messagebox.showinfo("信息", "中转表数据已更新")
            else:
                messagebox.showinfo("信息", "已取消修改")
            
        except Exception as e:
            messagebox.showerror("错误", f"打开中转表预览失败: {str(e)}")

    def create_sub_pdf_frame(self, parent, index):
        # 创建子PDF配置框架
        frame = ttk.LabelFrame(parent, text=f"子PDF {index+1}", padding="5")
        frame.pack(fill=tk.X, pady=3)
        
        # 页面范围
        range_frame = ttk.Frame(frame)
        range_frame.pack(fill=tk.X, pady=3)
        
        range_label = ttk.Label(range_frame, text="页面范围（如：1-20）:")
        range_label.pack(side=tk.LEFT, padx=5)
        
        range_var = tk.StringVar(value=f"{index*20+1}-{(index+1)*20}")
        range_entry = ttk.Entry(range_frame, textvariable=range_var, width=20)
        range_entry.pack(side=tk.LEFT, padx=5)
        
        # 文件名
        name_frame = ttk.Frame(frame)
        name_frame.pack(fill=tk.X, pady=3)
        
        name_label = ttk.Label(name_frame, text="文件名:")
        name_label.pack(side=tk.LEFT, padx=5)
        
        name_var = tk.StringVar(value=f"子PDF_{index+1}")
        name_entry = ttk.Entry(name_frame, textvariable=name_var, width=30)
        name_entry.pack(side=tk.LEFT, padx=5)
        
        # 保存路径
        path_frame = ttk.Frame(frame)
        path_frame.pack(fill=tk.X, pady=3)
        
        path_label = ttk.Label(path_frame, text="保存路径:")
        path_label.pack(side=tk.LEFT, padx=5)
        
        path_var = tk.StringVar()
        path_entry = ttk.Entry(path_frame, textvariable=path_var, state="readonly", width=40)
        path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        path_btn = ttk.Button(path_frame, text="选择路径", command=lambda idx=index: self.select_sub_pdf_path(idx))
        path_btn.pack(side=tk.RIGHT, padx=5)
        
        # 存储控件引用
        self.sub_pdf_frames.append(frame)
        self.sub_pdf_range_vars.append(range_var)
        self.sub_pdf_name_vars.append(name_var)
        self.sub_pdf_path_vars.append(path_var)

    def add_sub_pdf(self):
        # 添加子PDF配置
        if len(self.sub_pdf_frames) < 10:  # 限制最多10个子PDF
            parent = self.sub_pdf_frames[0].master
            index = len(self.sub_pdf_frames)
            self.create_sub_pdf_frame(parent, index)
            self.sub_pdf_count_var.set(str(len(self.sub_pdf_frames)))
            # 更新滚动区域
            self.update_scroll_region()

    def remove_sub_pdf(self):
        # 移除子PDF配置
        if len(self.sub_pdf_frames) > 1:  # 至少保留1个子PDF
            frame = self.sub_pdf_frames.pop()
            frame.destroy()
            self.sub_pdf_range_vars.pop()
            self.sub_pdf_name_vars.pop()
            self.sub_pdf_path_vars.pop()
            self.sub_pdf_count_var.set(str(len(self.sub_pdf_frames)))
            # 更新滚动区域
            self.update_scroll_region()

    def select_sub_pdf_path(self, index):
        # 选择子PDF保存路径
        path = filedialog.askdirectory(title=f"选择子PDF {index+1} 保存路径")
        if path:
            self.sub_pdf_path_vars[index].set(path)

    def parse_export_pages(self, pages_str):
        # 解析导出页面字符串，支持格式：1-6,1,2,3,4
        pages = []
        parts = pages_str.split(',')
        
        for part in parts:
            part = part.strip()
            if '-' in part:
                # 处理范围，如1-6
                start, end = part.split('-')
                try:
                    start_num = int(start.strip())
                    end_num = int(end.strip())
                    pages.extend(range(start_num, end_num + 1))
                except ValueError:
                    pass
            else:
                # 处理单个页面，如1,2,3,4
                try:
                    page_num = int(part)
                    pages.append(page_num)
                except ValueError:
                    pass
        
        return sorted(list(set(pages)))  # 去重并排序

    def process_export(self):
        # 指定导出页逻辑（支持批量处理）
        # 检查是否有子PDF配置
        if hasattr(self, 'sub_pdf_frames') and len(self.sub_pdf_frames) > 0:
            # 处理子PDF分割（批量版）
            self.process_sub_pdf_split_batch()
        else:
            # 批量页面导出逻辑
            output_dir = filedialog.askdirectory(title="选择保存目录")
            if not output_dir:
                return
            
            # 检查是否选择了PDF文件
            if not self.export_pdfs:
                messagebox.showerror("错误", "请选择要导出的PDF文件")
                return
            
            # 解析导出页面
            try:
                export_pages = self.parse_export_pages(self.export_pages_var.get())
                if not export_pages:
                    messagebox.showerror("错误", "请输入有效的导出页面")
                    return
            except Exception as e:
                messagebox.showerror("错误", f"解析导出页面失败: {str(e)}")
                return
            
            # 获取Excel数据（如果选择了Excel文件）
            name_id_mapping = []
            if self.export_excel_file and self.export_name_header and self.export_id_header:
                name_id_mapping = self.get_export_mapping()
            
            # 是否使用原文件名输出
            use_original_name = self.export_use_original_name_var.get() if hasattr(self, 'export_use_original_name_var') else False
            
            # 批量处理所有PDF文件
            total_success = 0
            total_failed = 0
            fail_messages = []
            total_files = len(self.export_pdfs)
            page_range_str = self.export_pages_var.get()
            
            for pdf_index, pdf_path in enumerate(self.export_pdfs):
                try:
                    with open(pdf_path, 'rb') as f:
                        reader = PyPDF2.PdfReader(f)
                        total_pages = len(reader.pages)
                        
                        # 验证页面范围
                        invalid_pages = [page for page in export_pages if page < 1 or page > total_pages]
                        if invalid_pages:
                            fail_msg = f"{os.path.basename(pdf_path)}: 页面超出范围"
                            fail_messages.append(fail_msg)
                            total_failed += 1
                            continue
                        
                        # 导出页面
                        for i, page_num in enumerate(export_pages):
                            try:
                                # 创建输出PDF
                                output_writer = PyPDF2.PdfWriter()
                                output_writer.add_page(reader.pages[page_num - 1])  # 页面索引从0开始
                                
                                # 生成文件名
                                if use_original_name:
                                    # 以原文件名输出
                                    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                                    filename = f"{base_name}_page{page_num}.pdf"
                                elif name_id_mapping and i < len(name_id_mapping):
                                    name, id_value = name_id_mapping[i]
                                    if id_value:
                                        filename = f"{name}序号{id_value}.pdf"
                                    else:
                                        filename = f"{name}.pdf"
                                else:
                                    filename = f"page_{page_num}.pdf"
                                
                                # 保存文件
                                output_path = os.path.join(output_dir, filename)
                                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                                
                                with open(output_path, 'wb') as f_out:
                                    output_writer.write(f_out)
                                total_success += 1
                                
                            except PermissionError:
                                fail_msg = f"{os.path.basename(pdf_path)}_page{page_num}: 没有写入权限"
                                fail_messages.append(fail_msg)
                                total_failed += 1
                                continue
                            except Exception as e:
                                fail_msg = f"{os.path.basename(pdf_path)}_page{page_num}: {str(e)}"
                                fail_messages.append(fail_msg)
                                total_failed += 1
                                continue
                        
                except Exception as e:
                    fail_msg = f"{os.path.basename(pdf_path)}: {str(e)}"
                    fail_messages.append(fail_msg)
                    total_failed += 1
                
                # 显示进度
                progress = (pdf_index + 1) / total_files * 100
                self.status_var.set(f"处理中... {progress:.1f}%")
                self.root.update()
            
            # 显示结果
            result_msg = f"批量导出完成！\n\n处理文件数：{total_files}\n成功导出：{total_success}"
            if total_failed > 0:
                result_msg += f"\n失败：{total_failed}"
                result_msg += "\n\n失败详情：\n" + "\n".join(fail_messages[:5])
                if len(fail_messages) > 5:
                    result_msg += f"\n... 还有 {len(fail_messages) - 5} 个失败项"
            
            if total_failed > 0:
                messagebox.showwarning("处理完成", result_msg)
            else:
                messagebox.showinfo("成功", result_msg)

    def process_sub_pdf_split(self):
        # 处理子PDF分割
        try:
            with open(self.export_pdf, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                total_pages = len(reader.pages)
            
            # 处理每个子PDF
            total_sub_pdfs = len(self.sub_pdf_frames)
            processed_count = 0
            
            for i in range(total_sub_pdfs):
                try:
                    # 获取子PDF配置
                    range_str = self.sub_pdf_range_vars[i].get()
                    name = self.sub_pdf_name_vars[i].get()
                    path = self.sub_pdf_path_vars[i].get()
                    
                    # 验证配置
                    if not range_str:
                        messagebox.showerror("错误", f"请输入子PDF {i+1} 的页面范围")
                        continue
                    if not name:
                        messagebox.showerror("错误", f"请输入子PDF {i+1} 的文件名")
                        continue
                    if not path:
                        messagebox.showerror("错误", f"请选择子PDF {i+1} 的保存路径")
                        continue
                    
                    # 解析页面范围
                    pages = self.parse_export_pages(range_str)
                    if not pages:
                        messagebox.showerror("错误", f"子PDF {i+1} 的页面范围无效")
                        continue
                    
                    # 验证页面范围
                    invalid_pages = [page for page in pages if page < 1 or page > total_pages]
                    if invalid_pages:
                        messagebox.showerror("错误", f"子PDF {i+1} 页面超出范围: {invalid_pages}\nPDF总页数: {total_pages}")
                        continue
                    
                    # 创建输出PDF
                    output_writer = PyPDF2.PdfWriter()
                    
                    # 添加页面
                    for page_num in pages:
                        with open(self.export_pdf, 'rb') as f:
                            reader = PyPDF2.PdfReader(f)
                            output_writer.add_page(reader.pages[page_num - 1])  # 页面索引从0开始
                    
                    # 保存文件
                    output_path = os.path.join(path, f"{name}.pdf")
                    os.makedirs(os.path.dirname(output_path), exist_ok=True)
                    
                    try:
                        with open(output_path, 'wb') as f_out:
                            output_writer.write(f_out)
                        processed_count += 1
                    except PermissionError:
                        messagebox.showerror("错误", f"没有写入权限：{output_path}\n请选择一个有写入权限的目录")
                        continue
                    except Exception as e:
                        messagebox.showerror("错误", f"保存文件时出错：{str(e)}")
                        continue
                    
                    # 显示进度
                    progress = (i + 1) / total_sub_pdfs * 100
                    self.status_var.set(f"处理中... {progress:.1f}%")
                    self.root.update()
                    
                except Exception as e:
                    messagebox.showerror("错误", f"处理子PDF {i+1} 时出错: {str(e)}")
            
            messagebox.showinfo("成功", f"子PDF分割完成！\n处理子PDF数：{processed_count}/{total_sub_pdfs}")
            
        except Exception as e:
            messagebox.showerror("错误", f"读取PDF文件失败: {str(e)}")
            return

    def process_sub_pdf_split_batch(self):
        # 批量处理子PDF分割
        # 检查是否选择了PDF文件
        if not self.export_pdfs:
            messagebox.showerror("错误", "请选择要导出的PDF文件")
            return
        
        # 是否使用原文件名输出
        use_original_name = self.export_use_original_name_var.get() if hasattr(self, 'export_use_original_name_var') else False
        
        total_success = 0
        total_failed = 0
        fail_messages = []
        total_files = len(self.export_pdfs)
        
        for pdf_index, pdf_path in enumerate(self.export_pdfs):
            try:
                with open(pdf_path, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    total_pages = len(reader.pages)
                
                # 获取PDF基础名称（用于原文件名输出）
                pdf_base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                
                # 处理每个子PDF配置
                total_sub_pdfs = len(self.sub_pdf_frames)
                
                for i in range(total_sub_pdfs):
                    try:
                        # 获取子PDF配置
                        range_str = self.sub_pdf_range_vars[i].get()
                        name = self.sub_pdf_name_vars[i].get()
                        path = self.sub_pdf_path_vars[i].get()
                        
                        # 验证配置
                        if not range_str:
                            fail_msg = f"{pdf_base_name}: 子PDF {i+1} 页面范围为空"
                            fail_messages.append(fail_msg)
                            total_failed += 1
                            continue
                        if not name:
                            fail_msg = f"{pdf_base_name}: 子PDF {i+1} 文件名为空"
                            fail_messages.append(fail_msg)
                            total_failed += 1
                            continue
                        if not path:
                            fail_msg = f"{pdf_base_name}: 子PDF {i+1} 保存路径为空"
                            fail_messages.append(fail_msg)
                            total_failed += 1
                            continue
                        
                        # 解析页面范围
                        pages = self.parse_export_pages(range_str)
                        if not pages:
                            fail_msg = f"{pdf_base_name}: 子PDF {i+1} 页面范围无效"
                            fail_messages.append(fail_msg)
                            total_failed += 1
                            continue
                        
                        # 验证页面范围
                        invalid_pages = [page for page in pages if page < 1 or page > total_pages]
                        if invalid_pages:
                            fail_msg = f"{pdf_base_name}: 子PDF {i+1} 页面超出范围"
                            fail_messages.append(fail_msg)
                            total_failed += 1
                            continue
                        
                        # 创建输出PDF
                        output_writer = PyPDF2.PdfWriter()
                        
                        # 添加页面
                        for page_num in pages:
                            output_writer.add_page(reader.pages[page_num - 1])  # 页面索引从0开始
                        
                        # 生成文件名
                        if use_original_name:
                            output_name = f"{pdf_base_name}_{name}"
                        else:
                            output_name = name
                        
                        # 保存文件
                        output_path = os.path.join(path, f"{output_name}.pdf")
                        os.makedirs(os.path.dirname(output_path), exist_ok=True)
                        
                        with open(output_path, 'wb') as f_out:
                            output_writer.write(f_out)
                        total_success += 1
                        
                    except Exception as e:
                        fail_msg = f"{pdf_base_name}: 子PDF {i+1} 处理失败 - {str(e)}"
                        fail_messages.append(fail_msg)
                        total_failed += 1
                
            except Exception as e:
                fail_msg = f"{pdf_base_name}: 读取失败 - {str(e)}"
                fail_messages.append(fail_msg)
                total_failed += 1
            
            # 显示进度
            progress = (pdf_index + 1) / total_files * 100
            self.status_var.set(f"处理中... {progress:.1f}%")
            self.root.update()
        
        # 显示结果
        result_msg = f"批量子PDF分割完成！\n\n处理文件数：{total_files}\n成功分割：{total_success}"
        if total_failed > 0:
            result_msg += f"\n失败：{total_failed}"
            result_msg += "\n\n失败详情：\n" + "\n".join(fail_messages[:5])
            if len(fail_messages) > 5:
                result_msg += f"\n... 还有 {len(fail_messages) - 5} 个失败项"
        
        if total_failed > 0:
            messagebox.showwarning("处理完成", result_msg)
        else:
            messagebox.showinfo("成功", result_msg)

    def get_export_mapping(self):
        # 从预览窗口确认的数据中获取姓名和序号映射
        try:
            # 如果有预览窗口确认的数据，直接使用
            if hasattr(self, 'export_data') and self.export_data:
                return self.export_data
            
            # 如果没有预览数据，从Excel文件中读取
            # 获取文件扩展名
            ext = os.path.splitext(self.export_excel_file)[1].lower()
            mapping = []
            
            if ext == '.xlsx':
                # 使用openpyxl读取xlsx文件
                wb = openpyxl.load_workbook(self.export_excel_file)
                ws = wb.active
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 1  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(1, min(max_check_rows, ws.max_row) + 1):
                    non_empty_cells = 0
                    for cell in ws[row]:
                        if cell.value:
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 查找姓名表头所在列
                name_col = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=header_row, column=col).value == self.export_name_header:
                        name_col = col
                        break
                
                if not name_col:
                    messagebox.showerror("错误", "未找到选中的姓名表头")
                    return []
                
                # 查找序号表头所在列
                id_col = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=header_row, column=col).value == self.export_id_header:
                        id_col = col
                        break
                
                if not id_col:
                    messagebox.showerror("错误", "未找到选中的序号表头")
                    return []
                
                # 提取数据
                for row in range(header_row + 1, ws.max_row + 1):
                    name_value = ws.cell(row=row, column=name_col).value
                    id_value = ws.cell(row=row, column=id_col).value
                    if name_value and id_value:
                        mapping.append((str(name_value), str(id_value)))
            elif ext == '.xls':
                # 使用xlrd读取xls文件
                wb = xlrd.open_workbook(self.export_excel_file)
                ws = wb.sheet_by_index(0)
                
                # 查找实际的表头行（跳过合并单元格行）
                header_row = 0  # 默认第一行为表头
                max_check_rows = 5  # 最多检查前5行
                
                # 检查前几行，找到第一个包含多个非空值的行作为表头
                for row in range(min(max_check_rows, ws.nrows)):
                    non_empty_cells = 0
                    for col in range(ws.ncols):
                        if ws.cell_value(row, col):
                            non_empty_cells += 1
                    if non_empty_cells >= 2:  # 如果一行中有2个或更多非空值，认为是表头
                        header_row = row
                        break
                
                # 查找姓名表头所在列
                name_col = None
                for col in range(ws.ncols):
                    if ws.cell_value(header_row, col) == self.export_name_header:
                        name_col = col
                        break
                
                if not name_col:
                    messagebox.showerror("错误", "未找到选中的姓名表头")
                    return []
                
                # 查找序号表头所在列
                id_col = None
                for col in range(ws.ncols):
                    if ws.cell_value(header_row, col) == self.export_id_header:
                        id_col = col
                        break
                
                if not id_col:
                    messagebox.showerror("错误", "未找到选中的序号表头")
                    return []
                
                # 提取数据
                for row in range(header_row + 1, ws.nrows):
                    name_value = ws.cell_value(row, name_col)
                    id_value = ws.cell_value(row, id_col)
                    if name_value and id_value:
                        mapping.append((str(name_value), str(id_value)))
            else:
                messagebox.showerror("错误", "不支持的Excel文件格式")
                return []
            
            return mapping
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")
            return []

    def preview_export(self):
        # 预览导出结果
        if not self.export_pdf:
            messagebox.showerror("错误", "请选择要导出的PDF文件")
            return
        
        try:
            # 读取PDF文件
            with open(self.export_pdf, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                total_pages = len(reader.pages)
            
            # 检查是否有子PDF配置
            if hasattr(self, 'sub_pdf_frames') and len(self.sub_pdf_frames) > 0:
                # 预览子PDF分割
                self.preview_sub_pdf_split(total_pages)
            else:
                # 原有的单个页面导出预览
                if not self.export_pages_var.get():
                    messagebox.showerror("错误", "请输入导出页面")
                    return
                
                # 解析导出页面
                export_pages = self.parse_export_pages(self.export_pages_var.get())
                if not export_pages:
                    messagebox.showerror("错误", "请输入有效的导出页面")
                    return
                
                # 验证页面范围
                invalid_pages = [page for page in export_pages if page < 1 or page > total_pages]
                if invalid_pages:
                    messagebox.showerror("错误", f"页面超出范围: {invalid_pages}\nPDF总页数: {total_pages}")
                    return
                
                # 获取Excel数据（如果选择了Excel文件）
                name_id_mapping = []
                if self.export_excel_file and self.export_name_header and self.export_id_header:
                    name_id_mapping = self.get_export_mapping()
                
                # 创建预览窗口
                preview_window = tk.Toplevel(self.root)
                preview_window.title("指定导出页预览")
                preview_window.geometry("600x400")
                
                # 创建滚动条
                scrollbar = ttk.Scrollbar(preview_window)
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                
                # 创建文本框
                text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('Courier New', 10))
                text.pack(fill=tk.BOTH, expand=True)
                
                # 配置滚动条
                scrollbar.config(command=text.yview)
                
                # 显示标题
                text.insert(tk.END, "指定导出页预览\n")
                text.insert(tk.END, "=" * 60 + "\n")
                text.insert(tk.END, f"PDF文件：{os.path.basename(self.export_pdf)}\n")
                text.insert(tk.END, f"PDF总页数：{total_pages}\n")
                text.insert(tk.END, f"导出页面：{self.export_pages_var.get()}\n")
                text.insert(tk.END, f"解析后页面：{export_pages}\n\n")
                
                # 显示Excel信息
                if name_id_mapping:
                    text.insert(tk.END, "Excel信息：\n")
                    text.insert(tk.END, f"  姓名表头：{self.export_name_header}\n")
                    text.insert(tk.END, f"  序号表头：{self.export_id_header}\n")
                    text.insert(tk.END, f"  数据条数：{len(name_id_mapping)}\n\n")
                
                # 显示导出文件列表
                text.insert(tk.END, "导出文件列表：\n")
                for i, page_num in enumerate(export_pages):
                    if name_id_mapping and i < len(name_id_mapping):
                        name, id_value = name_id_mapping[i]
                        if id_value:
                            filename = f"{name}序号{id_value}.pdf"
                        else:
                            filename = f"{name}.pdf"
                    else:
                        filename = f"page_{page_num}.pdf"
                    text.insert(tk.END, f"  {i+1}. 页面{page_num} -> {filename}\n")
                
                text.insert(tk.END, "\n")
                
                # 显示统计信息
                text.insert(tk.END, "=" * 60 + "\n")
                text.insert(tk.END, f"总导出页面数：{len(export_pages)}\n")
                if name_id_mapping:
                    text.insert(tk.END, f"使用Excel重命名：是\n")
                else:
                    text.insert(tk.END, f"使用Excel重命名：否\n")
                
                # 禁用文本编辑
                text.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("错误", f"预览失败: {str(e)}")

    def preview_sub_pdf_split(self, total_pages):
        # 预览子PDF分割结果
        try:
            # 创建预览窗口
            preview_window = tk.Toplevel(self.root)
            preview_window.title("子PDF分割预览")
            preview_window.geometry("600x400")
            
            # 创建滚动条
            scrollbar = ttk.Scrollbar(preview_window)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 创建文本框
            text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('Courier New', 10))
            text.pack(fill=tk.BOTH, expand=True)
            
            # 配置滚动条
            scrollbar.config(command=text.yview)
            
            # 显示标题
            text.insert(tk.END, "子PDF分割预览\n")
            text.insert(tk.END, "=" * 60 + "\n")
            text.insert(tk.END, f"PDF文件：{os.path.basename(self.export_pdf)}\n")
            text.insert(tk.END, f"PDF总页数：{total_pages}\n")
            text.insert(tk.END, f"子PDF数量：{len(self.sub_pdf_frames)}\n\n")
            
            # 显示子PDF配置
            text.insert(tk.END, "子PDF配置列表：\n")
            for i in range(len(self.sub_pdf_frames)):
                range_str = self.sub_pdf_range_vars[i].get()
                name = self.sub_pdf_name_vars[i].get()
                path = self.sub_pdf_path_vars[i].get()
                
                # 解析页面范围
                pages = self.parse_export_pages(range_str)
                
                # 验证页面范围
                invalid_pages = [page for page in pages if page < 1 or page > total_pages]
                
                text.insert(tk.END, f"子PDF {i+1}:\n")
                text.insert(tk.END, f"  页面范围：{range_str}\n")
                text.insert(tk.END, f"  解析后页面：{pages}\n")
                text.insert(tk.END, f"  文件名：{name}.pdf\n")
                text.insert(tk.END, f"  保存路径：{path}\n")
                if invalid_pages:
                    text.insert(tk.END, f"  错误：页面超出范围: {invalid_pages}\n")
                text.insert(tk.END, "\n")
            
            # 显示统计信息
            text.insert(tk.END, "=" * 60 + "\n")
            text.insert(tk.END, f"总子PDF数：{len(self.sub_pdf_frames)}\n")
            
            # 禁用文本编辑
            text.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("错误", f"预览失败: {str(e)}")

    def select_page_pdfs(self):
        file_paths = filedialog.askopenfilenames(
            title="选择PDF文件（最多100个）",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_paths:
            if len(file_paths) > 100:
                messagebox.showwarning("警告", "最多只能选择100个文件，已自动选择前100个")
                file_paths = file_paths[:100]
            
            valid_pdfs = []
            for path in file_paths:
                if self.is_valid_pdf(path):
                    valid_pdfs.append(path)
                else:
                    messagebox.showerror("错误", f"文件 {os.path.basename(path)} 不是有效的PDF文件")
            
            if valid_pdfs:
                self.page_pdfs = valid_pdfs
                self.page_pdf_var.set(f"已选择 {len(valid_pdfs)} 个文件")
                self.page_pdf_count_var.set(f"已选择 {len(valid_pdfs)} 个文件")
                self.update_button_state()

    def select_batch_extract_pdfs(self):
        file_paths = filedialog.askopenfilenames(
            title="选择PDF文件（批量提取）",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_paths:
            valid_pdfs = []
            for path in file_paths:
                if self.is_valid_pdf(path):
                    valid_pdfs.append(path)
                else:
                    messagebox.showerror("错误", f"文件 {os.path.basename(path)} 不是有效的PDF文件")
            
            if valid_pdfs:
                self.batch_extract_pdfs = valid_pdfs
                self.batch_extract_pdf_var.set(f"已选择 {len(valid_pdfs)} 个文件")
                self.update_button_state()

    def select_batch_extract_output_dir(self):
        dir_path = filedialog.askdirectory(
            title="选择输出目录"
        )
        if dir_path:
            self.batch_extract_output_dir = dir_path
            self.batch_extract_output_var.set(dir_path)
            self.update_button_state()

    def select_batch_append_source_pdf(self):
        file_path = filedialog.askopenfilename(
            title="选择源PDF文件（批量追加）",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_path:
            if self.is_valid_pdf(file_path):
                self.batch_append_source_pdf = file_path
                self.batch_append_source_var.set(file_path)
                self.update_button_state()
            else:
                messagebox.showerror("错误", "选择的文件不是有效的PDF文件")

    def reverse_pages(self):
        if not self.page_pdfs:
            messagebox.showwarning("警告", "请先选择PDF文件")
            return
        
        try:
            for pdf_path in self.page_pdfs:
                reader = PyPDF2.PdfReader(pdf_path)
                writer = PyPDF2.PdfWriter()
                
                total_pages = len(reader.pages)
                for i in range(total_pages - 1, -1, -1):
                    writer.add_page(reader.pages[i])
                
                output_path = pdf_path.replace('.pdf', '_reversed.pdf')
                with open(output_path, 'wb') as output_file:
                    writer.write(output_file)
            
            messagebox.showinfo("成功", f"已成功颠倒 {len(self.page_pdfs)} 个PDF文件的页面顺序")
            self.status_var.set(f"颠倒页面顺序完成，共处理 {len(self.page_pdfs)} 个文件")
        except Exception as e:
            messagebox.showerror("错误", f"颠倒页面顺序时出错：{str(e)}")

    def open_rotation_preview(self):
        if not self.page_pdfs:
            messagebox.showwarning("警告", "请先选择PDF文件")
            return
        
        RotationPreviewWindow(self.root, self.page_pdfs, self.apply_rotation_scheme)

    def apply_rotation_scheme(self, rotation_scheme):
        # 应用页面旋转方案
        try:
            for pdf_path, rotation in rotation_scheme.items():
                reader = PyPDF2.PdfReader(pdf_path)
                writer = PyPDF2.PdfWriter()
                
                for i, page in enumerate(reader.pages):
                    if i in rotation:
                        page.rotate(rotation[i])
                    writer.add_page(page)
                
                output_path = pdf_path.replace('.pdf', '_rotated.pdf')
                with open(output_path, 'wb') as output_file:
                    writer.write(output_file)
            
            messagebox.showinfo("成功", f"已成功旋转 {len(rotation_scheme)} 个PDF文件的页面")
            self.status_var.set(f"页面旋转完成，共处理 {len(rotation_scheme)} 个文件")
        except Exception as e:
            messagebox.showerror("错误", f"应用旋转方案时出错：{str(e)}")

    def select_batch_extract_excel_files(self):
        file_paths = filedialog.askopenfilenames(
            title="选择Excel文件（最多5个）",
            filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")]
        )
        if file_paths:
            # 限制选择5个文件
            if len(file_paths) > 5:
                messagebox.showwarning("警告", "最多只能选择5个Excel文件")
                file_paths = file_paths[:5]
            
            self.batch_extract_excel_files = file_paths
            self.batch_extract_excel_var.set(f"已选择 {len(file_paths)} 个文件")
            self.update_button_state()

    def read_excel_headers_with_merged_cells(self, excel_file):
        """读取Excel文件的表头，支持前三行和合并单元格"""
        headers = []
        try:
            if excel_file.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active
                
                # 获取合并单元格信息
                merged_cells = sheet.merged_cells
                
                # 读取前三行的数据
                for row_idx in range(1, 4):
                    row_headers = []
                    for col_idx in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        
                        # 检查是否是合并单元格
                        is_merged = False
                        for merged_range in merged_cells:
                            if cell.coordinate in merged_range:
                                is_merged = True
                                # 合并单元格只取左上角单元格的值
                                if cell.coordinate == merged_range.start_cell.coordinate:
                                    if cell.value is not None:
                                        row_headers.append(str(cell.value))
                                    else:
                                        row_headers.append("")
                                break
                        
                        if not is_merged:
                            if cell.value is not None:
                                row_headers.append(str(cell.value))
                            else:
                                row_headers.append("")
                    
                    headers.extend([h for h in row_headers if h])
            else:
                workbook = xlrd.open_workbook(excel_file)
                sheet = workbook.sheet_by_index(0)
                
                # 读取前三行的数据
                for row_idx in range(3):
                    for col_idx in range(sheet.ncols):
                        value = sheet.cell(row_idx, col_idx).value
                        if value is not None and str(value).strip():
                            headers.append(str(value))
            
            # 去重但保持顺序
            seen = set()
            unique_headers = []
            for h in headers:
                if h not in seen:
                    seen.add(h)
                    unique_headers.append(h)
            
            return unique_headers
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel表头时出错：{str(e)}")
            return []

    def select_batch_extract_name_header(self):
        if not self.batch_extract_excel_files:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return
        
        # 读取第一个Excel文件的表头
        try:
            headers = self.read_excel_headers_with_merged_cells(self.batch_extract_excel_files[0])
            
            if not headers:
                messagebox.showwarning("警告", "未能读取到有效的表头")
                return
            
            # 创建表头选择窗口
            header_window = tk.Toplevel(self.root)
            header_window.title("选择姓名表头")
            header_window.geometry("400x300")
            
            # 创建列表框
            listbox = tk.Listbox(header_window, selectmode=tk.SINGLE, width=50)
            listbox.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
            
            # 填充表头
            for header in headers:
                listbox.insert(tk.END, header)
            
            # 确定按钮
            def on_select():
                if listbox.curselection():
                    selected = listbox.get(listbox.curselection())
                    self.batch_extract_name_header = selected
                    self.batch_extract_name_header_var.set(selected)
                    header_window.destroy()
            
            ttk.Button(header_window, text="确定", command=on_select).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件时出错：{str(e)}")

    def preview_excel_content_batch_extract(self):
        """预览批量提取功能中的Excel表内容"""
        if not self.batch_extract_excel_files:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return
        
        # 预览第一个Excel文件
        excel_file = self.batch_extract_excel_files[0]
        self.preview_excel_content(excel_file)

    def preview_excel_content_batch_append(self):
        """预览批量追加页功能中的Excel表内容"""
        if not self.batch_append_excel_files:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return
        
        # 预览第一个Excel文件
        excel_file = self.batch_append_excel_files[0]
        self.preview_excel_content(excel_file)

    def preview_excel_content(self, excel_file):
        """预览Excel表内容"""
        try:
            # 读取Excel文件
            if excel_file.endswith('.xlsx'):
                import openpyxl
                workbook = openpyxl.load_workbook(excel_file, data_only=True)
                sheet = workbook.active
                
                # 获取合并单元格信息
                merged_cells = sheet.merged_cells
                
                # 获取最大行列数
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # 创建预览窗口
                preview_window = tk.Toplevel(self.root)
                preview_window.title(f"Excel预览 - {os.path.basename(excel_file)}")
                preview_window.geometry("1000x600")
                
                # 创建主框架
                main_frame = ttk.Frame(preview_window)
                main_frame.pack(fill=tk.BOTH, expand=True)
                
                # 创建横向滚动条
                h_scrollbar = ttk.Scrollbar(main_frame, orient=tk.HORIZONTAL)
                h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
                
                # 创建纵向滚动条
                v_scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL)
                v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                
                # 创建树状图显示Excel内容
                tree = ttk.Treeview(main_frame, columns=[f"col{i}" for i in range(1, max_col + 1)], 
                                   show="headings", xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)
                
                # 配置滚动条
                h_scrollbar.config(command=tree.xview)
                v_scrollbar.config(command=tree.yview)
                
                # 设置列宽
                for i in range(1, max_col + 1):
                    tree.heading(f"col{i}", text=f"列{i}")
                    tree.column(f"col{i}", width=100)
                
                # 处理表头三行
                for row_idx in range(1, min(4, max_row + 1)):
                    row_values = []
                    for col_idx in range(1, max_col + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        
                        # 检查是否是合并单元格
                        is_merged = False
                        for merged_range in merged_cells:
                            if cell.coordinate in merged_range:
                                is_merged = True
                                if cell.coordinate == merged_range.start_cell.coordinate:
                                    row_values.append(str(cell.value) if cell.value else "")
                                else:
                                    row_values.append("")
                                break
                        
                        if not is_merged:
                            row_values.append(str(cell.value) if cell.value else "")
                    
                    tree.insert("", tk.END, values=row_values, tags=("header",))
                
                # 处理数据行
                for row_idx in range(4, min(50, max_row + 1)):  # 只显示前50行
                    row_values = []
                    for col_idx in range(1, max_col + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        row_values.append(str(cell.value) if cell.value else "")
                    
                    tree.insert("", tk.END, values=row_values)
                
                # 显示树状图
                tree.pack(fill=tk.BOTH, expand=True)
                
                # 添加样式
                tree.tag_configure("header", background="lightblue")
                
                if max_row > 50:
                    # 添加提示标签
                    info_label = ttk.Label(preview_window, text=f"... 还有 {max_row - 50} 行未显示")
                    info_label.pack(pady=5)
            
            elif excel_file.endswith('.xls'):
                import xlrd
                workbook = xlrd.open_workbook(excel_file)
                sheet = workbook.sheet_by_index(0)
                
                # 获取最大行列数
                max_row = sheet.nrows
                max_col = sheet.ncols
                
                # 创建预览窗口
                preview_window = tk.Toplevel(self.root)
                preview_window.title(f"Excel预览 - {os.path.basename(excel_file)}")
                preview_window.geometry("1000x600")
                
                # 创建主框架
                main_frame = ttk.Frame(preview_window)
                main_frame.pack(fill=tk.BOTH, expand=True)
                
                # 创建横向滚动条
                h_scrollbar = ttk.Scrollbar(main_frame, orient=tk.HORIZONTAL)
                h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
                
                # 创建纵向滚动条
                v_scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL)
                v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                
                # 创建树状图显示Excel内容
                tree = ttk.Treeview(main_frame, columns=[f"col{i}" for i in range(1, max_col + 1)], 
                                   show="headings", xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)
                
                # 配置滚动条
                h_scrollbar.config(command=tree.xview)
                v_scrollbar.config(command=tree.yview)
                
                # 设置列宽
                for i in range(1, max_col + 1):
                    tree.heading(f"col{i}", text=f"列{i}")
                    tree.column(f"col{i}", width=100)
                
                # 处理表头三行
                for row_idx in range(min(3, max_row)):
                    row_values = []
                    for col_idx in range(max_col):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        row_values.append(str(cell_value) if cell_value else "")
                    
                    tree.insert("", tk.END, values=row_values, tags=("header",))
                
                # 处理数据行
                for row_idx in range(3, min(50, max_row)):  # 只显示前50行
                    row_values = []
                    for col_idx in range(max_col):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        row_values.append(str(cell_value) if cell_value else "")
                    
                    tree.insert("", tk.END, values=row_values)
                
                # 显示树状图
                tree.pack(fill=tk.BOTH, expand=True)
                
                # 添加样式
                tree.tag_configure("header", background="lightblue")
                
                if max_row > 50:
                    # 添加提示标签
                    info_label = ttk.Label(preview_window, text=f"... 还有 {max_row - 50} 行未显示")
                    info_label.pack(pady=5)
            else:
                messagebox.showwarning("警告", "不支持的Excel文件格式")
                return
        
        except Exception as e:
            messagebox.showerror("错误", f"预览Excel文件失败: {str(e)}")

    def select_batch_extract_class_header(self):
        if not self.batch_extract_excel_files:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return
        
        # 读取第一个Excel文件的表头
        try:
            headers = self.read_excel_headers_with_merged_cells(self.batch_extract_excel_files[0])
            
            if not headers:
                messagebox.showwarning("警告", "未能读取到有效的表头")
                return
            
            # 创建表头选择窗口
            header_window = tk.Toplevel(self.root)
            header_window.title("选择班级表头")
            header_window.geometry("400x300")
            
            # 创建列表框
            listbox = tk.Listbox(header_window, selectmode=tk.SINGLE, width=50)
            listbox.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
            
            # 填充表头
            for header in headers:
                listbox.insert(tk.END, header)
            
            # 确定按钮
            def on_select():
                if listbox.curselection():
                    selected = listbox.get(listbox.curselection())
                    self.batch_extract_class_header = selected
                    self.batch_extract_class_header_var.set(selected)
                    header_window.destroy()
            
            ttk.Button(header_window, text="确定", command=on_select).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件时出错：{str(e)}")

    def select_batch_append_excel_files(self):
        file_paths = filedialog.askopenfilenames(
            title="选择Excel文件（最多5个）",
            filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")]
        )
        if file_paths:
            # 限制选择5个文件
            if len(file_paths) > 5:
                messagebox.showwarning("警告", "最多只能选择5个Excel文件")
                file_paths = file_paths[:5]
            
            self.batch_append_excel_files = file_paths
            self.batch_append_excel_var.set(f"已选择 {len(file_paths)} 个文件")
            self.update_button_state()

    def select_batch_append_name_header(self):
        if not self.batch_append_excel_files:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return
        
        # 读取第一个Excel文件的表头
        try:
            headers = self.read_excel_headers_with_merged_cells(self.batch_append_excel_files[0])
            
            if not headers:
                messagebox.showwarning("警告", "未能读取到有效的表头")
                return
            
            # 创建表头选择窗口
            header_window = tk.Toplevel(self.root)
            header_window.title("选择姓名表头")
            header_window.geometry("400x300")
            
            # 创建列表框
            listbox = tk.Listbox(header_window, selectmode=tk.SINGLE, width=50)
            listbox.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
            
            # 填充表头
            for header in headers:
                listbox.insert(tk.END, header)
            
            # 确定按钮
            def on_select():
                if listbox.curselection():
                    selected = listbox.get(listbox.curselection())
                    self.batch_append_name_header = selected
                    self.batch_append_name_header_var.set(selected)
                    header_window.destroy()
            
            ttk.Button(header_window, text="确定", command=on_select).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件时出错：{str(e)}")

    def select_batch_append_class_header(self):
        if not self.batch_append_excel_files:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return
        
        # 读取第一个Excel文件的表头
        try:
            headers = self.read_excel_headers_with_merged_cells(self.batch_append_excel_files[0])
            
            if not headers:
                messagebox.showwarning("警告", "未能读取到有效的表头")
                return
            
            # 创建表头选择窗口
            header_window = tk.Toplevel(self.root)
            header_window.title("选择班级表头")
            header_window.geometry("400x300")
            
            # 创建列表框
            listbox = tk.Listbox(header_window, selectmode=tk.SINGLE, width=50)
            listbox.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
            
            # 填充表头
            for header in headers:
                listbox.insert(tk.END, header)
            
            # 确定按钮
            def on_select():
                if listbox.curselection():
                    selected = listbox.get(listbox.curselection())
                    self.batch_append_class_header = selected
                    self.batch_append_class_header_var.set(selected)
                    header_window.destroy()
            
            ttk.Button(header_window, text="确定", command=on_select).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件时出错：{str(e)}")

    def preview_batch_extract(self):
        """预览批量提取功能"""
        if not self.batch_extract_pdfs:
            messagebox.showwarning("警告", "请先选择PDF文件")
            return
        
        if not self.batch_extract_output_dir:
            messagebox.showwarning("警告", "请先选择输出目录")
            return
        
        try:
            # 如果还没有提取姓名，先自动提取
            if not self.batch_extract_extracted_names:
                self.auto_extract_names_batch_extract()
            
            # 读取Excel数据（如果有）
            use_excel = bool(self.batch_extract_excel_files and self.batch_extract_name_header and self.batch_extract_class_header)
            data_index = {} if not use_excel else self.read_excel_data(
                self.batch_extract_excel_files, 
                self.batch_extract_name_header, 
                self.batch_extract_class_header
            )
            
            # 计算匹配率
            total_files = len(self.batch_extract_pdfs)
            matched_files = 0
            preview_data = []
            
            # 先计算所有文件的匹配率
            if use_excel:
                for pdf_path in self.batch_extract_pdfs:
                    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                    # 获取提取的姓名
                    extracted_name = self.batch_extract_extracted_names.get(pdf_path, base_name)
                    # 使用提取的姓名进行匹配
                    matched_name, matched_class = self.match_filename_with_excel(extracted_name, data_index)
                    if matched_name:
                        matched_files += 1
            
            # 预览前5个文件
            for pdf_path in self.batch_extract_pdfs[:5]:  # 只预览前5个文件
                base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                matched_name, matched_class = None, None
                
                # 获取提取的姓名（无论是否使用Excel）
                extracted_name = self.batch_extract_extracted_names.get(pdf_path, base_name)
                
                if use_excel:
                    # 优先使用提取的姓名进行匹配
                    matched_name, matched_class = self.match_filename_with_excel(extracted_name, data_index)
                
                # 生成输出文件名
                if use_excel and matched_name and matched_class:
                    output_filename = f"{matched_name}_{matched_class}.pdf"
                else:
                    output_filename = f"{base_name}-{self.batch_extract_page}页.pdf"
                
                preview_data.append((pdf_path, base_name, extracted_name, matched_name, matched_class, output_filename))
            
            # 计算匹配率
            match_rate = (matched_files / total_files * 100) if total_files > 0 else 0
            
            # 创建预览窗口
            preview_window = tk.Toplevel(self.root)
            preview_window.title("批量提取预览")
            preview_window.geometry("800x600")
            
            # 创建滚动条
            scrollbar = ttk.Scrollbar(preview_window)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 创建文本框
            text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('微软雅黑', 10))
            text.pack(fill=tk.BOTH, expand=True)
            
            # 配置滚动条
            scrollbar.config(command=text.yview)
            
            # 显示标题
            text.insert(tk.END, "批量提取预览\n")
            text.insert(tk.END, "=" * 80 + "\n")
            text.insert(tk.END, f"PDF文件数量：{total_files}\n")
            text.insert(tk.END, f"提取页码：{self.batch_extract_page}\n")
            text.insert(tk.END, f"输出目录：{self.batch_extract_output_dir}\n")
            text.insert(tk.END, f"Excel匹配率：{match_rate:.1f}%\n")
            text.insert(tk.END, "\n")
            
            # 显示预览数据
            text.insert(tk.END, "预览文件列表：\n")
            text.insert(tk.END, "-" * 80 + "\n")
            
            for pdf_path, base_name, extracted_name, matched_name, matched_class, output_filename in preview_data:
                text.insert(tk.END, f"原文件名：{os.path.basename(pdf_path)}\n")
                text.insert(tk.END, f"基础名称：{base_name}\n")
                text.insert(tk.END, f"提取姓名：{extracted_name}\n")
                if matched_name:
                    text.insert(tk.END, f"匹配姓名：{matched_name}\n")
                    text.insert(tk.END, f"匹配班级：{matched_class}\n")
                else:
                    text.insert(tk.END, "匹配姓名：未匹配\n")
                    text.insert(tk.END, "匹配班级：未匹配\n")
                text.insert(tk.END, f"输出文件名：{output_filename}\n")
                text.insert(tk.END, "-" * 80 + "\n")
            
            if total_files > 5:
                text.insert(tk.END, f"... 还有 {total_files - 5} 个文件未显示\n")
            
            # 禁用文本编辑
            text.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("错误", f"预览失败: {str(e)}")

    def extract_name_from_filename(self, filename, special_char=None):
        """从文件名中提取姓名"""
        import re
        
        # 移除文件扩展名
        base_name = os.path.splitext(filename)[0]
        
        if special_char:
            # 使用特殊符号提取
            if special_char in base_name:
                return base_name.split(special_char)[0].strip()
        else:
            # 自动提取：提取所有汉字部分
            chinese_chars = re.findall(r'[\u4e00-\u9fa5]+', base_name)
            if chinese_chars:
                return ''.join(chinese_chars)
        
        # 如果无法提取，返回原文件名
        return base_name

    def auto_extract_names_batch_extract(self):
        """自动提取批量提取功能中的姓名"""
        if not self.batch_extract_pdfs:
            messagebox.showwarning("警告", "请先选择PDF文件")
            return
        
        # 提取姓名
        self.batch_extract_extracted_names = {}
        for pdf_path in self.batch_extract_pdfs:
            filename = os.path.basename(pdf_path)
            extracted_name = self.extract_name_from_filename(filename)
            self.batch_extract_extracted_names[pdf_path] = extracted_name
        
        messagebox.showinfo("成功", f"已自动提取 {len(self.batch_extract_extracted_names)} 个文件的姓名")

    def custom_extract_names_batch_extract(self):
        """使用自定义特殊符号提取批量提取功能中的姓名"""
        if not self.batch_extract_pdfs:
            messagebox.showwarning("警告", "请先选择PDF文件")
            return
        
        # 获取特殊符号
        special_char = self.batch_extract_special_char_var.get()
        if not special_char:
            messagebox.showwarning("警告", "请输入特殊符号")
            return
        
        # 提取姓名
        self.batch_extract_extracted_names = {}
        for pdf_path in self.batch_extract_pdfs:
            filename = os.path.basename(pdf_path)
            extracted_name = self.extract_name_from_filename(filename, special_char)
            self.batch_extract_extracted_names[pdf_path] = extracted_name
        
        messagebox.showinfo("成功", f"已使用特殊符号 '{special_char}' 提取 {len(self.batch_extract_extracted_names)} 个文件的姓名")

    def select_batch_append_pdfs(self):
        file_paths = filedialog.askopenfilenames(
            title="选择PDF文件（批量追加）",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_paths:
            valid_pdfs = []
            for path in file_paths:
                if self.is_valid_pdf(path):
                    valid_pdfs.append(path)
                else:
                    messagebox.showerror("错误", f"文件 {os.path.basename(path)} 不是有效的PDF文件")
            
            if valid_pdfs:
                self.batch_append_pdfs = valid_pdfs
                self.batch_append_pdf_var.set(f"已选择 {len(valid_pdfs)} 个文件")
                self.update_button_state()

    def select_batch_append_output_dir(self):
        dir_path = filedialog.askdirectory(
            title="选择输出目录"
        )
        if dir_path:
            self.batch_append_output_dir = dir_path
            self.batch_append_output_var.set(dir_path)
            self.update_button_state()

    def select_batch_extract_output_dir(self):
        dir_path = filedialog.askdirectory(
            title="选择输出目录"
        )
        if dir_path:
            self.batch_extract_output_dir = dir_path
            self.batch_extract_output_var.set(dir_path)
            self.update_button_state()

    def select_batch_extract_pdfs(self):
        file_paths = filedialog.askopenfilenames(
            title="选择PDF文件（批量提取）",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_paths:
            valid_pdfs = []
            for path in file_paths:
                if self.is_valid_pdf(path):
                    valid_pdfs.append(path)
                else:
                    messagebox.showerror("错误", f"文件 {os.path.basename(path)} 不是有效的PDF文件")
            
            if valid_pdfs:
                self.batch_extract_pdfs = valid_pdfs
                self.batch_extract_pdf_var.set(f"已选择 {len(valid_pdfs)} 个文件")
                self.update_button_state()

    def auto_extract_names_batch_append(self):
        """自动提取批量追加页功能中的姓名"""
        if not self.batch_append_pdfs:
            messagebox.showwarning("警告", "请先选择PDF文件")
            return
        
        # 提取姓名
        self.batch_append_extracted_names = {}
        for pdf_path in self.batch_append_pdfs:
            filename = os.path.basename(pdf_path)
            extracted_name = self.extract_name_from_filename(filename)
            self.batch_append_extracted_names[pdf_path] = extracted_name
        
        messagebox.showinfo("成功", f"已自动提取 {len(self.batch_append_extracted_names)} 个文件的姓名")

    def custom_extract_names_batch_append(self):
        """使用自定义特殊符号提取批量追加页功能中的姓名"""
        if not self.batch_append_pdfs:
            messagebox.showwarning("警告", "请先选择PDF文件")
            return
        
        special_char = self.batch_append_special_char_var.get()
        if not special_char:
            messagebox.showwarning("警告", "请输入特殊符号")
            return
        
        # 提取姓名
        self.batch_append_extracted_names = {}
        for pdf_path in self.batch_append_pdfs:
            filename = os.path.basename(pdf_path)
            extracted_name = self.extract_name_from_filename(filename, special_char)
            self.batch_append_extracted_names[pdf_path] = extracted_name
        
        messagebox.showinfo("成功", f"已使用特殊符号 '{special_char}' 提取 {len(self.batch_append_extracted_names)} 个文件的姓名")

    def preview_name_extraction_batch_append(self):
        """预览批量追加页功能中的姓名提取结果"""
        if not self.batch_append_pdfs:
            messagebox.showwarning("警告", "请先选择PDF文件")
            return
        
        # 如果还没有提取姓名，先自动提取
        if not self.batch_append_extracted_names:
            self.auto_extract_names_batch_append()
        
        # 创建预览窗口
        preview_window = tk.Toplevel(self.root)
        preview_window.title("姓名提取预览")
        preview_window.geometry("600x400")
        
        # 创建滚动条
        scrollbar = ttk.Scrollbar(preview_window)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建文本框
        text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('微软雅黑', 10))
        text.pack(fill=tk.BOTH, expand=True)
        
        # 配置滚动条
        scrollbar.config(command=text.yview)
        
        # 显示标题
        text.insert(tk.END, "姓名提取预览\n")
        text.insert(tk.END, "=" * 60 + "\n")
        
        # 显示预览数据
        for pdf_path, extracted_name in self.batch_append_extracted_names.items():
            text.insert(tk.END, f"原文件名：{os.path.basename(pdf_path)}\n")
            text.insert(tk.END, f"提取姓名：{extracted_name}\n")
            text.insert(tk.END, "-" * 60 + "\n")
        
        # 禁用文本编辑
        text.config(state=tk.DISABLED)

    def preview_batch_append(self):
        """预览批量追加页功能"""

        if not self.batch_append_source_pdf:
            messagebox.showwarning("警告", "请先选择源PDF文件")
            return
        
        if not self.batch_append_pdfs:
            messagebox.showwarning("警告", "请先选择要追加的PDF文件")
            return
        
        if not self.batch_append_output_dir:
            messagebox.showwarning("警告", "请先选择输出目录")
            return
        
        try:
            # 如果还没有提取姓名，先自动提取
            if not self.batch_append_extracted_names:
                self.auto_extract_names_batch_append()
            
            # 读取Excel数据（如果有）
            use_excel = bool(self.batch_append_excel_files and self.batch_append_name_header and self.batch_append_class_header)
            data_index = {} if not use_excel else self.read_excel_data(
                self.batch_append_excel_files, 
                self.batch_append_name_header, 
                self.batch_append_class_header
            )
            
            # 计算匹配率
            total_files = len(self.batch_append_pdfs)
            matched_files = 0
            preview_data = []
            
            # 先计算所有文件的匹配率
            if use_excel:
                for pdf_path in self.batch_append_pdfs:
                    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                    # 获取提取的姓名
                    extracted_name = self.batch_append_extracted_names.get(pdf_path, base_name)
                    # 使用提取的姓名进行匹配
                    matched_name, matched_class = self.match_filename_with_excel(extracted_name, data_index)
                    if matched_name:
                        matched_files += 1
            
            # 预览前5个文件
            for pdf_path in self.batch_append_pdfs[:5]:  # 只预览前5个文件
                base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                matched_name, matched_class = None, None
                
                # 获取提取的姓名（无论是否使用Excel）
                extracted_name = self.batch_append_extracted_names.get(pdf_path, base_name)
                
                if use_excel:
                    # 使用提取的姓名进行匹配
                    matched_name, matched_class = self.match_filename_with_excel(extracted_name, data_index)
                
                # 生成输出文件名
                if use_excel and matched_name and matched_class:
                    output_filename = f"{matched_name}_{matched_class}.pdf"
                else:
                    output_filename = f"{base_name}-追加页.pdf"
                
                preview_data.append((pdf_path, base_name, extracted_name, matched_name, matched_class, output_filename))
            
            # 计算匹配率
            match_rate = (matched_files / total_files * 100) if total_files > 0 else 0
            
            # 创建预览窗口
            preview_window = tk.Toplevel(self.root)
            preview_window.title("批量追加页预览")
            preview_window.geometry("800x600")
            
            # 创建滚动条
            scrollbar = ttk.Scrollbar(preview_window)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 创建文本框
            text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('微软雅黑', 10))
            text.pack(fill=tk.BOTH, expand=True)
            
            # 配置滚动条
            scrollbar.config(command=text.yview)
            
            # 显示标题
            text.insert(tk.END, "批量追加页预览\n")
            text.insert(tk.END, "=" * 80 + "\n")
            text.insert(tk.END, f"源PDF文件：{os.path.basename(self.batch_append_source_pdf)}\n")
            text.insert(tk.END, f"目标PDF数量：{total_files}\n")
            text.insert(tk.END, f"输出目录：{self.batch_append_output_dir}\n")
            text.insert(tk.END, f"Excel匹配率：{match_rate:.1f}%\n")
            text.insert(tk.END, "\n")
            
            # 显示预览数据
            text.insert(tk.END, "预览文件列表：\n")
            text.insert(tk.END, "-" * 80 + "\n")
            
            for pdf_path, base_name, extracted_name, matched_name, matched_class, output_filename in preview_data:
                text.insert(tk.END, f"原文件名：{os.path.basename(pdf_path)}\n")
                text.insert(tk.END, f"基础名称：{base_name}\n")
                text.insert(tk.END, f"提取姓名：{extracted_name}\n")
                if matched_name:
                    text.insert(tk.END, f"匹配姓名：{matched_name}\n")
                    text.insert(tk.END, f"匹配班级：{matched_class}\n")
                else:
                    text.insert(tk.END, "匹配姓名：未匹配\n")
                    text.insert(tk.END, "匹配班级：未匹配\n")
                text.insert(tk.END, f"输出文件名：{output_filename}\n")
                text.insert(tk.END, "-" * 80 + "\n")
            
            if total_files > 5:
                text.insert(tk.END, f"... 还有 {total_files - 5} 个文件未显示\n")
            
            # 禁用文本编辑
            text.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("错误", f"预览失败: {str(e)}")

    def select_student_pdfs(self):
        # 选择学生PDF文件
        file_paths = filedialog.askopenfilenames(
            title="选择学生PDF文件（批量）",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_paths:
            self.student_pdfs = list(file_paths)
            self.student_pdf_var.set(f"已选择 {len(file_paths)} 个学生PDF文件")

    def select_class_pdfs(self):
        # 选择班级PDF文件
        file_paths = filedialog.askopenfilenames(
            title="选择班级PDF文件（批量）",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_paths:
            self.class_pdfs = list(file_paths)
            self.class_pdf_var.set(f"已选择 {len(file_paths)} 个班级PDF文件")

    def select_student_class_excel(self):
        # 选择学生班级对应关系Excel文件
        file_path = filedialog.askopenfilename(
            title="选择学生班级对应关系Excel文件",
            filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.student_class_excel = file_path
            self.student_class_excel_var.set(file_path)

    def select_student_class_name_header(self):
        # 选择姓名表头
        if not self.student_class_excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        try:
            headers = self.read_excel_headers_with_merged_cells(self.student_class_excel_file)
            if not headers:
                messagebox.showerror("错误", "无法读取Excel表头")
                return
            
            dialog = tk.Toplevel(self.root)
            dialog.title("选择姓名表头")
            dialog.geometry("300x200")
            dialog.transient(self.root)
            dialog.grab_set()
            
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(frame, text="选择姓名表头:").pack(anchor=tk.W, pady=5)
            header_var = tk.StringVar()
            header_combo = ttk.Combobox(frame, textvariable=header_var, values=headers, width=30)
            header_combo.pack(anchor=tk.W, pady=5)
            
            btn_frame = ttk.Frame(frame)
            btn_frame.pack(fill=tk.X, pady=10)
            
            def select_header():
                header = header_var.get()
                if header:
                    self.student_class_name_header = header
                    self.student_class_name_header_var.set(header)
                    dialog.destroy()
                else:
                    messagebox.showwarning("警告", "请选择表头")
            
            ttk.Button(btn_frame, text="确定", command=select_header).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel表头时出错：{str(e)}")

    def select_student_class_class_header(self):
        # 选择班级表头
        if not self.student_class_excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        try:
            headers = self.read_excel_headers_with_merged_cells(self.student_class_excel_file)
            if not headers:
                messagebox.showerror("错误", "无法读取Excel表头")
                return
            
            dialog = tk.Toplevel(self.root)
            dialog.title("选择班级表头")
            dialog.geometry("300x200")
            dialog.transient(self.root)
            dialog.grab_set()
            
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(frame, text="选择班级表头:").pack(anchor=tk.W, pady=5)
            header_var = tk.StringVar()
            header_combo = ttk.Combobox(frame, textvariable=header_var, values=headers, width=30)
            header_combo.pack(anchor=tk.W, pady=5)
            
            btn_frame = ttk.Frame(frame)
            btn_frame.pack(fill=tk.X, pady=10)
            
            def select_header():
                header = header_var.get()
                if header:
                    self.student_class_class_header = header
                    self.student_class_class_header_var.set(header)
                    dialog.destroy()
                else:
                    messagebox.showwarning("警告", "请选择表头")
            
            ttk.Button(btn_frame, text="确定", command=select_header).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
            
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel表头时出错：{str(e)}")

    def select_student_class_output_dir(self):
        # 选择输出目录
        dir_path = filedialog.askdirectory(title="选择输出目录")
        if dir_path:
            self.student_class_output_dir = dir_path
            self.student_class_output_var.set(dir_path)

    def preview_student_class_merge(self):
        # 预览学生班级PDF合并结果
        if not self.student_pdfs:
            messagebox.showerror("错误", "请先选择学生PDF文件")
            return
        
        if not self.class_pdfs:
            messagebox.showerror("错误", "请先选择班级PDF文件")
            return
        
        if not self.student_class_excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        if not self.student_class_name_header:
            messagebox.showerror("错误", "请先选择姓名表头")
            return
        
        if not self.student_class_class_header:
            messagebox.showerror("错误", "请先选择班级表头")
            return
        
        try:
            # 读取Excel数据
            excel_data = self.read_excel_data([self.student_class_excel_file], self.student_class_name_header, self.student_class_class_header)
            
            # 创建预览窗口
            preview_window = tk.Toplevel(self.root)
            preview_window.title("学生班级PDF合并预览")
            preview_window.geometry("800x600")
            
            # 创建滚动条
            scrollbar = ttk.Scrollbar(preview_window)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 创建文本框
            text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('Courier New', 10))
            text.pack(fill=tk.BOTH, expand=True)
            
            # 配置滚动条
            scrollbar.config(command=text.yview)
            
            # 显示标题
            text.insert(tk.END, "学生班级PDF合并预览\n")
            text.insert(tk.END, "=" * 60 + "\n")
            text.insert(tk.END, f"学生PDF文件数：{len(self.student_pdfs)}\n")
            text.insert(tk.END, f"班级PDF文件数：{len(self.class_pdfs)}\n")
            text.insert(tk.END, f"Excel数据条数：{len(excel_data)}\n\n")
            
            # 显示每个学生的合并预览
            for student_pdf in self.student_pdfs:
                # 提取学生姓名（处理"学生_档案编号.pdf"格式）
                student_filename = os.path.basename(student_pdf)
                student_name = os.path.splitext(student_filename)[0]
                # 如果文件名包含下划线，只取前面部分作为姓名
                if '_' in student_name:
                    student_name = student_name.split('_')[0]
                text.insert(tk.END, f"学生PDF：{student_filename}\n")
                
                # 查找对应的班级
                matched_class = None
                for name, class_name in excel_data.items():
                    if name == student_name:
                        matched_class = class_name
                        break
                
                if matched_class:
                    # 查找对应的班级PDF
                    class_pdf = None
                    for pdf in self.class_pdfs:
                        pdf_name = os.path.splitext(os.path.basename(pdf))[0]
                        if pdf_name == matched_class:
                            class_pdf = pdf
                            break
                    
                    if class_pdf:
                        text.insert(tk.END, f"  班级：{matched_class}\n")
                        text.insert(tk.END, f"  班级PDF：{os.path.basename(class_pdf)}\n")
                        text.insert(tk.END, f"  合并结果：{os.path.splitext(student_filename)[0]}_合并.pdf\n")
                    else:
                        text.insert(tk.END, f"  班级：{matched_class}\n")
                        text.insert(tk.END, f"  班级PDF：未找到\n")
                else:
                    text.insert(tk.END, f"  班级：未找到匹配\n")
                
                text.insert(tk.END, "\n")
            
            # 禁用文本编辑
            text.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("错误", f"预览失败: {str(e)}")

    def process_student_class_merge(self):
        # 执行学生班级PDF合并
        if not self.student_pdfs:
            messagebox.showerror("错误", "请先选择学生PDF文件")
            return
        
        if not self.class_pdfs:
            messagebox.showerror("错误", "请先选择班级PDF文件")
            return
        
        if not self.student_class_excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        if not self.student_class_name_header:
            messagebox.showerror("错误", "请先选择姓名表头")
            return
        
        if not self.student_class_class_header:
            messagebox.showerror("错误", "请先选择班级表头")
            return
        
        if not self.student_class_output_dir:
            messagebox.showerror("错误", "请先选择输出目录")
            return
        
        try:
            # 读取Excel数据
            excel_data = self.read_excel_data([self.student_class_excel_file], self.student_class_name_header, self.student_class_class_header)
            
            total_files = len(self.student_pdfs)
            processed_count = 0
            success_count = 0
            fail_count = 0
            fail_messages = []
            
            for student_pdf in self.student_pdfs:
                try:
                    # 提取学生姓名（处理"学生_档案编号.pdf"格式）
                    student_filename = os.path.basename(student_pdf)
                    student_name = os.path.splitext(student_filename)[0]
                    if '_' in student_name:
                        student_name = student_name.split('_')[0]
                    
                    # 查找对应的班级
                    matched_class = None
                    for name, class_name in excel_data.items():
                        if name == student_name:
                            matched_class = class_name
                            break
                    
                    if matched_class:
                        # 查找对应的班级PDF
                        class_pdf = None
                        for pdf in self.class_pdfs:
                            pdf_name = os.path.splitext(os.path.basename(pdf))[0]
                            if pdf_name == matched_class:
                                class_pdf = pdf
                                break
                        
                        if class_pdf:
                            # 合并PDF
                            output_filename = f"{os.path.splitext(student_filename)[0]}_合并.pdf"
                            output_path = os.path.join(self.student_class_output_dir, output_filename)
                            
                            # 使用PyPDF2合并
                            merger = PyPDF2.PdfMerger()
                            merger.append(student_pdf)
                            merger.append(class_pdf)
                            merger.write(output_path)
                            merger.close()
                            
                            success_count += 1
                        else:
                            fail_msg = f"{student_filename}: 未找到班级PDF"
                            fail_messages.append(fail_msg)
                            fail_count += 1
                    else:
                        fail_msg = f"{student_filename}: 未找到匹配的班级"
                        fail_messages.append(fail_msg)
                        fail_count += 1
                    
                    processed_count += 1
                    progress = processed_count / total_files * 100
                    self.status_var.set(f"处理中... {progress:.1f}%")
                    self.root.update()
                    
                except Exception as e:
                    processed_count += 1
                    fail_count += 1
                    fail_msg = f"{os.path.basename(student_pdf)}: {str(e)}"
                    fail_messages.append(fail_msg)
            
            # 显示结果
            result_msg = f"学生班级PDF合并完成！\n\n处理文件数：{total_files}\n成功合并：{success_count}"
            if fail_count > 0:
                result_msg += f"\n失败：{fail_count}"
                result_msg += "\n\n失败详情：\n" + "\n".join(fail_messages[:5])
                if len(fail_messages) > 5:
                    result_msg += f"\n... 还有 {len(fail_messages) - 5} 个失败文件"
            
            if fail_count > 0:
                messagebox.showwarning("处理完成", result_msg)
            else:
                messagebox.showinfo("成功", result_msg)
            
        except Exception as e:
            messagebox.showerror("错误", f"处理失败: {str(e)}")

    def select_excel_files(self):
        # 选择Excel文件（批量）
        file_paths = filedialog.askopenfilenames(
            title="选择Excel文件（批量）",
            filetypes=[("Excel文件", "*.xlsx"), ("Excel文件", "*.xls"), ("所有文件", "*.*")]
        )
        if file_paths:
            self.excel_files = list(file_paths)
            self.excel_files_var.set(f"已选择 {len(file_paths)} 个Excel文件")
            self.update_button_state()

    def select_excel_output_dir(self):
        # 选择Excel导出PDF的输出目录
        dir_path = filedialog.askdirectory(title="选择输出目录")
        if dir_path:
            self.excel_output_dir = dir_path
            self.excel_output_var.set(dir_path)
            self.update_button_state()

    def select_word_files(self):
        # 选择Word文件（批量）
        file_paths = filedialog.askopenfilenames(
            title="选择Word文件（批量）",
            filetypes=[("Word文件", "*.docx"), ("Word文件", "*.doc"), ("所有文件", "*.*")]
        )
        if file_paths:
            self.word_files = list(file_paths)
            self.word_files_var.set(f"已选择 {len(file_paths)} 个Word文件")
            self.update_button_state()

    def select_word_output_dir(self):
        # 选择Word导出PDF的输出目录
        dir_path = filedialog.askdirectory(title="选择输出目录")
        if dir_path:
            self.word_output_dir = dir_path
            self.word_output_var.set(dir_path)
            self.update_button_state()

    def process_pdf(self):
        try:
            self.status_var.set("处理中...")
            self.root.update()
            
            if self.function_var.get() == "insert":
                # PDF插入功能
                self.insert_interval = int(self.interval_var.get())
                if self.insert_interval < 1:
                    messagebox.showerror("错误", "插入间隔必须大于0")
                    return
                
                self.insert_pages = int(self.insert_pages_var.get())
                if self.insert_pages < 1:
                    messagebox.showerror("错误", "每次插入页数必须大于0")
                    return
                
                if self.mode_var.get() == "single":
                    self.process_single_insert()
                elif self.mode_var.get() == "batch":
                    self.process_batch_insert()
                else:
                    # 多PDF顺序插入
                    self.multi_count = int(self.multi_count_var.get())
                    if self.multi_count < 1:
                        messagebox.showerror("错误", "每次插入PDF数量必须大于0")
                        return
                    self.process_multi_insert()
            elif self.function_var.get() == "split":
                # PDF分割功能
                self.split_pages = int(self.split_pages_var.get())
                if self.split_pages < 1:
                    messagebox.showerror("错误", "分割页数必须大于0")
                    return
                
                if not self.split_pdf:
                    messagebox.showerror("错误", "请选择要分割的PDF文件")
                    return
                
                if self.naming_var.get() == "excel" and not self.excel_file:
                    messagebox.showerror("错误", "请选择Excel文件")
                    return
                
                self.process_split_pdf()
            elif self.function_var.get() == "swap":
                # PDF页面交换功能
                self.page_a = int(self.page_a_var.get())
                if self.page_a < 1:
                    messagebox.showerror("错误", "页面A必须大于0")
                    return
                
                self.page_b = int(self.page_b_var.get())
                if self.page_b < 1:
                    messagebox.showerror("错误", "页面B必须大于0")
                    return
                
                if not self.swap_pdfs:
                    messagebox.showerror("错误", "请选择要交换页面的PDF文件")
                    return
                
                self.process_swap_pdf()
            elif self.function_var.get() == "export":
                # 指定导出页功能
                self.export_pages = self.export_pages_var.get()
                if not self.export_pages:
                    messagebox.showerror("错误", "请输入导出页面")
                    return
                self.process_export()
                # 不显示通用成功消息，因为process_export已经显示了详细的成功消息
                self.status_var.set("处理完成")
            elif self.function_var.get() == "reorder":
                # PDF重排序功能
                if not self.reorder_pdfs:
                    messagebox.showerror("错误", "请选择要重排序的PDF文件")
                    return
                
                # Excel文件是可选的，只有需要按Excel数据重命名时才需要
                use_excel = self.reorder_excel_file and self.reorder_name_header and self.reorder_id_header
                
                # 执行重排序
                output_dir = filedialog.askdirectory(title="选择保存目录")
                if not output_dir:
                    return
                
                try:
                    total_files = len(self.reorder_pdfs)
                    processed_count = 0
                    success_count = 0
                    fail_count = 0
                    fail_messages = []
                    
                    # 获取Excel数据（如果使用）
                    name_to_id = {}
                    if use_excel:
                        name_id_mapping = self.get_export_mapping()
                        name_to_id = {name: id_value for name, id_value in name_id_mapping}
                    
                    # 处理PDF文件
                    pdf_info = []
                    for pdf_path in self.reorder_pdfs:
                        original_name = os.path.basename(pdf_path)
                        # 从文件名中提取姓名（假设文件名格式为"XXX_02.pdf"或"XXX-27.pdf"）
                        name_match = re.match(r'^([^_\-]+)[_\-].*\.pdf$', original_name)
                        if name_match:
                            extracted_name = name_match.group(1)
                        else:
                            extracted_name = original_name.split('.')[0]
                        
                        # 获取Excel中的序号（如果使用Excel）
                        if use_excel and extracted_name in name_to_id:
                            id_value = name_to_id[extracted_name]
                        else:
                            id_value = None
                        
                        pdf_info.append((extracted_name, id_value, original_name, pdf_path))
                    
                    # 按序号排序（如果使用Excel），否则按原顺序
                    if use_excel:
                        def get_sort_key(item):
                            id_value = item[1]
                            if id_value is None:
                                return (1, item[0])  # 未找到的放在最后
                            try:
                                return (0, int(id_value))
                            except (ValueError, TypeError):
                                return (0, id_value)
                        
                        pdf_info.sort(key=get_sort_key)
                    
                    # 保存排序后的PDF
                    for i, (name, id_value, original_name, pdf_path) in enumerate(pdf_info, 1):
                        try:
                            # 生成新文件名
                            if use_excel and id_value is not None:
                                try:
                                    id_num = int(id_value)
                                    new_name = f"{name}_{id_num:02d}.pdf"
                                except (ValueError, TypeError):
                                    new_name = f"{name}_{id_value}.pdf"
                            else:
                                # 不使用Excel时，按顺序编号
                                new_name = f"{name}_{i:02d}.pdf"
                            
                            # 复制文件到输出目录
                            output_path = os.path.join(output_dir, new_name)
                            shutil.copy2(pdf_path, output_path)
                            success_count += 1
                            
                        except Exception as e:
                            fail_msg = f"{original_name}: {str(e)}"
                            fail_messages.append(fail_msg)
                            fail_count += 1
                        
                        processed_count += 1
                        progress = processed_count / total_files * 100
                        self.status_var.set(f"处理中... {progress:.1f}%")
                        self.root.update()
                    
                    # 显示结果
                    result_msg = f"PDF重排序完成！\n\n处理文件数：{total_files}\n成功：{success_count}"
                    if fail_count > 0:
                        result_msg += f"\n失败：{fail_count}"
                        result_msg += "\n\n失败详情：\n" + "\n".join(fail_messages[:5])
                        if len(fail_messages) > 5:
                            result_msg += f"\n... 还有 {len(fail_messages) - 5} 个失败项"
                    
                    if fail_count > 0:
                        messagebox.showwarning("处理完成", result_msg)
                    else:
                        messagebox.showinfo("成功", result_msg)
                    
                except Exception as e:
                    messagebox.showerror("错误", f"处理失败: {str(e)}")
                    
                self.status_var.set("处理完成")
            elif self.function_var.get() == "batch_extract":
                # 批量提取功能
                if not self.batch_extract_pdfs:
                    messagebox.showerror("错误", "请选择要提取的PDF文件")
                    return
                
                if not self.batch_extract_output_dir:
                    messagebox.showerror("错误", "请选择输出目录")
                    return
                
                try:
                    self.batch_extract_page = int(self.batch_extract_page_var.get())
                    if self.batch_extract_page < 1:
                        messagebox.showerror("错误", "提取页码必须大于0")
                        return
                except ValueError:
                    messagebox.showerror("错误", "请输入有效的提取页码")
                    return
                
                self.process_batch_extract()
                self.status_var.set("处理完成")
                messagebox.showinfo("成功", "PDF批量提取操作完成！")
            elif self.function_var.get() == "batch_append":
                # 批量追加页功能
                if not self.batch_append_source_pdf:
                    messagebox.showerror("错误", "请选择源PDF文件")
                    return
                
                if not self.batch_append_pdfs:
                    messagebox.showerror("错误", "请选择要追加的PDF文件")
                    return
                
                if not self.batch_append_output_dir:
                    messagebox.showerror("错误", "请选择输出目录")
                    return
                
                self.process_batch_append()
                self.status_var.set("处理完成")
                messagebox.showinfo("成功", "PDF批量追加页操作完成！")
            elif self.function_var.get() == "student_class":
                # 学生班级PDF合并功能
                self.process_student_class_merge()
                self.status_var.set("处理完成")
                messagebox.showinfo("成功", "学生班级PDF合并操作完成！")
            elif self.function_var.get() == "excel_to_pdf":
                # Excel批量导出PDF功能
                if not self.excel_files:
                    messagebox.showerror("错误", "请选择Excel文件")
                    return
                if not self.excel_output_dir:
                    messagebox.showerror("错误", "请选择输出目录")
                    return
                self.process_excel_to_pdf()
            elif self.function_var.get() == "word_to_pdf":
                # Word批量导出PDF功能
                if not self.word_files:
                    messagebox.showerror("错误", "请选择Word文件")
                    return
                if not self.word_output_dir:
                    messagebox.showerror("错误", "请选择输出目录")
                    return
                self.process_word_to_pdf()
            else:
                # PDF批量重命名功能
                if not self.rename_pdfs:
                    messagebox.showerror("错误", "请选择要重命名的PDF文件")
                    return
                
                # 根据选择的重命名模式处理
                if self.rename_mode_var.get() == "excel":
                    # 传统Excel匹配重命名
                    if not self.rename_excel_file:
                        messagebox.showerror("错误", "请选择Excel文件")
                        return
                    
                    if not self.rename_name_header:
                        messagebox.showerror("错误", "请选择姓名表头")
                        return
                    
                    if not self.rename_id_header:
                        messagebox.showerror("错误", "请选择序号表头")
                        return
                    
                    self.process_rename_pdf()
                else:
                    # 按班级重命名
                    self.process_class_rename_pdf()
                
                # 按班级重命名模式下，不显示通用成功消息，因为process_class_rename_pdf已经显示了详细的成功消息
                if self.rename_mode_var.get() == "excel":
                    self.status_var.set("处理完成")
                    messagebox.showinfo("成功", "PDF操作完成！")
        except Exception as e:
            self.status_var.set("处理失败")
            messagebox.showerror("错误", f"处理过程中出现错误: {str(e)}")

    def clear_all(self):
        self.main_pdf = ""
        self.insert_pdfs = ["", "", "", "", ""]
        self.main_pdf_var.set("")
        for i in range(len(self.insert_pdf_vars)):
            self.insert_pdf_vars[i].set("")
        self.interval_var.set("1")
        self.insert_pages_var.set("1")
        self.mode_var.set("single")
        # 清空PDF信息显示
        self.main_pdf_info_var.set("")
        for i in range(len(self.insert_pdf_info_vars)):
            self.insert_pdf_info_vars[i].set("")
        # 清空一分多变量
        self.split_pdf = ""
        self.split_pages = 1
        self.excel_file = ""
        self.selected_header = None
        self.selected_id_header = None
        self.split_data = None
        self.split_pdf_var.set("")
        self.split_pages_var.set("1")
        self.excel_var.set("")
        self.header_var.set("未选择")
        self.id_header_var.set("未选择")
        self.naming_var.set("auto")
        # 清空PDF页面交换变量
        self.swap_pdfs = []
        self.page_a = 1
        self.page_b = 2
        self.swap_pdf_var.set("")
        self.page_a_var.set("1")
        self.page_b_var.set("2")
        # 清空PDF批量重命名变量
        self.rename_pdfs = []
        self.rename_excel_file = ""
        self.rename_name_header = None
        self.rename_id_header = None
        self.rename_data = None
        self.rename_pdf_var.set("")
        self.rename_excel_var.set("")
        self.rename_name_header_var.set("未选择")
        self.rename_id_header_var.set("未选择")
        # 清空PDF重排序变量
        self.reorder_pdfs = []
        self.reorder_excel_file = ""
        self.reorder_name_header = None
        self.reorder_id_header = None
        self.reorder_data = None
        self.reorder_pdf_var.set("")
        self.reorder_excel_var.set("")
        self.reorder_name_header_var.set("未选择")
        self.reorder_id_header_var.set("未选择")
        # 清空指定导出页变量
        self.export_pdf = ""
        self.export_pages = ""
        self.export_excel_file = ""
        self.export_name_header = None
        self.export_id_header = None
        self.export_data = None
        self.export_pdf_var.set("")
        self.export_pages_var.set("1-6,1,2,3,4")
        self.export_excel_var.set("")
        self.export_name_header_var.set("未选择")
        self.export_id_header_var.set("未选择")
        self.export_pdf_info_var.set("")
        # 清空PDF页面变量
        self.page_pdfs = []
        self.page_rotation_scheme = []
        self.page_pdf_var.set("")
        self.page_pdf_count_var.set("已选择 0 个文件")
        # 清空批量提取变量
        self.batch_extract_pdfs = []
        self.batch_extract_page = 1
        self.batch_extract_output_dir = ""
        self.batch_extract_pdf_var.set("")
        self.batch_extract_page_var.set("1")
        self.batch_extract_output_var.set("")
        # 清空批量追加页变量
        self.batch_append_pdfs = []
        self.batch_append_source_pdf = ""
        self.batch_append_output_dir = ""
        self.batch_append_source_var.set("")
        self.batch_append_pdf_var.set("")
        self.batch_append_output_var.set("")
        # 清空学生班级PDF合并变量
        self.student_pdfs = []
        self.class_pdfs = []
        self.student_class_excel_file = ""
        self.student_class_name_header = None
        self.student_class_class_header = None
        self.student_class_output_dir = ""
        self.student_pdf_var.set("")
        self.class_pdf_var.set("")
        self.student_class_excel_var.set("")
        self.student_class_name_header_var.set("未选择")
        self.student_class_class_header_var.set("未选择")
        self.student_class_output_var.set("")
        # 清空Excel批量导出PDF变量
        self.excel_files = []
        self.excel_output_dir = ""
        self.excel_files_var.set("未选择文件")
        self.excel_output_var.set("未选择目录")
        self.excel_selected_sheets = []
        if hasattr(self, 'excel_export_all_sheets_var'):
            self.excel_export_all_sheets_var.set(True)
        # 清空Word批量导出PDF变量
        self.word_files = []
        self.word_output_dir = ""
        self.word_files_var.set("未选择文件")
        self.word_output_var.set("未选择目录")
        if hasattr(self, 'word_export_pages_var'):
            self.word_export_pages_var.set("")
        if hasattr(self, 'word_export_all_pages_var'):
            self.word_export_all_pages_var.set(True)
        # 重置状态
        self.status_var.set("就绪")
        self.update_button_state()

    def save_settings(self):
        """保存设置到文件"""
        try:
            # 打开文件选择对话框
            file_path = filedialog.asksaveasfilename(
                title="保存设置",
                defaultextension=".csv",
                filetypes=[("CSV文件", "*.csv"), ("文本文件", "*.txt")]
            )
            
            if not file_path:
                return
            
            # 收集设置
            settings = {
                # 批量提取设置
                "batch_extract_page": self.batch_extract_page_var.get(),
                "batch_extract_output_dir": self.batch_extract_output_dir,
                "batch_extract_excel_files": ",".join(self.batch_extract_excel_files),
                "batch_extract_name_header": self.batch_extract_name_header_var.get(),
                "batch_extract_class_header": self.batch_extract_class_header_var.get(),
                "batch_extract_special_char": self.batch_extract_special_char_var.get(),
                
                # 批量追加页设置
                "batch_append_output_dir": self.batch_append_output_dir,
                "batch_append_excel_files": ",".join(self.batch_append_excel_files),
                "batch_append_name_header": self.batch_append_name_header_var.get(),
                "batch_append_class_header": self.batch_append_class_header_var.get(),
                "batch_append_special_char": self.batch_append_special_char_var.get()
            }
            
            # 保存到文件
            with open(file_path, 'w', encoding='utf-8') as f:
                for key, value in settings.items():
                    f.write(f"{key},{value}\n")
            
            messagebox.showinfo("成功", f"设置已保存到 {file_path}")
        
        except Exception as e:
            messagebox.showerror("错误", f"保存设置失败: {str(e)}")

    def load_settings(self):
        """从文件加载设置"""
        try:
            # 打开文件选择对话框
            file_path = filedialog.askopenfilename(
                title="加载设置",
                filetypes=[("CSV文件", "*.csv"), ("文本文件", "*.txt")]
            )
            
            if not file_path:
                return
            
            # 读取设置
            settings = {}
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line and ',' in line:
                        key, value = line.split(',', 1)
                        settings[key] = value
            
            # 应用设置
            # 批量提取设置
            if "batch_extract_page" in settings:
                self.batch_extract_page_var.set(settings["batch_extract_page"])
            
            if "batch_extract_output_dir" in settings:
                self.batch_extract_output_dir = settings["batch_extract_output_dir"]
                self.batch_extract_output_var.set(self.batch_extract_output_dir)
            
            if "batch_extract_excel_files" in settings:
                self.batch_extract_excel_files = settings["batch_extract_excel_files"].split(',') if settings["batch_extract_excel_files"] else []
                self.batch_extract_excel_var.set(f"已选择 {len(self.batch_extract_excel_files)} 个文件")
            
            if "batch_extract_name_header" in settings:
                self.batch_extract_name_header_var.set(settings["batch_extract_name_header"])
            
            if "batch_extract_class_header" in settings:
                self.batch_extract_class_header_var.set(settings["batch_extract_class_header"])
            
            if "batch_extract_special_char" in settings:
                self.batch_extract_special_char_var.set(settings["batch_extract_special_char"])
            
            # 批量追加页设置
            if "batch_append_output_dir" in settings:
                self.batch_append_output_dir = settings["batch_append_output_dir"]
                self.batch_append_output_var.set(self.batch_append_output_dir)
            
            if "batch_append_excel_files" in settings:
                self.batch_append_excel_files = settings["batch_append_excel_files"].split(',') if settings["batch_append_excel_files"] else []
                self.batch_append_excel_var.set(f"已选择 {len(self.batch_append_excel_files)} 个文件")
            
            if "batch_append_name_header" in settings:
                self.batch_append_name_header_var.set(settings["batch_append_name_header"])
            
            if "batch_append_class_header" in settings:
                self.batch_append_class_header_var.set(settings["batch_append_class_header"])
            
            if "batch_append_special_char" in settings:
                self.batch_append_special_char_var.set(settings["batch_append_special_char"])
            
            messagebox.showinfo("成功", f"设置已从 {file_path} 加载")
        
        except Exception as e:
            messagebox.showerror("错误", f"加载设置失败: {str(e)}")

    def update_button_state(self):
        if self.function_var.get() == "insert":
            # PDF插入功能
            # 检查主PDF和至少一个插入PDF是否已选择
            has_insert_pdfs = any(pdf for pdf in self.insert_pdfs)
            if self.main_pdf and has_insert_pdfs:
                self.process_btn.config(state=tk.NORMAL)
            else:
                self.process_btn.config(state=tk.DISABLED)
        elif self.function_var.get() == "split":
            # PDF分割功能
            if self.split_pdf:
                if self.naming_var.get() == "excel" and not self.excel_file:
                    self.process_btn.config(state=tk.DISABLED)
                else:
                    self.process_btn.config(state=tk.NORMAL)
            else:
                self.process_btn.config(state=tk.DISABLED)
        elif self.function_var.get() == "swap":
            # PDF页面交换功能
            if self.swap_pdfs:
                self.process_btn.config(state=tk.NORMAL)
            else:
                self.process_btn.config(state=tk.DISABLED)
        elif self.function_var.get() == "export":
                # 指定导出页功能
                if self.export_pdf:
                    # 检查是否有子PDF配置
                    if hasattr(self, 'sub_pdf_frames') and len(self.sub_pdf_frames) > 0:
                        # 检查所有子PDF是否都有有效的配置
                        all_valid = True
                        for i in range(len(self.sub_pdf_frames)):
                            range_str = self.sub_pdf_range_vars[i].get()
                            name = self.sub_pdf_name_vars[i].get()
                            path = self.sub_pdf_path_vars[i].get()
                            if not range_str or not name or not path:
                                all_valid = False
                                break
                        if all_valid:
                            self.process_btn.config(state=tk.NORMAL)
                        else:
                            self.process_btn.config(state=tk.DISABLED)
                    else:
                        # 原有的单个页面导出逻辑
                        if self.export_pages_var.get():
                            self.process_btn.config(state=tk.NORMAL)
                        else:
                            self.process_btn.config(state=tk.DISABLED)
                else:
                    self.process_btn.config(state=tk.DISABLED)
        elif self.function_var.get() == "reorder":
            # PDF重排序功能
            if self.reorder_pdfs and self.reorder_excel_file and self.reorder_name_header and self.reorder_id_header:
                self.process_btn.config(state=tk.NORMAL)
            else:
                self.process_btn.config(state=tk.DISABLED)
        elif self.function_var.get() == "page":
            # PDF页面功能
            # 页面功能有独立的按钮，不需要启用process_btn
            self.process_btn.config(state=tk.DISABLED)
        elif self.function_var.get() == "batch_extract":
            # 批量提取功能
            if self.batch_extract_pdfs and self.batch_extract_output_dir:
                self.process_btn.config(state=tk.NORMAL)
            else:
                self.process_btn.config(state=tk.DISABLED)
        elif self.function_var.get() == "batch_append":
            # 批量追加页功能
            if self.batch_append_source_pdf and self.batch_append_pdfs and self.batch_append_output_dir:
                self.process_btn.config(state=tk.NORMAL)
            else:
                self.process_btn.config(state=tk.DISABLED)
        elif self.function_var.get() == "student_class":
            # 学生班级PDF合并功能
            if (self.student_pdfs and self.class_pdfs and self.student_class_excel_file and 
                self.student_class_name_header and self.student_class_class_header and self.student_class_output_dir):
                self.process_btn.config(state=tk.NORMAL)
            else:
                self.process_btn.config(state=tk.DISABLED)
        elif self.function_var.get() == "excel_to_pdf":
            # Excel批量导出PDF功能
            if self.excel_files and self.excel_output_dir:
                self.process_btn.config(state=tk.NORMAL)
            else:
                self.process_btn.config(state=tk.DISABLED)
        elif self.function_var.get() == "word_to_pdf":
            # Word批量导出PDF功能
            if self.word_files and self.word_output_dir:
                self.process_btn.config(state=tk.NORMAL)
            else:
                self.process_btn.config(state=tk.DISABLED)
        else:
            # PDF批量重命名功能
            if self.rename_pdfs and self.rename_excel_file and self.rename_name_header and self.rename_id_header:
                self.process_btn.config(state=tk.NORMAL)
            else:
                self.process_btn.config(state=tk.DISABLED)

    def select_excel_sheets(self):
        """选择要导出的Excel Sheet"""
        if not self.excel_files:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return
        
        # 获取第一个Excel文件的所有sheet名称
        excel_file = self.excel_files[0]
        sheets = []
        
        try:
            ext = os.path.splitext(excel_file)[1].lower()
            if ext == '.xlsx':
                import openpyxl
                wb = openpyxl.load_workbook(excel_file, read_only=True)
                sheets = wb.sheetnames
                wb.close()
            elif ext == '.xls':
                import xlrd
                wb = xlrd.open_workbook(excel_file)
                sheets = wb.sheet_names()
            else:
                messagebox.showerror("错误", "不支持的Excel格式")
                return
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件时出错：{str(e)}")
            return
        
        # 创建选择窗口
        select_window = tk.Toplevel(self.root)
        select_window.title("选择Sheet")
        select_window.geometry("400x300")
        
        # 创建滚动条
        scrollbar = ttk.Scrollbar(select_window)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建列表框
        listbox = tk.Listbox(select_window, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set)
        listbox.pack(fill=tk.BOTH, expand=True)
        
        # 配置滚动条
        scrollbar.config(command=listbox.yview)
        
        # 添加sheet到列表框
        for sheet in sheets:
            listbox.insert(tk.END, sheet)
        
        # 确定按钮
        def confirm_selection():
            selected_indices = listbox.curselection()
            self.excel_selected_sheets = [sheets[i] for i in selected_indices]
            self.excel_export_all_sheets_var.set(False)
            messagebox.showinfo("成功", f"已选择 {len(self.excel_selected_sheets)} 个Sheet")
            select_window.destroy()
        
        confirm_btn = ttk.Button(select_window, text="确定", command=confirm_selection)
        confirm_btn.pack(pady=10)
        
        select_window.transient(self.root)
        select_window.grab_set()

    def process_excel_to_pdf(self):
        # Excel批量导出PDF逻辑（增强版，支持xls和xlsx混合，支持多Sheet）
        total_files = len(self.excel_files)
        processed_count = 0
        success_count = 0
        fail_count = 0
        fail_messages = []
        
        try:
            import win32com.client
            import pythoncom
            
            # 初始化Excel应用
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            for excel_file in self.excel_files:
                try:
                    # 打开Excel文件
                    workbook = excel.Workbooks.Open(excel_file)
                    
                    # 生成输出PDF文件名
                    base_name = os.path.splitext(os.path.basename(excel_file))[0]
                    pdf_path = os.path.join(self.excel_output_dir, f"{base_name}.pdf")
                    
                    # 根据选择决定导出哪些sheet
                    if self.excel_export_all_sheets_var.get():
                        # 导出所有sheet
                        workbook.ExportAsFixedFormat(0, pdf_path)  # 0 = xlTypePDF
                    else:
                        # 只导出选定的sheet
                        if not self.excel_selected_sheets:
                            workbook.ExportAsFixedFormat(0, pdf_path)
                        else:
                            # 先隐藏所有sheet，再显示选定的sheet
                            for sheet in workbook.Sheets:
                                sheet.Visible = False
                            
                            for sheet_name in self.excel_selected_sheets:
                                workbook.Sheets(sheet_name).Visible = True
                            
                            # 导出可见的sheet
                            workbook.ExportAsFixedFormat(0, pdf_path)
                            
                            # 恢复所有sheet的可见性
                            for sheet in workbook.Sheets:
                                sheet.Visible = True
                    
                    # 关闭工作簿
                    workbook.Close(SaveChanges=False)
                    
                    processed_count += 1
                    success_count += 1
                    
                    # 显示进度
                    progress = processed_count / total_files * 100
                    self.status_var.set(f"处理中... {progress:.1f}%")
                    self.root.update()
                    
                except Exception as e:
                    processed_count += 1
                    fail_count += 1
                    fail_msg = f"{os.path.basename(excel_file)}: {str(e)}"
                    fail_messages.append(fail_msg)
                finally:
                    # 确保工作簿关闭
                    try:
                        if 'workbook' in locals():
                            workbook.Close(SaveChanges=False)
                    except:
                        pass
            
            # 关闭Excel应用
            excel.Quit()
            pythoncom.CoUninitialize()
            
            self.status_var.set("处理完成")
            
            # 显示详细结果
            result_msg = f"Excel批量导出PDF完成！\n\n处理文件数：{total_files}\n成功导出：{success_count}"
            if fail_count > 0:
                result_msg += f"\n失败：{fail_count}"
                result_msg += "\n\n失败详情：\n" + "\n".join(fail_messages[:5])
                if len(fail_messages) > 5:
                    result_msg += f"\n... 还有 {len(fail_messages) - 5} 个失败文件"
            
            if fail_count > 0:
                messagebox.showwarning("处理完成", result_msg)
            else:
                messagebox.showinfo("成功", result_msg)
            
        except Exception as e:
            self.status_var.set("处理失败")
            messagebox.showerror("错误", f"处理过程中出现错误: {str(e)}")
            try:
                if 'excel' in locals():
                    excel.Quit()
                pythoncom.CoUninitialize()
            except:
                pass

    def process_word_to_pdf(self):
        # Word批量导出PDF逻辑（增强版，支持页面选择和错误重试）
        total_files = len(self.word_files)
        processed_count = 0
        success_count = 0
        fail_count = 0
        fail_messages = []
        
        # Word常量定义
        wdExportFormatPDF = 17  # 正确的PDF导出格式常量
        wdStatisticPages = 2
        
        try:
            import win32com.client
            import pythoncom
            import time
            
            # 初始化Word应用
            pythoncom.CoInitialize()
            
            # 尝试创建Word对象
            try:
                word = win32com.client.Dispatch("Word.Application")
            except Exception as e:
                # 尝试使用其他方式创建Word对象
                try:
                    word = win32com.client.DispatchEx("Word.Application")
                except Exception as e2:
                    raise Exception(f"无法启动Word应用程序: {str(e2)}")
            
            word.Visible = False
            word.DisplayAlerts = 0  # wdAlertsNone = 0
            
            for word_file in self.word_files:
                attempts = 0
                max_attempts = 3
                success = False
                
                while attempts < max_attempts and not success:
                    attempts += 1
                    doc = None
                    
                    try:
                        # 确保路径正确处理
                        word_file_path = os.path.abspath(word_file)
                        
                        # 检查文件是否存在
                        if not os.path.exists(word_file_path):
                            raise Exception("文件不存在")
                        
                        # 打开Word文件
                        # 使用完整参数打开，确保正确处理
                        doc = word.Documents.Open(
                            FileName=word_file_path,
                            ReadOnly=True,
                            Visible=False
                        )
                        
                        # 生成输出PDF文件名
                        base_name = os.path.splitext(os.path.basename(word_file))[0]
                        # 清理文件名中的特殊字符
                        base_name = self.clean_filename(base_name)
                        pdf_path = os.path.join(self.word_output_dir, f"{base_name}.pdf")
                        pdf_path = os.path.abspath(pdf_path)
                        
                        # 根据选择决定导出范围
                        if self.word_export_all_pages_var.get():
                            # 导出全部页面
                            doc.ExportAsFixedFormat(
                                OutputFileName=pdf_path,
                                ExportFormat=wdExportFormatPDF
                            )
                        else:
                            # 导出指定页面范围
                            page_range = self.word_export_pages_var.get().strip()
                            if not page_range:
                                doc.ExportAsFixedFormat(
                                    OutputFileName=pdf_path,
                                    ExportFormat=wdExportFormatPDF
                                )
                            else:
                                # 解析页面范围
                                pages = self.parse_page_range(page_range)
                                if pages:
                                    # 获取总页数
                                    total_pages = doc.ComputeStatistics(wdStatisticPages)
                                    
                                    # 过滤有效页码
                                    valid_pages = [p for p in pages if 1 <= p <= total_pages]
                                    
                                    if valid_pages:
                                        from_page = min(valid_pages)
                                        to_page = max(valid_pages)
                                        
                                        doc.ExportAsFixedFormat(
                                            OutputFileName=pdf_path,
                                            ExportFormat=wdExportFormatPDF,
                                            From=from_page,
                                            To=to_page
                                        )
                                    else:
                                        # 没有有效页码，导出全部
                                        doc.ExportAsFixedFormat(
                                            OutputFileName=pdf_path,
                                            ExportFormat=wdExportFormatPDF
                                        )
                                else:
                                    doc.ExportAsFixedFormat(
                                        OutputFileName=pdf_path,
                                        ExportFormat=wdExportFormatPDF
                                    )
                        
                        # 关闭文档
                        if doc:
                            doc.Close(SaveChanges=False)
                        
                        success = True
                        
                    except Exception as e:
                        # 清理文档对象
                        if doc:
                            try:
                                doc.Close(SaveChanges=False)
                            except:
                                pass
                        
                        # 如果是最后一次尝试，记录错误
                        if attempts >= max_attempts:
                            fail_msg = f"{os.path.basename(word_file)}: {str(e)}"
                            fail_messages.append(fail_msg)
                        else:
                            # 等待一下再重试
                            time.sleep(1)
                
                if success:
                    processed_count += 1
                    success_count += 1
                else:
                    processed_count += 1
                    fail_count += 1
                
                # 显示进度
                progress = processed_count / total_files * 100
                self.status_var.set(f"处理中... {progress:.1f}%")
                self.root.update()
            
            # 关闭Word应用
            try:
                word.Quit()
            except:
                pass
            pythoncom.CoUninitialize()
            
            self.status_var.set("处理完成")
            
            # 显示详细结果
            result_msg = f"Word批量导出PDF完成！\n\n处理文件数：{total_files}\n成功导出：{success_count}"
            if fail_count > 0:
                result_msg += f"\n失败：{fail_count}"
                result_msg += "\n\n失败详情：\n" + "\n".join(fail_messages[:5])
                if len(fail_messages) > 5:
                    result_msg += f"\n... 还有 {len(fail_messages) - 5} 个失败文件"
            
            if fail_count > 0:
                messagebox.showwarning("处理完成", result_msg)
            else:
                messagebox.showinfo("成功", result_msg)
            
        except Exception as e:
            self.status_var.set("处理失败")
            error_msg = f"处理过程中出现错误: {str(e)}\n\n可能的原因:\n1. Word未安装或版本不兼容\n2. Word文件被其他程序占用\n3. 文件路径包含特殊字符\n4. 输出目录权限不足"
            messagebox.showerror("错误", error_msg)
            try:
                if 'word' in locals():
                    try:
                        word.Quit()
                    except:
                        pass
                pythoncom.CoUninitialize()
            except:
                pass

    def clean_filename(self, filename):
        """清理文件名中的特殊字符"""
        import re
        # 移除或替换非法字符
        cleaned = re.sub(r'[\\/*?:"<>|]', '_', filename)
        # 移除多余的下划线
        cleaned = re.sub(r'__+', '_', cleaned)
        # 移除首尾下划线
        cleaned = cleaned.strip('_')
        return cleaned

    def parse_page_range(self, range_str):
        """解析页面范围字符串，如 "2-5,7" 返回 [2,3,4,5,7]"""
        pages = []
        try:
            parts = range_str.split(',')
            for part in parts:
                part = part.strip()
                if '-' in part:
                    start_end = part.split('-')
                    if len(start_end) == 2:
                        start = int(start_end[0].strip())
                        end = int(start_end[1].strip())
                        pages.extend(range(start, end + 1))
                elif part.isdigit():
                    pages.append(int(part))
            # 去重并排序
            pages = sorted(list(set(pages)))
        except:
            pages = []
        return pages

    def preview_insert(self):
        # 预览PDF插入结果
        if not self.main_pdf:
            messagebox.showerror("错误", "请选择主PDF文件")
            return
        
        # 检查是否有有效的插入PDF文件
        valid_insert_pdfs = [pdf for pdf in self.insert_pdfs if pdf]
        if not valid_insert_pdfs:
            messagebox.showerror("错误", "请选择要插入的PDF文件")
            return
        
        try:
            # 获取插入参数
            insert_interval = int(self.interval_var.get())
            insert_pages = int(self.insert_pages_var.get())
            mode = self.mode_var.get()
            
            if insert_interval < 1 or insert_pages < 1:
                messagebox.showerror("错误", "插入间隔和每次插入页数必须大于0")
                return
            
            # 读取主PDF信息
            with open(self.main_pdf, 'rb') as f:
                main_reader = PyPDF2.PdfReader(f)
                main_pages = len(main_reader.pages)
            
            # 读取插入PDF信息
            insert_info = []
            # 过滤掉空的PDF文件路径
            valid_insert_pdfs = [pdf for pdf in self.insert_pdfs if pdf]
            for pdf_path in valid_insert_pdfs:
                with open(pdf_path, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    insert_info.append((os.path.basename(pdf_path), len(reader.pages)))
            
            # 创建预览窗口
            preview_window = tk.Toplevel(self.root)
            preview_window.title("PDF插入预览")
            preview_window.geometry("600x400")
            
            # 创建滚动条
            scrollbar = ttk.Scrollbar(preview_window)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 创建文本框
            text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('Courier New', 10))
            text.pack(fill=tk.BOTH, expand=True)
            
            # 配置滚动条
            scrollbar.config(command=text.yview)
            
            # 显示标题
            text.insert(tk.END, "PDF插入预览\n")
            text.insert(tk.END, "=" * 60 + "\n")
            text.insert(tk.END, f"主PDF：{os.path.basename(self.main_pdf)}（{main_pages}页）\n")
            text.insert(tk.END, f"插入间隔：{insert_interval}页\n")
            text.insert(tk.END, f"每次插入页数：{insert_pages}页\n")
            text.insert(tk.END, f"操作模式：{'单个PDF插入（隔页）' if mode == 'single' else '批量PDF插入（每页依次）'}\n\n")
            
            # 显示插入PDF信息
            text.insert(tk.END, "插入PDF文件：\n")
            for pdf_name, pdf_pages in insert_info:
                text.insert(tk.END, f"  - {pdf_name}（{pdf_pages}页）\n")
            text.insert(tk.END, "\n")
            
            # 计算插入后的页数
            # 过滤掉空的PDF文件路径
            valid_insert_pdfs = [pdf for pdf in self.insert_pdfs if pdf]
            if not valid_insert_pdfs:
                messagebox.showerror("错误", "请至少选择一个插入PDF文件")
                return
            
            if mode == 'single':
                # 单个PDF插入模式
                insert_count = main_pages // insert_interval
                total_insert_pages = insert_count * insert_pages * len(valid_insert_pdfs)
                total_pages_after = main_pages + total_insert_pages
            elif mode == 'batch':
                # 批量PDF插入模式
                insert_count = min(main_pages // insert_interval, len(valid_insert_pdfs))
                total_insert_pages = insert_count * insert_pages
                total_pages_after = main_pages + total_insert_pages
            else:
                # 多PDF顺序插入模式
                multi_count = int(self.multi_count_var.get())
                insert_count = main_pages // insert_interval
                total_insert_pdfs = min(insert_count * multi_count, len(valid_insert_pdfs))
                total_insert_pages = total_insert_pdfs * insert_pages
                total_pages_after = main_pages + total_insert_pages
            
            text.insert(tk.END, "插入结果预览：\n")
            text.insert(tk.END, f"  插入次数：{insert_count}\n")
            text.insert(tk.END, f"  插入总页数：{total_insert_pages}\n")
            text.insert(tk.END, f"  插入后总页数：{total_pages_after}\n")
            
            if mode == 'multi':
                multi_count = int(self.multi_count_var.get())
                text.insert(tk.END, f"  每次插入PDF数量：{multi_count}\n")
                text.insert(tk.END, f"  预计插入PDF总数：{min(insert_count * multi_count, len(self.insert_pdfs))}\n")
            
            # 禁用文本编辑
            text.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("错误", f"预览失败: {str(e)}")

    def show_help(self):
        """显示使用说明窗口"""
        help_window = tk.Toplevel(self.root)
        help_window.title("使用说明")
        help_window.geometry("800x600")
        help_window.transient(self.root)
        help_window.grab_set()
        
        # 创建滚动条
        scrollbar = ttk.Scrollbar(help_window)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建文本框
        text = tk.Text(help_window, yscrollcommand=scrollbar.set, font=('Microsoft YaHei', 10))
        text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 配置滚动条
        scrollbar.config(command=text.yview)
        
        # 写入使用说明
        help_text = """PDF操作工具使用说明
===================

本工具提供了多种PDF操作功能，包括PDF插入、分割、页面交换、指定导出页、批量重命名、重排序和页面操作。

1. PDF插入功能
   - 选择主PDF文件
   - 选择要插入的PDF文件（最多5个）
   - 设置插入间隔和插入页数
   - 点击"开始插入"按钮执行操作

2. PDF分割功能
   - 选择要分割的PDF文件
   - 设置分割页数
   - 选择Excel文件和姓名表头
   - 点击"开始分割"按钮执行操作

3. PDF页面交换功能
   - 选择要交换页面的PDF文件
   - 输入要交换的两个页面编号
   - 点击"开始交换"按钮执行操作

4. 指定导出页功能
   - 选择PDF文件
   - 输入要导出的页面范围
   - 选择Excel文件和姓名表头
   - 点击"开始导出"按钮执行操作

5. PDF批量重命名功能
   - 选择要重命名的PDF文件
   - 选择Excel文件
   - 构建文件名规则：
     - 点击"添加文本"按钮添加固定文本
     - 点击"添加Excel列"按钮添加Excel数据列
     - 拖拽调整构建单元的顺序
   - 点击"预览重命名结果"按钮查看效果
   - 点击"开始重命名"按钮执行操作
   - 匹配规则：使用原始文件名匹配Excel表中的任意表头

6. PDF重排序功能
   - 选择要重排序的PDF文件
   - 选择Excel文件和姓名、序号表头
   - 点击"打开中转表预览"按钮确认数据
   - 点击"开始重排序"按钮执行操作

7. PDF页面功能
   - 选择PDF文件（最多100个）
   - 点击"一键颠倒页面顺序"按钮颠倒页面顺序
   - 点击"设置页面旋转方案"按钮设置旋转方案
   - 在预览窗口中为每一页设置旋转方向
   - 点击"应用旋转方案到所有PDF"按钮执行操作

快捷键：
- F11：最大化窗口
- Escape：恢复窗口大小

实现方法
=========

1. PDF插入：使用PyPDF2库读取PDF文件，按指定间隔插入页面
2. PDF分割：使用PyPDF2库读取PDF文件，按指定页数分割
3. 页面交换：使用PyPDF2库读取PDF文件，交换指定页面
4. 指定导出页：使用PyPDF2库读取PDF文件，导出指定页面范围
5. 批量重命名：
   - 使用openpyxl/xlrd库读取Excel数据
   - 根据原始文件名匹配Excel数据
   - 根据构建规则生成新文件名
   - 使用shutil库复制文件并重命名
6. 重排序：根据Excel数据中的序号对PDF文件进行重排序
7. 页面操作：
   - 颠倒页面顺序：使用PyPDF2库读取PDF文件，反向添加页面
   - 旋转页面：使用PyPDF2库读取PDF文件，设置页面旋转角度
   - 使用PyMuPDF库预览PDF页面

注意事项
=========
1. 确保PDF文件和Excel文件格式正确
2. 确保有足够的磁盘空间保存操作结果
3. 操作大文件时可能需要较长时间，请耐心等待
4. 如遇权限问题，请确保有足够的文件操作权限
"""
        
        text.insert(tk.END, help_text)
        text.config(state=tk.DISABLED)
        
        # 关闭按钮
        button_frame = ttk.Frame(help_window)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        close_btn = ttk.Button(button_frame, text="关闭", command=help_window.destroy)
        close_btn.pack(side=tk.RIGHT)

class ExcelDataPreviewWindow:
    def __init__(self, parent, data, title="Excel数据预览"):
        self.parent = parent
        self.original_data = data  # 保存原始数据
        self.data = data.copy()  # 工作数据
        self.result = None
        self.start_row = 1
        self.end_row = len(data)
        self.window = tk.Toplevel(parent)
        self.window.title(title)
        self.window.geometry("800x650")
        
        self.create_widgets()
        
    def create_widgets(self):
        # 创建主框架
        main_frame = ttk.Frame(self.window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建范围设置框架
        range_frame = ttk.LabelFrame(main_frame, text="行范围设置", padding="5")
        range_frame.pack(fill=tk.X, pady=5)
        
        # 起始行
        start_frame = ttk.Frame(range_frame)
        start_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        start_label = ttk.Label(start_frame, text="起始行:")
        start_label.pack(side=tk.LEFT, padx=5)
        
        self.start_var = tk.StringVar(value="1")
        start_entry = ttk.Entry(start_frame, textvariable=self.start_var, width=10)
        start_entry.pack(side=tk.LEFT, padx=5)
        
        # 结束行
        end_frame = ttk.Frame(range_frame)
        end_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        end_label = ttk.Label(end_frame, text="结束行:")
        end_label.pack(side=tk.LEFT, padx=5)
        
        self.end_var = tk.StringVar(value=str(len(self.original_data)))
        end_entry = ttk.Entry(end_frame, textvariable=self.end_var, width=10)
        end_entry.pack(side=tk.LEFT, padx=5)
        
        # 应用范围按钮
        apply_btn = ttk.Button(range_frame, text="应用范围", command=self.apply_range)
        apply_btn.pack(side=tk.LEFT, padx=5)
        
        # 创建表格框架
        table_frame = ttk.Frame(main_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # 创建表格
        self.tree = ttk.Treeview(table_frame, columns=("name", "id"), show="headings", height=15)
        self.tree.heading("name", text="姓名")
        self.tree.heading("id", text="序号")
        self.tree.column("name", width=300)
        self.tree.column("id", width=150)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 创建滚动条
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # 统计信息
        stats_frame = ttk.Frame(main_frame)
        stats_frame.pack(fill=tk.X, pady=5)
        
        self.stats_var = tk.StringVar(value=f"总记录数：{len(self.data)}")
        stats_label = ttk.Label(stats_frame, textvariable=self.stats_var)
        stats_label.pack(side=tk.LEFT, padx=5)
        
        # 创建按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        # 确定按钮
        confirm_btn = ttk.Button(button_frame, text="确定", command=self.on_confirm)
        confirm_btn.pack(side=tk.LEFT, padx=5)
        
        # 取消按钮
        cancel_btn = ttk.Button(button_frame, text="取消", command=self.on_cancel)
        cancel_btn.pack(side=tk.LEFT, padx=5)
        
        # 反转按钮（合并姓名和序号反转功能）
        reverse_btn = ttk.Button(button_frame, text="反转顺序", command=self.on_reverse)
        reverse_btn.pack(side=tk.LEFT, padx=5)
        
        # 填充数据
        self.refresh_table()
        
        # 设置窗口模态
        self.window.transient(self.parent)
        self.window.grab_set()
        self.window.wait_window()
        
    def refresh_table(self):
        # 清空表格
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 填充数据
        for i, (name, id_value) in enumerate(self.data):
            # 确保序号以整数形式显示
            try:
                id_value = int(id_value)
            except (ValueError, TypeError):
                pass
            self.tree.insert("", tk.END, values=(name, id_value))
        
        # 更新统计信息
        self.stats_var.set(f"总记录数：{len(self.data)}")
        
    def on_confirm(self):
        # 确认操作
        self.result = self.data
        self.window.destroy()
        
    def on_cancel(self):
        # 取消操作
        self.result = None
        self.window.destroy()
        
    def on_reverse(self):
        # 反转数据顺序
        self.data = [(name, id_value) for name, id_value in reversed(self.data)]
        self.refresh_table()
    
    def apply_range(self):
        # 应用行范围
        try:
            start = int(self.start_var.get())
            end = int(self.end_var.get())
            
            if start < 1 or end > len(self.original_data) or start > end:
                messagebox.showerror("错误", f"行范围无效，请输入1-{len(self.original_data)}之间的值")
                return
            
            # 应用范围过滤
            self.data = self.original_data[start-1:end]
            self.start_row = start
            self.end_row = end
            self.refresh_table()
            
        except ValueError:
            messagebox.showerror("错误", "请输入有效的数字")

class RotationPreviewWindow:
    def __init__(self, parent, pdf_paths, callback):
        self.parent = parent
        self.pdf_paths = pdf_paths
        self.callback = callback
        self.rotation_scheme = []
        self.pdf_document = None
        
        self.window = tk.Toplevel(parent)
        self.window.title("页面旋转预览")
        self.window.geometry("1000x700")
        
        self.create_widgets()
        self.load_first_pdf()
    
    def create_widgets(self):
        main_frame = ttk.Frame(self.window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        title_label = ttk.Label(main_frame, text="为第一个PDF的每一页设置旋转方向", font=('微软雅黑', 12, 'bold'))
        title_label.pack(pady=10)
        
        scroll_frame = ttk.Frame(main_frame)
        scroll_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(scroll_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        canvas = tk.Canvas(scroll_frame, yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar.config(command=canvas.yview)
        
        content_frame = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=content_frame, anchor=tk.NW)
        
        self.content_frame = content_frame
        self.canvas = canvas
        
        self.page_controls = []
        self.page_previews = []
        
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        apply_btn = ttk.Button(button_frame, text="应用旋转方案到所有PDF", command=self.apply_rotation)
        apply_btn.pack(side=tk.LEFT, padx=5)
        
        cancel_btn = ttk.Button(button_frame, text="取消", command=self.window.destroy)
        cancel_btn.pack(side=tk.LEFT, padx=5)
        
        self.window.transient(self.parent)
        self.window.grab_set()
    
    def load_first_pdf(self):
        try:
            # 打开PDF文档
            self.pdf_document = fitz.open(self.pdf_paths[0])
            total_pages = len(self.pdf_document)
            
            for i in range(total_pages):
                self.rotation_scheme.append(0)
                
                page_frame = ttk.LabelFrame(self.content_frame, text=f"第 {i+1} 页", padding="10")
                page_frame.pack(fill=tk.X, pady=10, padx=5)
                
                # 创建页面内容框架（左侧预览，右侧控制）
                content_frame = ttk.Frame(page_frame)
                content_frame.pack(fill=tk.X)
                
                # 左侧：PDF预览
                preview_frame = ttk.Frame(content_frame, width=300, height=400)
                preview_frame.pack(side=tk.LEFT, padx=10, fill=tk.BOTH, expand=True)
                preview_frame.pack_propagate(False)
                
                # 创建预览画布
                preview_canvas = tk.Canvas(preview_frame, bg="white")
                preview_canvas.pack(fill=tk.BOTH, expand=True)
                
                # 右侧：旋转控制
                control_frame = ttk.Frame(content_frame, width=300)
                control_frame.pack(side=tk.RIGHT, padx=10, fill=tk.Y)
                
                rotation_var = tk.StringVar(value="0")
                
                ttk.Label(control_frame, text="旋转方向:").pack(anchor=tk.W, pady=5)
                
                ttk.Radiobutton(control_frame, text="不旋转", variable=rotation_var, value="0", 
                               command=lambda idx=i, var=rotation_var: self.update_rotation(idx, var)).pack(anchor=tk.W, pady=2)
                ttk.Radiobutton(control_frame, text="顺时针90度", variable=rotation_var, value="90", 
                               command=lambda idx=i, var=rotation_var: self.update_rotation(idx, var)).pack(anchor=tk.W, pady=2)
                ttk.Radiobutton(control_frame, text="逆时针90度", variable=rotation_var, value="-90", 
                               command=lambda idx=i, var=rotation_var: self.update_rotation(idx, var)).pack(anchor=tk.W, pady=2)
                
                self.page_controls.append(rotation_var)
                self.page_previews.append((preview_canvas, i))
                
                # 初始渲染页面
                self.render_page(i, preview_canvas, 0)
            
            self.content_frame.update_idletasks()
            self.canvas.config(scrollregion=self.canvas.bbox("all"))
            
        except Exception as e:
            messagebox.showerror("错误", f"加载PDF文件时出错：{str(e)}")
            self.window.destroy()
    
    def render_page(self, page_num, canvas, rotation):
        """渲染PDF页面到画布"""
        try:
            # 清除画布
            canvas.delete("all")
            
            # 获取页面
            page = self.pdf_document[page_num]
            
            # 设置旋转
            page.set_rotation(rotation)
            
            # 渲染页面为图片
            zoom = 0.5  # 缩放比例
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            
            # 将图片转换为Tkinter可用的格式
            from PIL import Image, ImageTk
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            photo = ImageTk.PhotoImage(img)
            
            # 保存photo引用，防止被垃圾回收
            if not hasattr(self, 'photos'):
                self.photos = []
            # 确保photos列表足够大
            while len(self.photos) <= page_num:
                self.photos.append(None)
            self.photos[page_num] = photo
            
            # 计算居中位置
            canvas_width = canvas.winfo_width()
            canvas_height = canvas.winfo_height()
            x = (canvas_width - pix.width) // 2
            y = (canvas_height - pix.height) // 2
            
            # 绘制图片
            canvas.create_image(max(0, x), max(0, y), anchor=tk.NW, image=photo)
            
        except Exception as e:
            print(f"渲染页面时出错：{str(e)}")
    
    def update_rotation(self, page_idx, var):
        """更新页面旋转并重新渲染"""
        try:
            rotation = int(var.get())
            self.rotation_scheme[page_idx] = rotation
            
            # 重新渲染页面
            canvas, _ = self.page_previews[page_idx]
            self.render_page(page_idx, canvas, rotation)
        except Exception as e:
            print(f"更新旋转时出错：{str(e)}")
    
    def apply_rotation(self):
        self.rotation_scheme = []
        for control in self.page_controls:
            self.rotation_scheme.append(int(control.get()))
        
        # 关闭PDF文档
        if self.pdf_document:
            self.pdf_document.close()
        
        self.callback(self.rotation_scheme)
        self.window.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = PDFEditor(root)
    root.mainloop()
