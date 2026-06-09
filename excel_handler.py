#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel处理模块
包含Excel文件的读取和数据处理功能
"""

import os
import re
from tkinter import messagebox


def read_excel_headers_with_merged_cells(excel_file):
    """读取Excel文件的表头，支持前三行和合并单元格"""
    headers = []
    try:
        if excel_file.endswith('.xlsx'):
            from openpyxl import load_workbook
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            
            # 获取合并单元格信息
            merged_cells = sheet.merged_cells
            
            # 读取前三行的数据
            for row_idx in range(1, 4):
                row_headers = []
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    
                    # 检查是否是合并单元格
                    is_merged = False
                    for merged_range in merged_cells:
                        if cell.coordinate in merged_range:
                            is_merged = True
                            # 合并单元格只取左上角单元格的值
                            if cell.coordinate == merged_range.start_cell.coordinate:
                                if cell.value is not None:
                                    row_headers.append(str(cell.value))
                                else:
                                    row_headers.append("")
                            break
                    
                    if not is_merged:
                        if cell.value is not None:
                            row_headers.append(str(cell.value))
                        else:
                            row_headers.append("")
                
                headers.extend([h for h in row_headers if h])
        else:
            import xlrd
            workbook = xlrd.open_workbook(excel_file)
            sheet = workbook.sheet_by_index(0)
            
            # 读取前三行的数据
            for row_idx in range(3):
                for col_idx in range(sheet.ncols):
                    value = sheet.cell(row_idx, col_idx).value
                    if value is not None and str(value).strip():
                        headers.append(str(value))
        
        # 去重但保持顺序
        seen = set()
        unique_headers = []
        for h in headers:
            if h not in seen:
                seen.add(h)
                unique_headers.append(h)
        
        return unique_headers
    except Exception as e:
        messagebox.showerror("错误", f"读取Excel表头时出错：{str(e)}")
        return []


def read_excel_data(excel_files, name_header, class_header):
    """读取Excel数据并创建索引，支持前三行表头和合并单元格"""
    data_index = {}
    
    for excel_file in excel_files:
        try:
            if excel_file.endswith('.xlsx'):
                from openpyxl import load_workbook
                workbook = load_workbook(excel_file, data_only=False)
                sheet = workbook.active
                
                # 获取合并单元格信息
                merged_cells = sheet.merged_cells
                
                # 扫描前三行找到表头行
                header_row = -1
                headers = []
                
                for row_idx in range(1, 4):
                    row_headers = []
                    for col_idx in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        
                        # 检查是否是合并单元格
                        is_merged = False
                        for merged_range in merged_cells:
                            if cell.coordinate in merged_range:
                                is_merged = True
                                if cell.coordinate == merged_range.start_cell.coordinate:
                                    row_headers.append(str(cell.value) if cell.value else "")
                                else:
                                    row_headers.append("")
                                break
                        
                        if not is_merged:
                            row_headers.append(str(cell.value) if cell.value else "")
                    
                    # 检查是否包含姓名和班级表头
                    if name_header in row_headers and class_header in row_headers:
                        header_row = row_idx
                        headers = row_headers
                        break
                
                if header_row == -1:
                    continue
                
                # 查找姓名和班级列
                name_col = headers.index(name_header) + 1
                class_col = headers.index(class_header) + 1
                
                # 读取数据行
                for row_idx in range(header_row + 1, sheet.max_row + 1):
                    name_cell = sheet.cell(row=row_idx, column=name_col)
                    class_cell = sheet.cell(row=row_idx, column=class_col)
                    
                    name = str(name_cell.value).strip() if name_cell.value else ""
                    class_name = str(class_cell.value).strip() if class_cell.value else ""
                    
                    if name:
                        data_index[name] = class_name
            else:
                import xlrd
                workbook = xlrd.open_workbook(excel_file)
                sheet = workbook.sheet_by_index(0)
                
                # 扫描前三行找到表头行
                header_row = -1
                headers = []
                
                for row_idx in range(min(3, sheet.nrows)):
                    row_headers = [str(sheet.cell(row_idx, col_idx).value).strip() for col_idx in range(sheet.ncols)]
                    if name_header in row_headers and class_header in row_headers:
                        header_row = row_idx
                        headers = row_headers
                        break
                
                if header_row == -1:
                    continue
                
                # 查找姓名和班级列
                name_col = headers.index(name_header)
                class_col = headers.index(class_header)
                
                # 读取数据行
                for row_idx in range(header_row + 1, sheet.nrows):
                    name = str(sheet.cell(row_idx, name_col).value).strip()
                    class_name = str(sheet.cell(row_idx, class_col).value).strip()
                    
                    if name:
                        data_index[name] = class_name
        except Exception as e:
            messagebox.showwarning("警告", f"读取Excel文件 {os.path.basename(excel_file)} 时出错：{str(e)}")
    
    return data_index


def match_filename_with_excel(extracted_name, data_index):
    """使用提取的姓名匹配Excel数据"""
    if not extracted_name or not data_index:
        return None, None
    
    # 精确匹配
    if extracted_name in data_index:
        return extracted_name, data_index[extracted_name]
    
    # 模糊匹配（去除空格）
    extracted_name_clean = extracted_name.replace(' ', '').replace('　', '')
    for name in data_index:
        name_clean = name.replace(' ', '').replace('　', '')
        if extracted_name_clean == name_clean:
            return name, data_index[name]
    
    # 包含匹配
    for name in data_index:
        if extracted_name in name or name in extracted_name:
            return name, data_index[name]
    
    return None, None


def read_name_id_mapping(excel_file, name_header, id_header):
    """读取姓名和序号的映射关系"""
    mapping = []
    
    try:
        if excel_file.endswith('.xlsx'):
            from openpyxl import load_workbook
            workbook = load_workbook(excel_file, data_only=True)
            sheet = workbook.active
            
            # 扫描前三行找到表头行
            header_row = -1
            headers = []
            
            for row_idx in range(1, 4):
                row_headers = []
                for col_idx in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    row_headers.append(str(cell_value).strip() if cell_value else "")
                
                if name_header in row_headers and id_header in row_headers:
                    header_row = row_idx
                    headers = row_headers
                    break
            
            if header_row == -1:
                return mapping
            
            # 查找姓名和序号列
            name_col = headers.index(name_header) + 1
            id_col = headers.index(id_header) + 1
            
            # 读取数据行
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                name_cell = sheet.cell(row=row_idx, column=name_col)
                id_cell = sheet.cell(row=row_idx, column=id_col)
                
                name = str(name_cell.value).strip() if name_cell.value else ""
                id_value = str(id_cell.value).strip() if id_cell.value else ""
                
                if name:
                    mapping.append((name, id_value))
        else:
            import xlrd
            workbook = xlrd.open_workbook(excel_file)
            sheet = workbook.sheet_by_index(0)
            
            # 扫描前三行找到表头行
            header_row = -1
            headers = []
            
            for row_idx in range(min(3, sheet.nrows)):
                row_headers = [str(sheet.cell(row_idx, col_idx).value).strip() for col_idx in range(sheet.ncols)]
                if name_header in row_headers and id_header in row_headers:
                    header_row = row_idx
                    headers = row_headers
                    break
            
            if header_row == -1:
                return mapping
            
            # 查找姓名和序号列
            name_col = headers.index(name_header)
            id_col = headers.index(id_header)
            
            # 读取数据行
            for row_idx in range(header_row + 1, sheet.nrows):
                name = str(sheet.cell(row_idx, name_col).value).strip()
                id_value = str(sheet.cell(row_idx, id_col).value).strip()
                
                if name:
                    mapping.append((name, id_value))
    except Exception as e:
        messagebox.showerror("错误", f"读取Excel数据时出错：{str(e)}")
    
    return mapping