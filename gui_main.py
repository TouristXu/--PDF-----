#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PDF操作工具 - 主GUI模块
"""

import os
import re
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import shutil
import fitz  # PyMuPDF - 用于PDF渲染预览
import PyPDF2  # 用于PDF操作

# 导入自定义模块
from utils import is_valid_pdf, parse_page_range, clean_filename, show_error, show_info, show_warning
from excel_handler import read_excel_headers_with_merged_cells, read_excel_data, match_filename_with_excel, read_name_id_mapping
from pdf_handler import PDFHandler
from office_converter import OfficeConverter


class PDFEditorGUI:
    """PDF编辑器GUI主类"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("PDF操作工具 by XuXuQuan")
        
        # 获取屏幕尺寸并设置窗口大小为屏幕的80%
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        window_width = int(screen_width * 0.8)
        window_height = int(screen_height * 0.8)
        
        # 设置最小窗口大小
        self.root.minsize(800, 600)
        self.root.geometry(f"{window_width}x{window_height}")
        
        # 设置默认字体
        self.default_font = ('微软雅黑', 10)
        
        # 初始化变量
        self.init_variables()
        
        # 创建主界面
        self.create_main_frame()
        
        # 创建功能选择区
        self.create_function_selector()
        
        # 创建功能框架容器
        self.create_function_frames()
        
        # 创建状态栏
        self.create_status_bar()
        
        # 初始化显示默认功能
        self.on_function_change()
    
    def init_variables(self):
        """初始化所有变量"""
        # 功能选择
        self.function_var = tk.StringVar(value="insert")
        
        # PDF插入变量
        self.insert_pdf = ""
        self.insert_pages_var = tk.StringVar(value="1")
        self.interval_var = tk.StringVar(value="1")
        self.mode_var = tk.StringVar(value="single")
        self.multi_count_var = tk.StringVar(value="1")
        self.insert_output_dir = ""
        
        # 子PDF路径列表（支持动态添加）
        self.child_pdf_vars = []  # 存储子PDF路径的StringVar
        self.child_pdf_frames = []  # 存储子PDF控件框架
        
        # PDF分割变量
        self.split_pdf = ""
        self.split_pages_var = tk.StringVar(value="10")
        self.naming_var = tk.StringVar(value="number")
        self.excel_file = ""
        
        # PDF页面交换变量
        self.swap_pdfs = []
        self.swap_pdf_pages = {}  # 存储每个PDF的页数 {路径: 页数}
        self.page_a_var = tk.StringVar(value="1")
        self.page_b_var = tk.StringVar(value="2")
        
        # 指定导出页变量（支持批量处理）
        self.export_pdfs = []
        self.export_pages_var = tk.StringVar(value="1-6,7-10")
        self.export_excel_file = ""
        self.export_name_header = None
        self.export_id_header = None
        self.export_use_original_name_var = tk.BooleanVar(value=False)
        
        # PDF重命名变量
        self.rename_pdfs = []
        self.rename_mode_var = tk.StringVar(value="prefix")
        self.prefix_var = tk.StringVar(value="")
        self.suffix_var = tk.StringVar(value="")
        self.class_rename_special_char_var = tk.StringVar(value="_")
        self.filename_parts = []
        self.match_fields = []
        
        # PDF重排序变量
        self.reorder_pdfs = []
        self.reorder_excel_file = ""
        self.reorder_name_header = None
        self.reorder_id_header = None
        
        # PDF页面变量
        self.page_pdfs = []
        self.page_rotation_scheme = []
        
        # 批量提取变量
        self.batch_extract_pdfs = []
        self.batch_extract_page_var = tk.StringVar(value="2")
        self.batch_extract_output_dir = ""
        self.batch_extract_excel_files = []
        self.batch_extract_name_header = None
        self.batch_extract_class_header = None
        self.batch_extract_extracted_names = {}
        
        # 批量追加页变量
        self.batch_append_pdfs = []
        self.batch_append_source_pdf = ""
        self.batch_append_output_dir = ""
        self.batch_append_excel_files = []
        self.batch_append_name_header = None
        self.batch_append_class_header = None
        self.batch_append_extracted_names = {}
        
        # PDF合并变量
        self.merge_main_pdf = ""
        self.merge_append_pdfs = []  # 存储要追加的PDF路径
        self.merge_append_pdf_vars = []  # 存储对应的StringVar
        self.merge_append_frames = []  # 存储对应的框架
        self.merge_output_dir = ""
        
        # 批量单页追加变量
        self.batch_single_main_pdfs = []  # 主PDF文件列表
        self.batch_single_append_pdfs = []  # 要追加的PDF文件列表
        self.batch_single_page_var = tk.StringVar(value="1")  # 要提取的页码
        self.batch_single_output_dir = ""
        
        # PDF页面逆序变量
        self.reverse_pdfs = []  # 要逆序的PDF文件列表
        self.reverse_output_dir = ""
        
        # 交叉合并变量
        self.cross_pdf_a = ""  # PDF_A（正面扫描）
        self.cross_pdf_b = ""  # PDF_B（背面扫描，已逆序）
        self.cross_pages_a = tk.StringVar(value="1")  # PDF_A每次取页数
        self.cross_pages_b = tk.StringVar(value="1")  # PDF_B每次取页数
        self.cross_copies = tk.StringVar(value="79")  # 合并份数
        self.cross_output_dir = ""
        
        # PDF_A PDF_B合并变量
        self.student_pdfs = []
        self.class_pdfs = []
        self.student_class_excel_file = ""
        self.student_class_name_header = None
        self.student_class_class_header = None
        self.student_class_output_dir = ""
        
        # Excel批量导出PDF变量
        self.excel_files = []
        self.excel_output_dir = ""
        self.excel_selected_sheets = []
        self.excel_print_active_sheet_var = tk.BooleanVar(value=False)
        self.excel_paper_size_var = tk.StringVar(value="A4")
        self.excel_first_page_var = tk.StringVar(value="")
        self.excel_last_page_var = tk.StringVar(value="")
        
        # Word批量导出PDF变量
        self.word_files = []
        self.word_output_dir = ""
        self.word_export_all_pages_var = tk.BooleanVar(value=True)
        self.word_page_range_var = tk.StringVar(value="")
        
        # PDF预览相关变量
        self.preview_pdf_path = ""
        self.preview_current_page = 0
        self.preview_total_pages = 0
        
        # 状态变量
        self.status_var = tk.StringVar(value="就绪")
        
        # 功能框架引用
        self.insert_frame = None
        self.split_frame = None
        self.swap_frame = None
        self.export_frame = None
        self.rename_frame = None
        self.reorder_frame = None
        self.page_frame = None
        self.reverse_frame = None
        self.batch_extract_frame = None
        self.batch_append_frame = None
        self.merge_frame = None
        self.batch_single_append_frame = None
        self.cross_merge_frame = None
        self.student_class_frame = None
        self.excel_to_pdf_frame = None
        self.word_to_pdf_frame = None
    
    def create_main_frame(self):
        """创建主框架"""
        # 创建主容器
        self.main_container = ttk.Frame(self.root)
        self.main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 创建滚动区域
        self.canvas = tk.Canvas(self.main_container)
        self.scrollbar = ttk.Scrollbar(self.main_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
    
    def create_function_selector(self):
        """创建功能选择区域"""
        function_frame = ttk.LabelFrame(self.scrollable_frame, text="功能选择", padding="10")
        function_frame.pack(fill=tk.X, pady=5)
        
        functions = [
            ("PDF插入", "insert"),
            ("PDF分割", "split"),
            ("PDF页面交换", "swap"),
            ("指定导出页", "export"),
            ("PDF批量重命名", "rename"),
            ("PDF重排序", "reorder"),
            ("PDF页面", "page"),
            ("PDF页面逆序", "reverse"),
            ("批量提取", "batch_extract"),
            ("批量追加页", "batch_append"),
            ("PDF合并", "merge"),
            ("批量单页追加", "batch_single_append"),
            ("交叉合并", "cross_merge"),
            ("PDF_A与PDF_B合并", "student_class"),
            ("Excel批量导出PDF", "excel_to_pdf"),
            ("Word批量导出PDF", "word_to_pdf"),
        ]
        
        # 创建第一行功能按钮
        row1_frame = ttk.Frame(function_frame)
        row1_frame.pack(fill=tk.X, pady=3)
        
        # 创建第二行功能按钮
        row2_frame = ttk.Frame(function_frame)
        row2_frame.pack(fill=tk.X, pady=3)
        
        # 每行显示6个功能
        for i, (text, value) in enumerate(functions):
            radio = ttk.Radiobutton(
                row1_frame if i < 6 else row2_frame, 
                text=text, 
                variable=self.function_var, 
                value=value, 
                command=self.on_function_change
            )
            radio.pack(side=tk.LEFT, padx=8)
        
        # 使用说明按钮（单独一行）
        help_frame = ttk.Frame(function_frame)
        help_frame.pack(fill=tk.X, pady=3)
        help_btn = ttk.Button(help_frame, text="使用说明", command=self.show_help)
        help_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_function_frames(self):
        """创建各个功能的设置框架"""
        main_frame = ttk.Frame(self.scrollable_frame)
        main_frame.pack(fill=tk.X, pady=5)
        
        # PDF插入设置
        self.insert_frame = ttk.LabelFrame(main_frame, text="PDF插入设置", padding="10")
        
        # PDF分割设置
        self.split_frame = ttk.LabelFrame(main_frame, text="PDF分割设置", padding="10")
        
        # PDF页面交换设置
        self.swap_frame = ttk.LabelFrame(main_frame, text="PDF页面交换设置", padding="10")
        
        # 指定导出页设置
        self.export_frame = ttk.LabelFrame(main_frame, text="指定导出页设置", padding="10")
        
        # PDF重命名设置
        self.rename_frame = ttk.LabelFrame(main_frame, text="PDF重命名设置", padding="10")
        
        # PDF重排序设置
        self.reorder_frame = ttk.LabelFrame(main_frame, text="PDF重排序设置", padding="10")
        
        # PDF页面设置
        self.page_frame = ttk.LabelFrame(main_frame, text="PDF页面设置", padding="10")
        
        # PDF页面逆序设置
        self.reverse_frame = ttk.LabelFrame(main_frame, text="PDF页面逆序设置", padding="10")
        
        # 交叉合并设置
        self.cross_merge_frame = ttk.LabelFrame(main_frame, text="PDF交叉合并设置", padding="10")
        
        # 批量提取设置
        self.batch_extract_frame = ttk.LabelFrame(main_frame, text="批量提取设置", padding="10")
        
        # 批量追加页设置
        self.batch_append_frame = ttk.LabelFrame(main_frame, text="批量追加页设置", padding="10")
        
        # PDF合并设置
        self.merge_frame = ttk.LabelFrame(main_frame, text="PDF合并设置", padding="10")
        
        # 批量单页追加设置
        self.batch_single_append_frame = ttk.LabelFrame(main_frame, text="批量单页追加设置", padding="10")
        
        # PDF_A与PDF_B合并设置
        self.student_class_frame = ttk.LabelFrame(main_frame, text="PDF_A与PDF_B合并设置", padding="10")
        
        # Excel批量导出PDF设置
        self.excel_to_pdf_frame = ttk.LabelFrame(main_frame, text="Excel批量导出PDF设置", padding="10")
        
        # Word批量导出PDF设置
        self.word_to_pdf_frame = ttk.LabelFrame(main_frame, text="Word批量导出PDF设置", padding="10")
        
        # 创建各功能的具体控件
        self.create_insert_widgets()
        self.create_split_widgets()
        self.create_swap_widgets()
        self.create_export_widgets()
        self.create_rename_widgets()
        self.create_reorder_widgets()
        self.create_page_widgets()
        self.create_reverse_widgets()
        self.create_cross_merge_widgets()
        self.create_batch_extract_widgets()
        self.create_batch_append_widgets()
        self.create_merge_widgets()
        self.create_batch_single_append_widgets()
        self.create_student_class_widgets()
        self.create_excel_to_pdf_widgets()
        self.create_word_to_pdf_widgets()
        
        # 执行操作按钮
        action_frame = ttk.Frame(self.scrollable_frame)
        action_frame.pack(fill=tk.X, pady=10)
        
        execute_btn = ttk.Button(action_frame, text="执行操作", command=self.process_pdf)
        execute_btn.pack(side=tk.LEFT, padx=5)
        
        clear_btn = ttk.Button(action_frame, text="清空", command=self.clear_all)
        clear_btn.pack(side=tk.LEFT, padx=5)
        
        # 设置保存/加载按钮
        settings_frame = ttk.Frame(self.scrollable_frame)
        settings_frame.pack(fill=tk.X, pady=5)
        
        load_settings_btn = ttk.Button(settings_frame, text="加载设置", command=self.load_settings)
        load_settings_btn.pack(side=tk.RIGHT, padx=5)
        
        save_settings_btn = ttk.Button(settings_frame, text="保存设置", command=self.save_settings)
        save_settings_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_insert_widgets(self):
        """创建PDF插入功能的控件"""
        # 主PDF选择
        insert_pdf_frame = ttk.Frame(self.insert_frame)
        insert_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.insert_pdf_var = tk.StringVar()
        insert_pdf_entry = ttk.Entry(insert_pdf_frame, textvariable=self.insert_pdf_var, 
                                     state="readonly", width=60)
        insert_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        insert_pdf_btn = ttk.Button(insert_pdf_frame, text="选择主PDF文件", command=self.select_insert_pdf)
        insert_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # 子PDF路径框架（支持动态添加）
        self.child_pdfs_container = ttk.LabelFrame(self.insert_frame, text="子PDF文件（可添加多个）", padding="5")
        self.child_pdfs_container.pack(fill=tk.X, pady=5)
        
        # 初始化默认3个子PDF路径控件
        for i in range(3):
            self.add_child_pdf_widget(i)
        
        # 添加/删除按钮框架
        child_pdf_btn_frame = ttk.Frame(self.insert_frame)
        child_pdf_btn_frame.pack(fill=tk.X, pady=3)
        
        add_btn = ttk.Button(child_pdf_btn_frame, text="+ 添加子PDF", command=self.add_child_pdf_widget)
        add_btn.pack(side=tk.LEFT, padx=5)
        
        remove_btn = ttk.Button(child_pdf_btn_frame, text="- 删除最后一个", command=self.remove_child_pdf_widget)
        remove_btn.pack(side=tk.LEFT, padx=5)
        
        # 插入设置
        settings_frame = ttk.Frame(self.insert_frame)
        settings_frame.pack(fill=tk.X, pady=5)
        
        # 插入间隔
        interval_label = ttk.Label(settings_frame, text="插入间隔（页）：")
        interval_label.pack(side=tk.LEFT, padx=5)
        interval_entry = ttk.Entry(settings_frame, textvariable=self.interval_var, width=5)
        interval_entry.pack(side=tk.LEFT, padx=5)
        
        # 每页插入数量
        pages_label = ttk.Label(settings_frame, text="每页插入页数：")
        pages_label.pack(side=tk.LEFT, padx=5)
        pages_entry = ttk.Entry(settings_frame, textvariable=self.insert_pages_var, width=5)
        pages_entry.pack(side=tk.LEFT, padx=5)
        
        # 模式选择
        mode_frame = ttk.Frame(self.insert_frame)
        mode_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(mode_frame, text="单页插入", variable=self.mode_var, value="single").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(mode_frame, text="批量插入", variable=self.mode_var, value="batch").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(mode_frame, text="多PDF顺序插入", variable=self.mode_var, value="multi").pack(side=tk.LEFT, padx=5)
        
        # 多PDF插入数量
        multi_frame = ttk.Frame(self.insert_frame)
        multi_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(multi_frame, text="每次插入PDF数量：").pack(side=tk.LEFT, padx=5)
        ttk.Entry(multi_frame, textvariable=self.multi_count_var, width=5).pack(side=tk.LEFT, padx=5)
        
        # 输出目录
        output_frame = ttk.Frame(self.insert_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.insert_output_var = tk.StringVar()
        insert_output_entry = ttk.Entry(output_frame, textvariable=self.insert_output_var, 
                                        state="readonly", width=60)
        insert_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        insert_output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_insert_output_dir)
        insert_output_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_split_widgets(self):
        """创建PDF分割功能的控件"""
        # PDF文件选择
        split_pdf_frame = ttk.Frame(self.split_frame)
        split_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.split_pdf_var = tk.StringVar()
        split_pdf_entry = ttk.Entry(split_pdf_frame, textvariable=self.split_pdf_var, 
                                    state="readonly", width=60)
        split_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        split_pdf_btn = ttk.Button(split_pdf_frame, text="选择PDF文件", command=self.select_split_pdf)
        split_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # 分割页数
        split_pages_frame = ttk.Frame(self.split_frame)
        split_pages_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(split_pages_frame, text="分割页数：").pack(side=tk.LEFT, padx=5)
        ttk.Entry(split_pages_frame, textvariable=self.split_pages_var, width=10).pack(side=tk.LEFT, padx=5)
        
        # 命名方式
        naming_frame = ttk.Frame(self.split_frame)
        naming_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(naming_frame, text="数字编号", variable=self.naming_var, value="number").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(naming_frame, text="Excel命名", variable=self.naming_var, value="excel").pack(side=tk.LEFT, padx=5)
        
        # Excel文件选择（仅在Excel命名时显示）
        excel_frame = ttk.Frame(self.split_frame)
        excel_frame.pack(fill=tk.X, pady=5)
        
        self.split_excel_var = tk.StringVar()
        split_excel_entry = ttk.Entry(excel_frame, textvariable=self.split_excel_var, 
                                       state="readonly", width=50)
        split_excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        split_excel_btn = ttk.Button(excel_frame, text="选择Excel文件", command=self.select_split_excel)
        split_excel_btn.pack(side=tk.RIGHT, padx=5)
        
        # 输出目录
        output_frame = ttk.Frame(self.split_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.split_output_var = tk.StringVar()
        split_output_entry = ttk.Entry(output_frame, textvariable=self.split_output_var, 
                                        state="readonly", width=60)
        split_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        split_output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_split_output_dir)
        split_output_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_swap_widgets(self):
        """创建PDF页面交换功能的控件"""
        # PDF文件选择
        swap_pdf_frame = ttk.Frame(self.swap_frame)
        swap_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.swap_pdf_var = tk.StringVar()
        swap_pdf_entry = ttk.Entry(swap_pdf_frame, textvariable=self.swap_pdf_var, 
                                    state="readonly", width=60)
        swap_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        swap_pdf_btn = ttk.Button(swap_pdf_frame, text="选择PDF文件（可多选）", command=self.select_swap_pdfs)
        swap_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # PDF页数信息显示
        self.swap_pages_info_var = tk.StringVar(value="")
        swap_pages_info_label = ttk.Label(self.swap_frame, textvariable=self.swap_pages_info_var,
                                           font=('Arial', 9), foreground='#666666')
        swap_pages_info_label.pack(anchor=tk.W, padx=5, pady=2)
        
        # 页面顺序编辑按钮
        edit_btn_frame = ttk.Frame(self.swap_frame)
        edit_btn_frame.pack(fill=tk.X, pady=5)
        
        edit_order_btn = ttk.Button(edit_btn_frame, text="打开交换方案编辑窗口", command=self.open_swap_order_editor)
        edit_order_btn.pack(side=tk.LEFT, padx=5)
        
        # 当前页面顺序显示
        self.swap_order_var = tk.StringVar(value="当前顺序：未设置")
        order_label = ttk.Label(edit_btn_frame, textvariable=self.swap_order_var, 
                                font=('Arial', 10), foreground='#336699')
        order_label.pack(side=tk.LEFT, padx=10)
        
        # 输出目录
        output_frame = ttk.Frame(self.swap_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.swap_output_var = tk.StringVar()
        swap_output_entry = ttk.Entry(output_frame, textvariable=self.swap_output_var, 
                                       state="readonly", width=60)
        swap_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        swap_output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_swap_output_dir)
        swap_output_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_export_widgets(self):
        """创建指定导出页功能的控件"""
        # PDF文件选择（支持批量）
        export_pdf_frame = ttk.LabelFrame(self.export_frame, text="PDF文件（支持批量选择）", padding="5")
        export_pdf_frame.pack(fill=tk.X, pady=5)
        
        export_pdf_content = ttk.Frame(export_pdf_frame)
        export_pdf_content.pack(fill=tk.X, padx=5, pady=5)
        
        self.export_pdf_var = tk.StringVar(value="未选择文件")
        export_pdf_entry = ttk.Entry(export_pdf_content, textvariable=self.export_pdf_var, 
                                      state="readonly", width=60)
        export_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        export_pdf_btn = ttk.Button(export_pdf_content, text="选择文件（可多选）", command=self.select_export_pdf)
        export_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # PDF数量显示
        self.export_pdf_count_var = tk.StringVar(value="")
        export_pdf_count = ttk.Label(export_pdf_frame, textvariable=self.export_pdf_count_var, 
                                      font=('Arial', 9), foreground='#666666')
        export_pdf_count.pack(side=tk.LEFT, padx=10, pady=2, anchor=tk.W)
        
        # 导出页面设置
        export_pages_frame = ttk.Frame(self.export_frame)
        export_pages_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(export_pages_frame, text="导出页面范围（如：1-6,7-10，每组导出一个多页PDF）:").pack(side=tk.LEFT, padx=5)
        ttk.Entry(export_pages_frame, textvariable=self.export_pages_var, width=30).pack(side=tk.LEFT, padx=5)
        
        # 输出选项
        export_options_frame = ttk.Frame(self.export_frame)
        export_options_frame.pack(fill=tk.X, pady=5)
        
        ttk.Checkbutton(export_options_frame, text="以原PDF文件名输出", 
                        variable=self.export_use_original_name_var).pack(side=tk.LEFT, padx=5)
        
        # Excel文件选择（可选）
        export_excel_frame = ttk.Frame(self.export_frame)
        export_excel_frame.pack(fill=tk.X, pady=5)
        
        self.export_excel_var = tk.StringVar()
        export_excel_entry = ttk.Entry(export_excel_frame, textvariable=self.export_excel_var, 
                                        state="readonly", width=50)
        export_excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        export_excel_btn = ttk.Button(export_excel_frame, text="选择Excel文件", command=self.select_export_excel_file)
        export_excel_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_rename_widgets(self):
        """创建PDF重命名功能的控件"""
        # PDF文件选择
        rename_pdf_frame = ttk.Frame(self.rename_frame)
        rename_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.rename_pdf_var = tk.StringVar()
        rename_pdf_entry = ttk.Entry(rename_pdf_frame, textvariable=self.rename_pdf_var, 
                                      state="readonly", width=60)
        rename_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        rename_pdf_btn = ttk.Button(rename_pdf_frame, text="选择PDF文件（可多选）", command=self.select_rename_pdfs)
        rename_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # 重命名模式
        mode_frame = ttk.Frame(self.rename_frame)
        mode_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(mode_frame, text="前缀", variable=self.rename_mode_var, value="prefix").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(mode_frame, text="后缀", variable=self.rename_mode_var, value="suffix").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(mode_frame, text="按班级", variable=self.rename_mode_var, value="class").pack(side=tk.LEFT, padx=5)
        
        # 前缀/后缀输入
        text_frame = ttk.Frame(self.rename_frame)
        text_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(text_frame, text="前缀：").pack(side=tk.LEFT, padx=5)
        ttk.Entry(text_frame, textvariable=self.prefix_var, width=20).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(text_frame, text="后缀：").pack(side=tk.LEFT, padx=5)
        ttk.Entry(text_frame, textvariable=self.suffix_var, width=20).pack(side=tk.LEFT, padx=5)
        
        # 输出目录
        output_frame = ttk.Frame(self.rename_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.rename_output_var = tk.StringVar()
        rename_output_entry = ttk.Entry(output_frame, textvariable=self.rename_output_var, 
                                         state="readonly", width=60)
        rename_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        rename_output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_rename_output_dir)
        rename_output_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_reorder_widgets(self):
        """创建PDF重排序功能的控件"""
        # PDF文件选择
        reorder_pdf_frame = ttk.Frame(self.reorder_frame)
        reorder_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.reorder_pdf_var = tk.StringVar()
        reorder_pdf_entry = ttk.Entry(reorder_pdf_frame, textvariable=self.reorder_pdf_var, 
                                       state="readonly", width=50)
        reorder_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        reorder_pdf_btn = ttk.Button(reorder_pdf_frame, text="选择文件", command=self.select_reorder_pdfs)
        reorder_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # Excel文件选择（可选）
        reorder_excel_frame = ttk.Frame(self.reorder_frame)
        reorder_excel_frame.pack(fill=tk.X, pady=5)
        
        self.reorder_excel_var = tk.StringVar()
        reorder_excel_entry = ttk.Entry(reorder_excel_frame, textvariable=self.reorder_excel_var, 
                                        state="readonly", width=50)
        reorder_excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        reorder_excel_btn = ttk.Button(reorder_excel_frame, text="选择Excel文件", command=self.select_reorder_excel_file)
        reorder_excel_btn.pack(side=tk.RIGHT, padx=5)
        
        # 姓名表头选择
        reorder_name_header_frame = ttk.Frame(self.reorder_frame)
        reorder_name_header_frame.pack(fill=tk.X, pady=5)
        
        self.reorder_name_header_var = tk.StringVar(value="未选择")
        reorder_name_header_entry = ttk.Entry(reorder_name_header_frame, textvariable=self.reorder_name_header_var, 
                                               state="readonly", width=50)
        reorder_name_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        reorder_name_header_btn = ttk.Button(reorder_name_header_frame, text="选择姓名表头", command=self.select_reorder_name_header)
        reorder_name_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 序号表头选择
        reorder_id_header_frame = ttk.Frame(self.reorder_frame)
        reorder_id_header_frame.pack(fill=tk.X, pady=5)
        
        self.reorder_id_header_var = tk.StringVar(value="未选择")
        reorder_id_header_entry = ttk.Entry(reorder_id_header_frame, textvariable=self.reorder_id_header_var, 
                                             state="readonly", width=50)
        reorder_id_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        reorder_id_header_btn = ttk.Button(reorder_id_header_frame, text="选择序号表头", command=self.select_reorder_id_header)
        reorder_id_header_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_page_widgets(self):
        """创建PDF页面功能的控件"""
        # PDF文件选择
        page_pdf_frame = ttk.Frame(self.page_frame)
        page_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.page_pdf_var = tk.StringVar()
        page_pdf_entry = ttk.Entry(page_pdf_frame, textvariable=self.page_pdf_var, 
                                    state="readonly", width=60)
        page_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        page_pdf_btn = ttk.Button(page_pdf_frame, text="选择PDF文件（可多选）", command=self.select_page_pdfs)
        page_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # PDF预览按钮
        page_preview_btn = ttk.Button(page_pdf_frame, text="预览PDF", command=self.preview_page_pdf)
        page_preview_btn.pack(side=tk.RIGHT, padx=5)
        
        # 操作选择
        action_frame = ttk.Frame(self.page_frame)
        action_frame.pack(fill=tk.X, pady=5)
        
        self.page_action_var = tk.StringVar(value="reverse")
        ttk.Radiobutton(action_frame, text="反转页面顺序", variable=self.page_action_var, value="reverse").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(action_frame, text="旋转页面", variable=self.page_action_var, value="rotate").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(action_frame, text="页面交换", variable=self.page_action_var, value="swap").pack(side=tk.LEFT, padx=5)
        
        # 旋转角度设置
        rotate_frame = ttk.Frame(self.page_frame)
        rotate_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(rotate_frame, text="旋转角度（度）：").pack(side=tk.LEFT, padx=5)
        self.rotate_angle_var = tk.StringVar(value="90")
        ttk.Entry(rotate_frame, textvariable=self.rotate_angle_var, width=10).pack(side=tk.LEFT, padx=5)
        
        # 页面交换设置
        swap_frame = ttk.Frame(self.page_frame)
        swap_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(swap_frame, text="交换页码：").pack(side=tk.LEFT, padx=5)
        self.swap_page_a_var = tk.StringVar(value="1")
        ttk.Entry(swap_frame, textvariable=self.swap_page_a_var, width=5).pack(side=tk.LEFT, padx=5)
        ttk.Label(swap_frame, text=" ↔ ").pack(side=tk.LEFT, padx=5)
        self.swap_page_b_var = tk.StringVar(value="2")
        ttk.Entry(swap_frame, textvariable=self.swap_page_b_var, width=5).pack(side=tk.LEFT, padx=5)
        
        # 输出目录
        output_frame = ttk.Frame(self.page_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.page_output_var = tk.StringVar()
        page_output_entry = ttk.Entry(output_frame, textvariable=self.page_output_var, 
                                       state="readonly", width=60)
        page_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        page_output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_page_output_dir)
        page_output_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_reverse_widgets(self):
        """创建PDF页面逆序功能的控件"""
        # PDF文件选择
        reverse_pdf_frame = ttk.Frame(self.reverse_frame)
        reverse_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.reverse_pdf_var = tk.StringVar()
        reverse_pdf_entry = ttk.Entry(reverse_pdf_frame, textvariable=self.reverse_pdf_var, 
                                     state="readonly", width=60)
        reverse_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        reverse_pdf_btn = ttk.Button(reverse_pdf_frame, text="选择PDF文件（可多选）", command=self.select_reverse_pdfs)
        reverse_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # 输出目录
        output_frame = ttk.Frame(self.reverse_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.reverse_output_var = tk.StringVar()
        reverse_output_entry = ttk.Entry(output_frame, textvariable=self.reverse_output_var, 
                                         state="readonly", width=60)
        reverse_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        reverse_output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_reverse_output_dir)
        reverse_output_btn.pack(side=tk.RIGHT, padx=5)
        
        # 使用说明
        note_frame = ttk.LabelFrame(self.reverse_frame, text="使用说明", padding="5")
        note_frame.pack(fill=tk.X, pady=5)
        
        note_text = """• 选择一个或多个PDF文件，程序将反转其页面顺序
• 输出文件名格式：原文件名_reversed.pdf"""
        note_label = ttk.Label(note_frame, text=note_text, font=('微软雅黑', 9), justify=tk.LEFT)
        note_label.pack(padx=5, pady=5)
    
    def create_cross_merge_widgets(self):
        """创建PDF交叉合并功能的控件"""
        # PDF_A选择（正面扫描）
        pdf_a_frame = ttk.Frame(self.cross_merge_frame)
        pdf_a_frame.pack(fill=tk.X, pady=5)
        
        self.cross_pdf_a_var = tk.StringVar()
        pdf_a_entry = ttk.Entry(pdf_a_frame, textvariable=self.cross_pdf_a_var, 
                                state="readonly", width=60)
        pdf_a_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        pdf_a_btn = ttk.Button(pdf_a_frame, text="选择PDF_A（正面扫描）", command=self.select_cross_pdf_a)
        pdf_a_btn.pack(side=tk.RIGHT, padx=5)
        
        # PDF_B选择（背面扫描，已逆序）
        pdf_b_frame = ttk.Frame(self.cross_merge_frame)
        pdf_b_frame.pack(fill=tk.X, pady=5)
        
        self.cross_pdf_b_var = tk.StringVar()
        pdf_b_entry = ttk.Entry(pdf_b_frame, textvariable=self.cross_pdf_b_var, 
                                state="readonly", width=60)
        pdf_b_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        pdf_b_btn = ttk.Button(pdf_b_frame, text="选择PDF_B（背面扫描，已逆序）", command=self.select_cross_pdf_b)
        pdf_b_btn.pack(side=tk.RIGHT, padx=5)
        
        # 取页设置
        pages_frame = ttk.Frame(self.cross_merge_frame)
        pages_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(pages_frame, text="PDF_A每次取页数：").pack(side=tk.LEFT, padx=5)
        ttk.Entry(pages_frame, textvariable=self.cross_pages_a, width=5).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(pages_frame, text="PDF_B每次取页数：").pack(side=tk.LEFT, padx=5)
        ttk.Entry(pages_frame, textvariable=self.cross_pages_b, width=5).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(pages_frame, text="合并份数：").pack(side=tk.LEFT, padx=5)
        ttk.Entry(pages_frame, textvariable=self.cross_copies, width=5).pack(side=tk.LEFT, padx=5)
        
        # 输出目录
        output_frame = ttk.Frame(self.cross_merge_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.cross_output_var = tk.StringVar()
        cross_output_entry = ttk.Entry(output_frame, textvariable=self.cross_output_var, 
                                       state="readonly", width=60)
        cross_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        cross_output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_cross_output_dir)
        cross_output_btn.pack(side=tk.RIGHT, padx=5)
        
        # 预览按钮
        preview_btn = ttk.Button(self.cross_merge_frame, text="预览合并方案", command=self.preview_cross_merge)
        preview_btn.pack(fill=tk.X, pady=5)
        
        # 使用说明
        note_frame = ttk.LabelFrame(self.cross_merge_frame, text="使用说明", padding="5")
        note_frame.pack(fill=tk.X, pady=5)
        
        note_text = """• PDF_A：正面扫描件（如封面正面、申请表正面）
• PDF_B：背面扫描件，需先逆序处理（如封面背面、申请表背面）
• 交叉合并模式：A取N页 → B取M页 → A取N页 → B取M页...
• 示例（79人档案）：
  - PDF_A每次取1页，PDF_B每次取1页，共79份
  - 输出：封面 → 封面背面 → 申请表正面 → 申请表背面（每人4页）"""
        note_label = ttk.Label(note_frame, text=note_text, font=('微软雅黑', 9), justify=tk.LEFT)
        note_label.pack(padx=5, pady=5)
    
    def create_batch_extract_widgets(self):
        """创建批量提取功能的控件"""
        # PDF文件选择
        extract_pdf_frame = ttk.Frame(self.batch_extract_frame)
        extract_pdf_frame.pack(fill=tk.X, pady=5)
        
        self.extract_pdf_var = tk.StringVar()
        extract_pdf_entry = ttk.Entry(extract_pdf_frame, textvariable=self.extract_pdf_var, 
                                       state="readonly", width=60)
        extract_pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        extract_pdf_btn = ttk.Button(extract_pdf_frame, text="选择PDF文件（可多选）", command=self.select_batch_extract_pdfs)
        extract_pdf_btn.pack(side=tk.RIGHT, padx=5)
        
        # 提取页码
        page_frame = ttk.Frame(self.batch_extract_frame)
        page_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(page_frame, text="提取页码：").pack(side=tk.LEFT, padx=5)
        ttk.Entry(page_frame, textvariable=self.batch_extract_page_var, width=10).pack(side=tk.LEFT, padx=5)
        
        # 输出目录
        output_frame = ttk.Frame(self.batch_extract_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.extract_output_var = tk.StringVar()
        extract_output_entry = ttk.Entry(output_frame, textvariable=self.extract_output_var, 
                                          state="readonly", width=60)
        extract_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        extract_output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_batch_extract_output_dir)
        extract_output_btn.pack(side=tk.RIGHT, padx=5)
        
        # Excel匹配设置（可选）
        excel_frame = ttk.LabelFrame(self.batch_extract_frame, text="Excel匹配设置（可选）", padding="5")
        excel_frame.pack(fill=tk.X, pady=5)
        
        # Excel文件选择
        excel_file_frame = ttk.Frame(excel_frame)
        excel_file_frame.pack(fill=tk.X, pady=5)
        
        self.extract_excel_var = tk.StringVar()
        extract_excel_entry = ttk.Entry(excel_file_frame, textvariable=self.extract_excel_var, 
                                         state="readonly", width=50)
        extract_excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        extract_excel_btn = ttk.Button(excel_file_frame, text="选择Excel文件", command=self.select_batch_extract_excel_files)
        extract_excel_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_batch_append_widgets(self):
        """创建批量追加页功能的控件"""
        # 源PDF选择
        source_frame = ttk.Frame(self.batch_append_frame)
        source_frame.pack(fill=tk.X, pady=5)
        
        self.append_source_var = tk.StringVar()
        source_entry = ttk.Entry(source_frame, textvariable=self.append_source_var, 
                                  state="readonly", width=60)
        source_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        source_btn = ttk.Button(source_frame, text="选择源PDF文件", command=self.select_batch_append_source_pdf)
        source_btn.pack(side=tk.RIGHT, padx=5)
        
        # 目标PDF选择
        target_frame = ttk.Frame(self.batch_append_frame)
        target_frame.pack(fill=tk.X, pady=5)
        
        self.append_target_var = tk.StringVar()
        target_entry = ttk.Entry(target_frame, textvariable=self.append_target_var, 
                                  state="readonly", width=60)
        target_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        target_btn = ttk.Button(target_frame, text="选择PDF文件（可多选）", command=self.select_batch_append_pdfs)
        target_btn.pack(side=tk.RIGHT, padx=5)
        
        # 输出目录
        output_frame = ttk.Frame(self.batch_append_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.append_output_var = tk.StringVar()
        output_entry = ttk.Entry(output_frame, textvariable=self.append_output_var, 
                                  state="readonly", width=60)
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_batch_append_output_dir)
        output_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_merge_widgets(self):
        """创建PDF合并功能的控件"""
        # 主PDF选择
        main_frame = ttk.Frame(self.merge_frame)
        main_frame.pack(fill=tk.X, pady=5)
        
        self.merge_main_pdf_var = tk.StringVar()
        main_entry = ttk.Entry(main_frame, textvariable=self.merge_main_pdf_var, 
                               state="readonly", width=60)
        main_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        main_btn = ttk.Button(main_frame, text="选择主PDF文件", command=self.select_merge_main_pdf)
        main_btn.pack(side=tk.RIGHT, padx=5)
        
        # 子PDF路径框架（支持动态添加）
        self.merge_append_container = ttk.LabelFrame(self.merge_frame, text="要追加的PDF文件（可添加多个）", padding="5")
        self.merge_append_container.pack(fill=tk.X, pady=5)
        
        # 初始化默认1个子PDF路径控件
        self.add_merge_append_pdf_widget(0)
        
        # 添加/删除按钮框架
        merge_btn_frame = ttk.Frame(self.merge_frame)
        merge_btn_frame.pack(fill=tk.X, pady=3)
        
        add_btn = ttk.Button(merge_btn_frame, text="+ 添加要追加的PDF", command=self.add_merge_append_pdf_widget)
        add_btn.pack(side=tk.LEFT, padx=5)
        
        remove_btn = ttk.Button(merge_btn_frame, text="- 删除最后一个", command=self.remove_merge_append_pdf_widget)
        remove_btn.pack(side=tk.LEFT, padx=5)
        
        # 输出目录
        output_frame = ttk.Frame(self.merge_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.merge_output_var = tk.StringVar()
        output_entry = ttk.Entry(output_frame, textvariable=self.merge_output_var, 
                                 state="readonly", width=60)
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_merge_output_dir)
        output_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_batch_single_append_widgets(self):
        """创建批量单页追加功能的控件"""
        # 主PDF文件选择（可多选）
        main_frame = ttk.LabelFrame(self.batch_single_append_frame, text="主PDF文件（可多选，按顺序对应）", padding="5")
        main_frame.pack(fill=tk.X, pady=5)
        
        main_content = ttk.Frame(main_frame)
        main_content.pack(fill=tk.X, padx=5, pady=5)
        
        self.batch_single_main_var = tk.StringVar(value="未选择文件")
        main_entry = ttk.Entry(main_content, textvariable=self.batch_single_main_var, 
                               state="readonly", width=60)
        main_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        main_btn = ttk.Button(main_content, text="选择主PDF文件（可多选）", command=self.select_batch_single_main_pdfs)
        main_btn.pack(side=tk.RIGHT, padx=5)
        
        # 要追加的PDF文件选择（可多选）
        append_frame = ttk.LabelFrame(self.batch_single_append_frame, text="要追加的PDF文件（可多选）", padding="5")
        append_frame.pack(fill=tk.X, pady=5)
        
        append_content = ttk.Frame(append_frame)
        append_content.pack(fill=tk.X, padx=5, pady=5)
        
        self.batch_single_append_var = tk.StringVar(value="未选择文件")
        append_entry = ttk.Entry(append_content, textvariable=self.batch_single_append_var, 
                                 state="readonly", width=60)
        append_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        append_btn = ttk.Button(append_content, text="选择要追加的PDF文件（可多选）", command=self.select_batch_single_append_pdfs)
        append_btn.pack(side=tk.RIGHT, padx=5)
        
        # 提取页码设置
        page_frame = ttk.Frame(self.batch_single_append_frame)
        page_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(page_frame, text="提取页码：").pack(side=tk.LEFT, padx=5)
        ttk.Entry(page_frame, textvariable=self.batch_single_page_var, width=10).pack(side=tk.LEFT, padx=5)
        
        # 输出目录
        output_frame = ttk.Frame(self.batch_single_append_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.batch_single_output_var = tk.StringVar()
        output_entry = ttk.Entry(output_frame, textvariable=self.batch_single_output_var, 
                                  state="readonly", width=60)
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_batch_single_output_dir)
        output_btn.pack(side=tk.RIGHT, padx=5)
        
        # 预览按钮
        preview_frame = ttk.Frame(self.batch_single_append_frame)
        preview_frame.pack(fill=tk.X, pady=5)
        
        preview_btn = ttk.Button(preview_frame, text="预览文件列表", command=self.preview_batch_single_files)
        preview_btn.pack(side=tk.LEFT, padx=5)
        
        # 使用说明
        note_frame = ttk.LabelFrame(self.batch_single_append_frame, text="使用说明", padding="5")
        note_frame.pack(fill=tk.X, pady=5)
        
        note_text = """• 选择多个主PDF文件（如5个），再选择多个要追加的PDF文件（如79个）
• 程序会将每个追加PDF的指定页依次追加到对应的主PDF后面
• 如果主PDF数量少于追加PDF数量，会循环使用主PDF
• 输出文件名格式：主PDF文件名_追加PDF文件名.pdf"""
        note_label = ttk.Label(note_frame, text=note_text, font=('微软雅黑', 9), justify=tk.LEFT)
        note_label.pack(padx=5, pady=5)
    
    def create_student_class_widgets(self):
        """创建PDF_A与PDF_B合并功能的控件"""
        # PDF_A选择
        student_frame = ttk.Frame(self.student_class_frame)
        student_frame.pack(fill=tk.X, pady=5)
        
        self.student_pdf_var = tk.StringVar()
        student_entry = ttk.Entry(student_frame, textvariable=self.student_pdf_var, 
                                   state="readonly", width=60)
        student_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        student_btn = ttk.Button(student_frame, text="选择PDF_A文件（可多选）", command=self.select_student_pdfs)
        student_btn.pack(side=tk.RIGHT, padx=5)
        
        # PDF_B选择
        class_frame = ttk.Frame(self.student_class_frame)
        class_frame.pack(fill=tk.X, pady=5)
        
        self.class_pdf_var = tk.StringVar()
        class_entry = ttk.Entry(class_frame, textvariable=self.class_pdf_var, 
                                  state="readonly", width=60)
        class_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        class_btn = ttk.Button(class_frame, text="选择PDF_B文件（可多选）", command=self.select_class_pdfs)
        class_btn.pack(side=tk.RIGHT, padx=5)
        
        # Excel文件选择
        excel_frame = ttk.Frame(self.student_class_frame)
        excel_frame.pack(fill=tk.X, pady=5)
        
        self.student_class_excel_var = tk.StringVar()
        excel_entry = ttk.Entry(excel_frame, textvariable=self.student_class_excel_var, 
                                 state="readonly", width=50)
        excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        excel_btn = ttk.Button(excel_frame, text="选择Excel文件", command=self.select_student_class_excel)
        excel_btn.pack(side=tk.RIGHT, padx=5)
        
        # 表头选择
        headers_frame = ttk.Frame(self.student_class_frame)
        headers_frame.pack(fill=tk.X, pady=5)
        
        # 姓名表头
        name_header_frame = ttk.Frame(headers_frame)
        name_header_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.student_class_name_header_var = tk.StringVar(value="未选择")
        name_header_entry = ttk.Entry(name_header_frame, textvariable=self.student_class_name_header_var, 
                                       state="readonly")
        name_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        name_header_btn = ttk.Button(name_header_frame, text="选择姓名表头", command=self.select_student_class_name_header)
        name_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # PDF_B表头
        class_header_frame = ttk.Frame(headers_frame)
        class_header_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.student_class_class_header_var = tk.StringVar(value="未选择")
        class_header_entry = ttk.Entry(class_header_frame, textvariable=self.student_class_class_header_var, 
                                        state="readonly")
        class_header_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        class_header_btn = ttk.Button(class_header_frame, text="选择PDF_B表头", command=self.select_student_class_class_header)
        class_header_btn.pack(side=tk.RIGHT, padx=5)
        
        # 输出目录
        output_frame = ttk.Frame(self.student_class_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.student_class_output_var = tk.StringVar()
        output_entry = ttk.Entry(output_frame, textvariable=self.student_class_output_var, 
                                   state="readonly", width=60)
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_student_class_output_dir)
        output_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_excel_to_pdf_widgets(self):
        """创建Excel批量导出PDF功能的控件"""
        # Excel文件选择
        excel_frame = ttk.Frame(self.excel_to_pdf_frame)
        excel_frame.pack(fill=tk.X, pady=5)
        
        self.excel_files_var = tk.StringVar(value="未选择文件")
        excel_entry = ttk.Entry(excel_frame, textvariable=self.excel_files_var, 
                                 state="readonly", width=60)
        excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        excel_btn = ttk.Button(excel_frame, text="选择Excel文件（可多选）", command=self.select_excel_files)
        excel_btn.pack(side=tk.RIGHT, padx=5)
        
        # PDF预览按钮
        preview_btn = ttk.Button(excel_frame, text="预览PDF效果", command=self.preview_excel_to_pdf)
        preview_btn.pack(side=tk.RIGHT, padx=5)
        
        # 打印设置
        print_settings_frame = ttk.LabelFrame(self.excel_to_pdf_frame, text="打印设置", padding="5")
        print_settings_frame.pack(fill=tk.X, pady=5)
        
        # 打印范围设置
        print_range_frame = ttk.Frame(print_settings_frame)
        print_range_frame.pack(fill=tk.X, pady=3)
        ttk.Label(print_range_frame, text="打印范围：").pack(side=tk.LEFT, padx=5)
        ttk.Checkbutton(print_range_frame, text="仅打印活动工作表", 
                        variable=self.excel_print_active_sheet_var).pack(side=tk.LEFT, padx=5)
        
        # 页码范围
        page_range_frame = ttk.Frame(print_settings_frame)
        page_range_frame.pack(fill=tk.X, pady=3)
        ttk.Label(page_range_frame, text="页码范围：").pack(side=tk.LEFT, padx=5)
        ttk.Label(page_range_frame, text="从").pack(side=tk.LEFT, padx=2)
        ttk.Entry(page_range_frame, textvariable=self.excel_first_page_var, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(page_range_frame, text="到").pack(side=tk.LEFT, padx=2)
        ttk.Entry(page_range_frame, textvariable=self.excel_last_page_var, width=8).pack(side=tk.LEFT, padx=2)
        
        # 页面方向
        orientation_frame = ttk.Frame(print_settings_frame)
        orientation_frame.pack(fill=tk.X, pady=3)
        ttk.Label(orientation_frame, text="页面方向：").pack(side=tk.LEFT, padx=5)
        self.excel_orientation_var = tk.StringVar(value="landscape")
        ttk.Radiobutton(orientation_frame, text="横向", variable=self.excel_orientation_var, 
                        value="landscape").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(orientation_frame, text="纵向", variable=self.excel_orientation_var, 
                        value="portrait").pack(side=tk.LEFT, padx=5)
        
        # 纸张大小
        paper_size_frame = ttk.Frame(print_settings_frame)
        paper_size_frame.pack(fill=tk.X, pady=3)
        ttk.Label(paper_size_frame, text="纸张大小：").pack(side=tk.LEFT, padx=5)
        paper_size_options = ['A4', 'A3', 'Letter', 'Legal', 'B4', 'B5']
        ttk.Combobox(paper_size_frame, textvariable=self.excel_paper_size_var, 
                     values=paper_size_options, state='readonly', width=10).pack(side=tk.LEFT, padx=5)
        
        # 缩放选项
        scale_frame = ttk.Frame(print_settings_frame)
        scale_frame.pack(fill=tk.X, pady=3)
        ttk.Label(scale_frame, text="缩放方式：").pack(side=tk.LEFT, padx=5)
        self.excel_scale_mode_var = tk.StringVar(value="fit_width")
        ttk.Radiobutton(scale_frame, text="适应宽度（列不拆分）", variable=self.excel_scale_mode_var, 
                        value="fit_width").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(scale_frame, text="适应页面", variable=self.excel_scale_mode_var, 
                        value="fit_all").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(scale_frame, text="正常大小", variable=self.excel_scale_mode_var, 
                        value="normal").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(scale_frame, text="自定义缩放", variable=self.excel_scale_mode_var, 
                        value="custom").pack(side=tk.LEFT, padx=5)
        
        # 自定义缩放比例
        custom_scale_frame = ttk.Frame(print_settings_frame)
        custom_scale_frame.pack(fill=tk.X, pady=3)
        ttk.Label(custom_scale_frame, text="缩放比例：").pack(side=tk.LEFT, padx=5)
        self.excel_scale_var = tk.StringVar(value="70")
        ttk.Entry(custom_scale_frame, textvariable=self.excel_scale_var, width=10).pack(side=tk.LEFT, padx=2)
        ttk.Label(custom_scale_frame, text="%").pack(side=tk.LEFT, padx=2)
        
        # 输出目录
        output_frame = ttk.Frame(self.excel_to_pdf_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.excel_output_var = tk.StringVar()
        output_entry = ttk.Entry(output_frame, textvariable=self.excel_output_var, 
                                   state="readonly", width=60)
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_excel_output_dir)
        output_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_word_to_pdf_widgets(self):
        """创建Word批量导出PDF功能的控件"""
        # Word文件选择
        word_frame = ttk.Frame(self.word_to_pdf_frame)
        word_frame.pack(fill=tk.X, pady=5)
        
        self.word_files_var = tk.StringVar(value="未选择文件")
        word_entry = ttk.Entry(word_frame, textvariable=self.word_files_var, 
                                state="readonly", width=60)
        word_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        word_btn = ttk.Button(word_frame, text="选择Word文件（可多选）", command=self.select_word_files)
        word_btn.pack(side=tk.RIGHT, padx=5)
        
        # 输出目录
        output_frame = ttk.Frame(self.word_to_pdf_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.word_output_var = tk.StringVar()
        output_entry = ttk.Entry(output_frame, textvariable=self.word_output_var, 
                                  state="readonly", width=60)
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        output_btn = ttk.Button(output_frame, text="选择输出目录", command=self.select_word_output_dir)
        output_btn.pack(side=tk.RIGHT, padx=5)
        
        # 页面范围设置
        page_range_frame = ttk.Frame(self.word_to_pdf_frame)
        page_range_frame.pack(fill=tk.X, pady=5)
        
        ttk.Checkbutton(page_range_frame, text="导出全部页面", 
                        variable=self.word_export_all_pages_var).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(page_range_frame, text="页码范围（如：2-5,7）：").pack(side=tk.LEFT, padx=5)
        ttk.Entry(page_range_frame, textvariable=self.word_page_range_var, width=20).pack(side=tk.LEFT, padx=5)
    
    def create_status_bar(self):
        """创建状态栏"""
        status_frame = ttk.Frame(self.root)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=10, pady=5)
        
        ttk.Label(status_frame, textvariable=self.status_var, font=('微软雅黑', 9)).pack(side=tk.LEFT)
    
    def on_function_change(self):
        """功能切换时的处理"""
        # 隐藏所有框架
        all_frames = [
            self.insert_frame,
            self.split_frame,
            self.swap_frame,
            self.export_frame,
            self.rename_frame,
            self.reorder_frame,
            self.page_frame,
            self.reverse_frame,
            self.cross_merge_frame,
            self.batch_extract_frame,
            self.batch_append_frame,
            self.merge_frame,
            self.batch_single_append_frame,
            self.student_class_frame,
            self.excel_to_pdf_frame,
            self.word_to_pdf_frame
        ]
        
        for frame in all_frames:
            frame.pack_forget()
        
        # 显示选中的功能框架
        selected_func = self.function_var.get()
        if selected_func == "insert":
            self.insert_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "split":
            self.split_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "swap":
            self.swap_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "export":
            self.export_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "rename":
            self.rename_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "reorder":
            self.reorder_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "page":
            self.page_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "reverse":
            self.reverse_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "cross_merge":
            self.cross_merge_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "batch_extract":
            self.batch_extract_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "batch_append":
            self.batch_append_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "merge":
            self.merge_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "batch_single_append":
            self.batch_single_append_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "student_class":
            self.student_class_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "excel_to_pdf":
            self.excel_to_pdf_frame.pack(fill=tk.X, pady=5)
        elif selected_func == "word_to_pdf":
            self.word_to_pdf_frame.pack(fill=tk.X, pady=5)
    
    def process_pdf(self):
        """主处理入口"""
        self.status_var.set("处理中...")
        self.root.update()
        
        try:
            func = self.function_var.get()
            
            if func == "insert":
                self.process_insert()
            elif func == "split":
                self.process_split()
            elif func == "swap":
                self.process_swap()
            elif func == "export":
                self.process_export()
            elif func == "rename":
                self.process_rename()
            elif func == "reorder":
                self.process_reorder()
            elif func == "page":
                self.process_page()
            elif func == "reverse":
                self.process_reverse()
            elif func == "cross_merge":
                self.process_cross_merge()
            elif func == "batch_extract":
                self.process_batch_extract()
            elif func == "batch_append":
                self.process_batch_append()
            elif func == "merge":
                self.process_merge()
            elif func == "batch_single_append":
                self.process_batch_single_append()
            elif func == "student_class":
                self.process_student_class_merge()
            elif func == "excel_to_pdf":
                self.process_excel_to_pdf()
            elif func == "word_to_pdf":
                self.process_word_to_pdf()
            
            self.status_var.set("处理完成")
        except Exception as e:
            show_error(f"处理失败: {str(e)}")
            self.status_var.set("就绪")
    
    # ==================== 文件选择方法 ====================
    
    def select_insert_pdf(self):
        path = filedialog.askopenfilename(filetypes=[("PDF文件", "*.pdf")])
        if path:
            self.insert_pdf = path
            self.insert_pdf_var.set(path)
    
    def add_child_pdf_widget(self, index=None):
        """添加子PDF路径控件"""
        if index is None:
            index = len(self.child_pdf_vars)
        
        frame = ttk.Frame(self.child_pdfs_container)
        frame.pack(fill=tk.X, pady=2)
        
        var = tk.StringVar()
        entry = ttk.Entry(frame, textvariable=var, state="readonly", width=50)
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        btn = ttk.Button(frame, text=f"选择子PDF {index + 1}", 
                         command=lambda v=var: self.select_child_pdf(v))
        btn.pack(side=tk.RIGHT, padx=5)
        
        self.child_pdf_vars.append(var)
        self.child_pdf_frames.append(frame)
    
    def remove_child_pdf_widget(self):
        """删除最后一个子PDF路径控件"""
        if len(self.child_pdf_frames) > 1:  # 至少保留1个
            frame = self.child_pdf_frames.pop()
            frame.destroy()
            self.child_pdf_vars.pop()
    
    def select_child_pdf(self, var):
        """选择子PDF文件"""
        path = filedialog.askopenfilename(filetypes=[("PDF文件", "*.pdf")])
        if path:
            var.set(path)
    
    def select_insert_pages(self):
        # 保持兼容，使用第一个子PDF控件
        if self.child_pdf_vars:
            path = filedialog.askopenfilename(filetypes=[("PDF文件", "*.pdf")])
            if path:
                self.child_pdf_vars[0].set(path)
    
    def select_insert_output_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.insert_output_dir = path
            self.insert_output_var.set(path)
    
    def select_split_pdf(self):
        path = filedialog.askopenfilename(filetypes=[("PDF文件", "*.pdf")])
        if path:
            self.split_pdf = path
            self.split_pdf_var.set(path)
    
    def select_split_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx;*.xls")])
        if path:
            self.excel_file = path
            self.split_excel_var.set(path)
    
    def select_split_output_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.split_output_dir = path
            self.split_output_var.set(path)
    
    def select_swap_pdfs(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF文件", "*.pdf")])
        if paths:
            self.swap_pdfs = list(paths)
            self.swap_pdf_var.set(f"已选择 {len(paths)} 个文件")
            
            # 读取每个PDF的页数信息
            pages_info = []
            self.swap_pdf_pages = {}
            max_pages = 0
            for pdf_path in self.swap_pdfs:
                page_count = PDFHandler.get_page_count(pdf_path)
                self.swap_pdf_pages[pdf_path] = page_count
                file_name = os.path.basename(pdf_path)
                pages_info.append(f"{file_name} ({page_count}页)")
                if page_count > max_pages:
                    max_pages = page_count
            
            if pages_info:
                self.swap_pages_info_var.set("；".join(pages_info))
            else:
                self.swap_pages_info_var.set("")
            
            # 初始化页面顺序（默认按原始顺序）
            self.swap_page_order = list(range(1, max_pages + 1))
            self.swap_order_var.set(f"当前顺序：{', '.join(map(str, self.swap_page_order))}")
    
    def select_swap_output_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.swap_output_dir = path
            self.swap_output_var.set(path)
    
    def open_swap_order_editor(self):
        """打开页面顺序编辑窗口"""
        if not self.swap_pdfs:
            show_error("请先选择PDF文件")
            return
        
        # 获取最大页数
        max_pages = max(self.swap_pdf_pages.values()) if self.swap_pdf_pages else 0
        if max_pages == 0:
            show_error("无法读取PDF页数")
            return
        
        # 如果超过10页，弹出提示
        if max_pages > 10:
            show_info(f"页面顺序超过10页（共{max_pages}页）\n非全屏模式下只显示前10页，点击全屏按钮可查看全部")
        
        # 创建编辑窗口（放大1.5倍）
        editor_window = tk.Toplevel(self.root)
        editor_window.title("页面顺序编辑")
        editor_window.geometry("750x675")  # 500x450 * 1.5
        editor_window.resizable(True, True)
        
        # 当前是否全屏
        is_fullscreen = [False]
        # 当前显示模式（全部/前10页）
        show_all = [max_pages <= 10]
        
        # 初始化页面顺序
        if not hasattr(self, 'swap_page_order') or len(self.swap_page_order) != max_pages:
            self.swap_page_order = list(range(1, max_pages + 1))
        
        # 预加载PDF缩略图（保持引用防止被垃圾回收）
        page_images = []
        page_orientations = []  # 记录每页的方向
        first_pdf = self.swap_pdfs[0] if self.swap_pdfs else None
        
        def load_page_thumbnails():
            if not first_pdf:
                return
            
            try:
                doc = fitz.open(first_pdf)
                for page_num in range(1, min(max_pages + 1, 51)):
                    if page_num <= len(doc):
                        try:
                            page = doc.load_page(page_num - 1)
                            # 检测页面方向
                            page_rect = page.rect
                            is_landscape = page_rect.width > page_rect.height
                            page_orientations.append(is_landscape)
                            pix = page.get_pixmap(matrix=fitz.Matrix(0.2, 0.2))
                            img_data = pix.tobytes("ppm")
                            photo = tk.PhotoImage(data=img_data)
                            page_images.append(photo)
                        except Exception as e:
                            page_images.append(None)
                            page_orientations.append(False)
                doc.close()
            except Exception as e:
                print(f"加载PDF缩略图失败: {e}")
        
        load_page_thumbnails()
        
        # 页面顺序显示区域（带滚动条）
        display_frame = ttk.Frame(editor_window, padding="10")
        display_frame.pack(fill=tk.BOTH, expand=True)
        
        # 滚动条
        v_scrollbar = ttk.Scrollbar(display_frame, orient=tk.VERTICAL)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        h_scrollbar = ttk.Scrollbar(display_frame, orient=tk.HORIZONTAL)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 页面画布区域
        pages_canvas = tk.Canvas(display_frame, bg="#f5f5f5", relief=tk.SUNKEN,
                                 yscrollcommand=v_scrollbar.set,
                                 xscrollcommand=h_scrollbar.set)
        pages_canvas.pack(fill=tk.BOTH, expand=True, pady=10)
        
        v_scrollbar.config(command=pages_canvas.yview)
        h_scrollbar.config(command=pages_canvas.xview)
        
        # 选中的页面索引
        selected_index = [0]
        
        # 绘制页面方框
        def draw_pages():
            pages_canvas.delete("all")
            cell_width = 144  # 80 * 1.5 * 1.2
            cell_height = 180  # 100 * 1.5 * 1.2
            padding = 18  # 10 * 1.5 * 1.2
            cols = 10  # 一行显示10个页面
            
            # 确定要显示的页数
            display_pages = max_pages if show_all[0] else min(max_pages, 10)
            rows = (display_pages + cols - 1) // cols
            
            canvas_width = cols * (cell_width + padding) + padding + 40
            canvas_height = rows * (cell_height + padding) + padding + 40
            pages_canvas.config(scrollregion=(0, 0, canvas_width, canvas_height))
            
            for i in range(display_pages):
                page_num = self.swap_page_order[i]
                row = i // cols
                col = i % cols
                x = col * (cell_width + padding) + padding
                y = row * (cell_height + padding) + padding
                
                # 绘制方框背景
                bg_color = "#ffffff" if i != selected_index[0] else "#fff3cd"
                pages_canvas.create_rectangle(x, y, x + cell_width, y + cell_height,
                                             outline="#3366cc" if i != selected_index[0] else "#ff6600", 
                                             width=2 if i != selected_index[0] else 3, 
                                             fill=bg_color)
                
                # 尝试绘制PDF缩略图
                # 使用实际页码获取对应的缩略图，而不是显示位置索引
                if (page_num - 1) < len(page_images) and page_images[page_num - 1] is not None:
                    try:
                        img = page_images[page_num - 1]
                        # 计算图像位置使其居中
                        img_width = img.width()
                        img_height = img.height()
                        scale = min((cell_width - 8) / img_width, (cell_height - 30) / img_height)
                        if scale < 1:
                            # 使用 zoom 而不是 subsample，避免创建新对象导致引用丢失
                            zoom_factor = max(int(1/scale), 1)
                            # 直接显示原始图像，让 Canvas 自动缩放
                            img_x = x + (cell_width - img_width) // 2
                            img_y = y + 5
                        else:
                            img_x = x + (cell_width - img_width) // 2
                            img_y = y + 5
                        pages_canvas.create_image(img_x, img_y, image=img, anchor=tk.NW)
                    except Exception as e:
                        # 如果图像显示失败，显示页码
                        pages_canvas.create_text(x + cell_width // 2, y + cell_height // 2,
                                                 text=str(page_num), font=('Arial', 18, 'bold'),
                                                 fill="#3366cc")
                else:
                    # 显示页码
                    pages_canvas.create_text(x + cell_width // 2, y + cell_height // 2,
                                             text=str(page_num), font=('Arial', 18, 'bold'),
                                             fill="#3366cc")
                
                # 绘制顺序标记（当前位置）
                pages_canvas.create_text(x + 8, y + 8, text=f"#{i+1}", 
                                         font=('Arial', 9, 'bold'), fill="#ff6600")
                
                # 绘制原始位置标记
                original_pos = self.swap_page_order.index(page_num) + 1
                pages_canvas.create_text(x + cell_width // 2, y + cell_height - 8,
                                         text=f"原#{original_pos}", font=('Arial', 8),
                                         fill="#999999")
        
        draw_pages()
        
        # 按钮控制区域
        control_frame = ttk.Frame(editor_window, padding="10")
        control_frame.pack(fill=tk.X)
        
        ttk.Label(control_frame, text="选中页面:").pack(side=tk.LEFT, padx=5)
        selected_label = ttk.Label(control_frame, text=f"第 {self.swap_page_order[0]} 页", 
                                   font=('Arial', 12, 'bold'))
        selected_label.pack(side=tk.LEFT, padx=5)
        
        # 上下移动按钮
        btn_frame = ttk.Frame(control_frame)
        btn_frame.pack(side=tk.RIGHT)
        
        up_btn = ttk.Button(btn_frame, text="↑ 上移", command=lambda: move_page(-1))
        up_btn.pack(side=tk.LEFT, padx=5)
        
        down_btn = ttk.Button(btn_frame, text="↓ 下移", command=lambda: move_page(1))
        down_btn.pack(side=tk.LEFT, padx=5)
        
        # 全屏按钮
        fullscreen_btn = ttk.Button(btn_frame, text="全屏显示" if not is_fullscreen[0] else "退出全屏", 
                                    command=lambda: toggle_fullscreen())
        fullscreen_btn.pack(side=tk.LEFT, padx=5)
        
        def toggle_fullscreen():
            is_fullscreen[0] = not is_fullscreen[0]
            if is_fullscreen[0]:
                editor_window.attributes('-fullscreen', True)
                show_all[0] = True
                fullscreen_btn.config(text="退出全屏")
            else:
                editor_window.attributes('-fullscreen', False)
                editor_window.geometry("750x675")
                fullscreen_btn.config(text="全屏显示")
            draw_pages()
        
        def move_page(direction):
            idx = selected_index[0]
            new_idx = idx + direction
            
            # 限制在显示范围内
            display_pages = max_pages if show_all[0] else min(max_pages, 10)
            
            if 0 <= new_idx < display_pages:
                self.swap_page_order[idx], self.swap_page_order[new_idx] = \
                    self.swap_page_order[new_idx], self.swap_page_order[idx]
                selected_index[0] = new_idx
                selected_label.config(text=f"第 {self.swap_page_order[selected_index[0]]} 页")
                draw_pages()
        
        # 页码点击选择
        def on_canvas_click(event):
            cell_width = 144  # 80 * 1.5 * 1.2
            cell_height = 180  # 100 * 1.5 * 1.2
            padding = 18  # 10 * 1.5 * 1.2
            cols = 10
            
            x = pages_canvas.canvasx(event.x)
            y = pages_canvas.canvasy(event.y)
            
            col = int(x // (cell_width + padding))
            row = int(y // (cell_height + padding))
            idx = row * cols + col
            
            display_pages = max_pages if show_all[0] else min(max_pages, 10)
            
            if 0 <= idx < display_pages:
                selected_index[0] = idx
                selected_label.config(text=f"第 {self.swap_page_order[idx]} 页")
                draw_pages()
        
        pages_canvas.bind("<Button-1>", on_canvas_click)
        
        # 当前顺序显示
        order_frame = ttk.Frame(editor_window, padding="0 0 10 10")
        order_frame.pack(fill=tk.X)
        
        ttk.Label(order_frame, text="当前顺序:").pack(side=tk.LEFT, padx=10)
        order_text = tk.Text(order_frame, height=2, width=80, font=('Courier New', 10))
        order_text.pack(side=tk.LEFT, padx=5)
        
        def update_order_text():
            order_text.delete(1.0, tk.END)
            order_text.insert(tk.END, ", ".join(map(str, self.swap_page_order)))
        
        update_order_text()
        
        # 确认和取消按钮
        btn_frame2 = ttk.Frame(editor_window, padding="10")
        btn_frame2.pack(fill=tk.X)
        
        confirm_btn = ttk.Button(btn_frame2, text="确认", 
                                 command=lambda: confirm_order())
        confirm_btn.pack(side=tk.RIGHT, padx=5)
        
        cancel_btn = ttk.Button(btn_frame2, text="取消", 
                                 command=lambda: editor_window.destroy())
        cancel_btn.pack(side=tk.RIGHT, padx=5)
        
        def confirm_order():
            self.swap_order_var.set(f"当前顺序：{', '.join(map(str, self.swap_page_order))}")
            editor_window.destroy()
        
        # 定时更新顺序显示
        def refresh_order():
            update_order_text()
            editor_window.after(100, refresh_order)
        
        refresh_order()
    
    def validate_swap_pages(self):
        """校验页码输入是否有效"""
        if not self.swap_pdfs:
            self.swap_validate_var.set("")
            return
        
        try:
            page_a = int(self.page_a_var.get())
            page_b = int(self.page_b_var.get())
        except ValueError:
            self.swap_validate_var.set("请输入有效的页码")
            return
        
        if page_a < 1 or page_b < 1:
            self.swap_validate_var.set("页码必须大于0")
            return
        
        # 检查是否超出任何PDF的页数
        overflow_files = []
        for pdf_path in self.swap_pdfs:
            page_count = PDFHandler.get_page_count(pdf_path)
            if page_a > page_count or page_b > page_count:
                file_name = os.path.basename(pdf_path)
                overflow_files.append(f"{file_name} ({page_count}页)")
        
        if overflow_files:
            self.swap_validate_var.set(f"警告：以下文件页数不足：{'; '.join(overflow_files)}")
        else:
            self.swap_validate_var.set("")
    
    def preview_swap_result(self):
        """预览页面交换结果"""
        if not self.swap_pdfs:
            show_error("请选择要交换页面的PDF文件")
            return
        
        try:
            page_a = int(self.page_a_var.get())
            page_b = int(self.page_b_var.get())
        except ValueError:
            show_error("请输入有效的页码")
            return
        
        if page_a < 1 or page_b < 1:
            show_error("页码必须大于0")
            return
        
        # 创建预览窗口
        preview_window = tk.Toplevel(self.root)
        preview_window.title("页面交换预览")
        preview_window.geometry("600x400")
        
        # 创建滚动条
        scrollbar = ttk.Scrollbar(preview_window)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建文本框
        text = tk.Text(preview_window, yscrollcommand=scrollbar.set, font=('Courier New', 10))
        text.pack(fill=tk.BOTH, expand=True)
        
        # 配置滚动条
        scrollbar.config(command=text.yview)
        
        # 显示标题
        text.insert(tk.END, "PDF文件页面交换预览\n")
        text.insert(tk.END, "=" * 60 + "\n")
        text.insert(tk.END, f"交换页面：{page_a} ↔ {page_b}\n\n")
        
        # 显示每个文件的预览
        for pdf_path in self.swap_pdfs:
            pdf_name = os.path.basename(pdf_path)
            text.insert(tk.END, f"文件：{pdf_name}\n")
            
            page_count = PDFHandler.get_page_count(pdf_path)
            
            if page_a > page_count or page_b > page_count:
                text.insert(tk.END, f"  错误：文件页数不足（共{page_count}页）\n\n")
            else:
                # 生成交换前的页面顺序（只显示前20页）
                text.insert(tk.END, f"  总页数：{page_count}页\n")
                text.insert(tk.END, f"  交换前：")
                if page_count <= 20:
                    pages_before = list(range(1, page_count + 1))
                else:
                    pages_before = list(range(1, 11)) + ['...'] + list(range(page_count - 8, page_count + 1))
                
                text.insert(tk.END, "[" + ", ".join(str(p) for p in pages_before) + "]\n")
                
                # 生成交换后的页面顺序
                text.insert(tk.END, f"  交换后：")
                if page_count <= 20:
                    pages_after = list(range(1, page_count + 1))
                    pages_after[page_a - 1], pages_after[page_b - 1] = pages_after[page_b - 1], pages_after[page_a - 1]
                else:
                    pages_after = []
                    for i in range(1, page_count + 1):
                        if i <= 10 or i > page_count - 9:
                            if i == page_a:
                                pages_after.append(page_b)
                            elif i == page_b:
                                pages_after.append(page_a)
                            else:
                                pages_after.append(i)
                        elif i == 11:
                            pages_after.append('...')
                
                text.insert(tk.END, "[" + ", ".join(str(p) for p in pages_after) + "]\n\n")
        
        # 显示统计信息
        text.insert(tk.END, "=" * 60 + "\n")
        text.insert(tk.END, f"总文件数：{len(self.swap_pdfs)}\n")
        
        # 禁用文本编辑
        text.config(state=tk.DISABLED)
    
    def select_export_pdf(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF文件", "*.pdf")])
        if paths:
            valid_pdfs = [p for p in paths if is_valid_pdf(p)]
            self.export_pdfs = valid_pdfs
            self.export_pdf_var.set(f"已选择 {len(valid_pdfs)} 个文件" if len(valid_pdfs) > 1 else valid_pdfs[0])
            
            total_pages = sum(PDFHandler.get_page_count(p) for p in valid_pdfs)
            self.export_pdf_count_var.set(f"共 {len(valid_pdfs)} 个文件，总页数: {total_pages}页")
    
    def select_export_excel_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx;*.xls")])
        if path:
            self.export_excel_file = path
            self.export_excel_var.set(path)
    
    def select_rename_pdfs(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF文件", "*.pdf")])
        if paths:
            self.rename_pdfs = list(paths)
            self.rename_pdf_var.set(f"已选择 {len(paths)} 个文件")
    
    def select_rename_output_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.rename_output_dir = path
            self.rename_output_var.set(path)
    
    def select_reorder_pdfs(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF文件", "*.pdf")])
        if paths:
            self.reorder_pdfs = list(paths)
            self.reorder_pdf_var.set(f"已选择 {len(paths)} 个文件")
    
    def select_reorder_excel_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx;*.xls")])
        if path:
            self.reorder_excel_file = path
            self.reorder_excel_var.set(path)
    
    def select_reorder_name_header(self):
        # 选择姓名表头
        if not self.reorder_excel_file:
            show_error("请先选择Excel文件")
            return
        
        try:
            headers = read_excel_headers_with_merged_cells(self.reorder_excel_file)
            if not headers:
                show_error("无法读取Excel表头")
                return
            
            dialog = tk.Toplevel(self.root)
            dialog.title("选择姓名表头")
            dialog.geometry("300x200")
            dialog.transient(self.root)
            dialog.grab_set()
            
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(frame, text="选择姓名表头:").pack(anchor=tk.W, pady=5)
            header_var = tk.StringVar()
            header_combo = ttk.Combobox(frame, textvariable=header_var, values=headers, width=30)
            header_combo.pack(anchor=tk.W, pady=5)
            header_combo.current(0)  # 默认选中第一个
            
            btn_frame = ttk.Frame(frame)
            btn_frame.pack(fill=tk.X, pady=10)
            
            def select_header():
                header = header_var.get()
                if header:
                    self.reorder_name_header = header
                    self.reorder_name_header_var.set(header)
                    dialog.destroy()
            
            ttk.Button(btn_frame, text="确定", command=select_header).pack(side=tk.RIGHT, padx=5)
            ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.RIGHT)
            
            dialog.wait_window()
        except Exception as e:
            show_error(f"选择表头失败: {str(e)}")
    
    def select_reorder_id_header(self):
        # 选择编号表头
        if not self.reorder_excel_file:
            show_error("请先选择Excel文件")
            return
        
        try:
            headers = read_excel_headers_with_merged_cells(self.reorder_excel_file)
            if not headers:
                show_error("无法读取Excel表头")
                return
            
            dialog = tk.Toplevel(self.root)
            dialog.title("选择编号表头")
            dialog.geometry("300x200")
            dialog.transient(self.root)
            dialog.grab_set()
            
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(frame, text="选择编号表头:").pack(anchor=tk.W, pady=5)
            header_var = tk.StringVar()
            header_combo = ttk.Combobox(frame, textvariable=header_var, values=headers, width=30)
            header_combo.pack(anchor=tk.W, pady=5)
            # 默认选中第二个（如果有），否则选中第一个
            header_combo.current(1 if len(headers) > 1 else 0)
            
            btn_frame = ttk.Frame(frame)
            btn_frame.pack(fill=tk.X, pady=10)
            
            def select_header():
                header = header_var.get()
                if header:
                    self.reorder_id_header = header
                    self.reorder_id_header_var.set(header)
                    dialog.destroy()
            
            ttk.Button(btn_frame, text="确定", command=select_header).pack(side=tk.RIGHT, padx=5)
            ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.RIGHT)
            
            dialog.wait_window()
        except Exception as e:
            show_error(f"选择表头失败: {str(e)}")
    
    def select_page_pdfs(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF文件", "*.pdf")])
        if paths:
            self.page_pdfs = list(paths)
            self.page_pdf_var.set(f"已选择 {len(paths)} 个文件")
    
    def select_page_output_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.page_output_dir = path
            self.page_output_var.set(path)
    
    def select_reverse_pdfs(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF文件", "*.pdf")])
        if paths:
            self.reverse_pdfs = list(paths)
            self.reverse_pdf_var.set(f"已选择 {len(paths)} 个文件")
    
    def select_reverse_output_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.reverse_output_dir = path
            self.reverse_output_var.set(path)
    
    def select_cross_pdf_a(self):
        """选择PDF_A（正面扫描）"""
        path = filedialog.askopenfilename(title="选择PDF_A（正面扫描）", filetypes=[("PDF文件", "*.pdf")])
        if path:
            self.cross_pdf_a = path
            self.cross_pdf_a_var.set(path)
    
    def select_cross_pdf_b(self):
        """选择PDF_B（背面扫描，已逆序）"""
        path = filedialog.askopenfilename(title="选择PDF_B（背面扫描，已逆序）", filetypes=[("PDF文件", "*.pdf")])
        if path:
            self.cross_pdf_b = path
            self.cross_pdf_b_var.set(path)
    
    def select_cross_output_dir(self):
        """选择交叉合并的输出目录"""
        path = filedialog.askdirectory(title="选择输出目录")
        if path:
            self.cross_output_dir = path
            self.cross_output_var.set(path)
    
    def preview_cross_merge(self):
        """预览交叉合并方案"""
        # 尝试从变量获取路径（防止变量不同步）
        if not self.cross_pdf_a:
            self.cross_pdf_a = self.cross_pdf_a_var.get()
        if not self.cross_pdf_b:
            self.cross_pdf_b = self.cross_pdf_b_var.get()
        
        if not self.cross_pdf_a or not self.cross_pdf_b:
            show_error("请先选择PDF_A和PDF_B")
            return
        
        try:
            pages_a = int(self.cross_pages_a.get())
            pages_b = int(self.cross_pages_b.get())
            copies = int(self.cross_copies.get())
        except ValueError:
            show_error("页数和份数必须是数字")
            return
        
        if pages_a <= 0 or pages_b <= 0 or copies <= 0:
            show_error("页数和份数必须大于0")
            return
        
        # 获取PDF页数
        try:
            with open(self.cross_pdf_a, 'rb') as f:
                reader_a = PyPDF2.PdfReader(f)
                total_pages_a = len(reader_a.pages)
            
            with open(self.cross_pdf_b, 'rb') as f:
                reader_b = PyPDF2.PdfReader(f)
                total_pages_b = len(reader_b.pages)
        except Exception as e:
            show_error(f"读取PDF文件失败：{str(e)}")
            return
        
        total_needed_a = pages_a * copies
        total_needed_b = pages_b * copies
        
        preview_text = f"=== 交叉合并预览 ===\n\n"
        preview_text += f"PDF_A（正面扫描）: {os.path.basename(self.cross_pdf_a)}\n"
        preview_text += f"  - 总页数: {total_pages_a}\n"
        preview_text += f"  - 每次取页: {pages_a}页\n"
        preview_text += f"  - 共需页数: {total_needed_a}页\n"
        if total_needed_a > total_pages_a:
            preview_text += f"  ⚠️ 页数不足！\n"
        preview_text += f"\n"
        
        preview_text += f"PDF_B（背面扫描，已逆序）: {os.path.basename(self.cross_pdf_b)}\n"
        preview_text += f"  - 总页数: {total_pages_b}\n"
        preview_text += f"  - 每次取页: {pages_b}页\n"
        preview_text += f"  - 共需页数: {total_needed_b}页\n"
        if total_needed_b > total_pages_b:
            preview_text += f"  ⚠️ 页数不足！\n"
        preview_text += f"\n"
        
        preview_text += f"合并份数: {copies}份\n"
        preview_text += f"每页输出页数: {pages_a + pages_b}页\n"
        preview_text += f"总输出页数: {copies * (pages_a + pages_b)}页\n"
        preview_text += f"\n"
        
        preview_text += f"合并模式: A取{pages_a}页 → B取{pages_b}页 → A取{pages_a}页 → B取{pages_b}页...\n"
        preview_text += f"\n"
        
        preview_text += f"输出目录: {self.cross_output_dir if self.cross_output_dir else '未选择'}\n"
        
        show_info(preview_text, title="交叉合并预览")
    
    def process_cross_merge(self):
        """执行交叉合并"""
        # 调试信息
        print(f"\n=== 调试信息 ===")
        print(f"self.cross_pdf_a: '{self.cross_pdf_a}'")
        print(f"self.cross_pdf_b: '{self.cross_pdf_b}'")
        print(f"self.cross_pages_a.get(): '{self.cross_pages_a.get()}'")
        print(f"self.cross_pages_b.get(): '{self.cross_pages_b.get()}'")
        print(f"self.cross_copies.get(): '{self.cross_copies.get()}'")
        print(f"self.cross_output_dir: '{self.cross_output_dir}'")
        print("===============")
        
        # 尝试从变量获取路径（防止变量不同步）
        if not self.cross_pdf_a:
            self.cross_pdf_a = self.cross_pdf_a_var.get()
        if not self.cross_pdf_b:
            self.cross_pdf_b = self.cross_pdf_b_var.get()
        
        if not self.cross_output_dir:
            self.cross_output_dir = self.cross_output_var.get()
        
        if not self.cross_pdf_a or not self.cross_pdf_b:
            show_error("请先选择PDF_A和PDF_B")
            return
        
        if not self.cross_output_dir:
            show_error("请选择输出目录")
            return
        
        try:
            pages_a = int(self.cross_pages_a.get())
            pages_b = int(self.cross_pages_b.get())
            copies = int(self.cross_copies.get())
        except ValueError:
            show_error("页数和份数必须是数字")
            return
        
        if pages_a <= 0 or pages_b <= 0 or copies <= 0:
            show_error("页数和份数必须大于0")
            return
        
        # 先获取页数（不保持文件打开）
        try:
            with open(self.cross_pdf_a, 'rb') as f:
                reader_a_check = PyPDF2.PdfReader(f)
                total_pages_a = len(reader_a_check.pages)
            
            with open(self.cross_pdf_b, 'rb') as f:
                reader_b_check = PyPDF2.PdfReader(f)
                total_pages_b = len(reader_b_check.pages)
        except Exception as e:
            show_error(f"读取PDF文件失败：{str(e)}")
            return
        
        total_needed_a = pages_a * copies
        total_needed_b = pages_b * copies
        
        if total_needed_a > total_pages_a:
            show_error(f"PDF_A页数不足！需要{total_needed_a}页，实际只有{total_pages_a}页")
            return
        
        if total_needed_b > total_pages_b:
            show_error(f"PDF_B页数不足！需要{total_needed_b}页，实际只有{total_pages_b}页")
            return
        
        self.status_var.set("正在执行交叉合并...")
        self.root.update_idletasks()
        
        start_time = time.time()
        print(f"\n=== PDF交叉合并开始 ===")
        print(f"开始时间: {time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"PDF_A: {os.path.basename(self.cross_pdf_a)} ({total_pages_a}页)")
        print(f"PDF_B: {os.path.basename(self.cross_pdf_b)} ({total_pages_b}页)")
        print(f"每次取页: A={pages_a}页, B={pages_b}页")
        print(f"合并份数: {copies}份")
        print(f"输出目录: {self.cross_output_dir}")
        print("-" * 60)
        
        try:
            # 保持文件打开状态进行处理
            with open(self.cross_pdf_a, 'rb') as f_a, open(self.cross_pdf_b, 'rb') as f_b:
                reader_a = PyPDF2.PdfReader(f_a)
                reader_b = PyPDF2.PdfReader(f_b)
                
                output_writer = PyPDF2.PdfWriter()
                page_idx_a = 0
                page_idx_b = 0
                
                for i in range(copies):
                    # 从PDF_A取页
                    for _ in range(pages_a):
                        if page_idx_a < total_pages_a:
                            output_writer.add_page(reader_a.pages[page_idx_a])
                            page_idx_a += 1
                    
                    # 从PDF_B取页
                    for _ in range(pages_b):
                        if page_idx_b < total_pages_b:
                            output_writer.add_page(reader_b.pages[page_idx_b])
                            page_idx_b += 1
                    
                    # 每10份打印一次进度
                    if (i + 1) % 10 == 0 or (i + 1) == copies:
                        elapsed = time.time() - start_time
                        avg_time = elapsed / (i + 1)
                        remaining = avg_time * (copies - i - 1)
                        print(f"进度: {i+1}/{copies} 已处理页数: {page_idx_a + page_idx_b}")
                        print(f"  已耗时: {elapsed:.2f}秒, 预计剩余: {remaining:.2f}秒")
            
            # 输出文件名
            base_name = os.path.splitext(os.path.basename(self.cross_pdf_a))[0]
            output_path = os.path.join(self.cross_output_dir, f"{base_name}_交叉合并.pdf")
            
            with open(output_path, 'wb') as f:
                output_writer.write(f)
            
            total_time = time.time() - start_time
            print("-" * 60)
            print(f"=== PDF交叉合并完成 ===")
            print(f"结束时间: {time.strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"总耗时: {total_time:.2f}秒")
            print(f"输出文件: {output_path}")
            print(f"总页数: {len(output_writer.pages)}页")
            
            self.status_var.set("就绪")
            show_info(f"PDF交叉合并成功！\n\n输出文件：{output_path}\n总页数：{len(output_writer.pages)}页\n耗时：{total_time:.2f}秒")
        
        except Exception as e:
            self.status_var.set("就绪")
            print(f"合并失败: {str(e)}")
            show_error(f"PDF交叉合并失败：{str(e)}")
    
    def select_batch_extract_pdfs(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF文件", "*.pdf")])
        if paths:
            self.batch_extract_pdfs = list(paths)
            self.extract_pdf_var.set(f"已选择 {len(paths)} 个文件")
    
    def select_batch_extract_output_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.batch_extract_output_dir = path
            self.extract_output_var.set(path)
    
    def select_batch_extract_excel_files(self):
        paths = filedialog.askopenfilenames(filetypes=[("Excel文件", "*.xlsx;*.xls")])
        if paths:
            self.batch_extract_excel_files = list(paths)[:5]  # 最多5个
            self.extract_excel_var.set(f"已选择 {len(self.batch_extract_excel_files)} 个文件")
    
    def select_batch_append_source_pdf(self):
        path = filedialog.askopenfilename(filetypes=[("PDF文件", "*.pdf")])
        if path:
            self.batch_append_source_pdf = path
            self.append_source_var.set(path)
    
    def select_batch_append_pdfs(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF文件", "*.pdf")])
        if paths:
            self.batch_append_pdfs = list(paths)
            self.append_target_var.set(f"已选择 {len(paths)} 个文件")
    
    def select_batch_append_output_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.batch_append_output_dir = path
            self.append_output_var.set(path)
    
    def add_merge_append_pdf_widget(self, index=None):
        """添加要追加的PDF控件"""
        if index is None:
            index = len(self.merge_append_pdf_vars)
        
        frame = ttk.Frame(self.merge_append_container)
        frame.pack(fill=tk.X, pady=2)
        
        var = tk.StringVar()
        entry = ttk.Entry(frame, textvariable=var, state="readonly", width=50)
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        btn = ttk.Button(frame, text=f"选择要追加的PDF {index + 1}", 
                         command=lambda v=var: self.select_merge_append_pdf(v))
        btn.pack(side=tk.RIGHT, padx=5)
        
        self.merge_append_pdf_vars.append(var)
        self.merge_append_frames.append(frame)
    
    def remove_merge_append_pdf_widget(self):
        """删除最后一个要追加的PDF控件"""
        if len(self.merge_append_frames) > 1:  # 至少保留1个
            frame = self.merge_append_frames.pop()
            frame.destroy()
            self.merge_append_pdf_vars.pop()
    
    def select_merge_append_pdf(self, var):
        """选择要追加的PDF文件"""
        path = filedialog.askopenfilename(filetypes=[("PDF文件", "*.pdf")])
        if path:
            var.set(path)
    
    def select_merge_main_pdf(self):
        """选择主PDF文件"""
        path = filedialog.askopenfilename(filetypes=[("PDF文件", "*.pdf")])
        if path:
            self.merge_main_pdf = path
            self.merge_main_pdf_var.set(path)
    
    def select_merge_output_dir(self):
        """选择输出目录"""
        path = filedialog.askdirectory()
        if path:
            self.merge_output_dir = path
            self.merge_output_var.set(path)
    
    def select_batch_single_main_pdfs(self):
        """选择批量单页追加的主PDF文件"""
        paths = filedialog.askopenfilenames(filetypes=[("PDF文件", "*.pdf")])
        if paths:
            self.batch_single_main_pdfs = list(paths)
            self.batch_single_main_var.set(f"已选择 {len(paths)} 个文件")
    
    def select_batch_single_append_pdfs(self):
        """选择批量单页追加的要追加PDF文件"""
        paths = filedialog.askopenfilenames(filetypes=[("PDF文件", "*.pdf")])
        if paths:
            self.batch_single_append_pdfs = list(paths)
            self.batch_single_append_var.set(f"已选择 {len(paths)} 个文件")
    
    def select_batch_single_output_dir(self):
        """选择批量单页追加的输出目录"""
        path = filedialog.askdirectory()
        if path:
            self.batch_single_output_dir = path
            self.batch_single_output_var.set(path)
    
    def preview_batch_single_files(self):
        """预览批量单页追加的文件列表"""
        if not self.batch_single_main_pdfs:
            show_warning("请先选择主PDF文件")
            return
        
        if not self.batch_single_append_pdfs:
            show_warning("请先选择要追加的PDF文件")
            return
        
        # 创建预览窗口
        preview_window = tk.Toplevel(self.root)
        preview_window.title("批量单页追加 - 文件预览")
        preview_window.geometry("800x600")
        
        # 创建滚动区域
        canvas = tk.Canvas(preview_window)
        scrollbar = ttk.Scrollbar(preview_window, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 显示主PDF文件
        main_label = ttk.Label(scrollable_frame, text="主PDF文件：", font=('微软雅黑', 12, 'bold'))
        main_label.pack(anchor="w", padx=10, pady=(10, 5))
        
        for i, pdf in enumerate(self.batch_single_main_pdfs):
            page_count = PDFHandler.get_page_count(pdf)
            ttk.Label(scrollable_frame, text=f"  [{i+1}] {os.path.basename(pdf)} ({page_count}页)").pack(anchor="w", padx=20)
        
        # 显示要追加的PDF文件
        append_label = ttk.Label(scrollable_frame, text="\n要追加的PDF文件：", font=('微软雅黑', 12, 'bold'))
        append_label.pack(anchor="w", padx=10, pady=(10, 5))
        
        for i, pdf in enumerate(self.batch_single_append_pdfs):
            page_count = PDFHandler.get_page_count(pdf)
            ttk.Label(scrollable_frame, text=f"  [{i+1}] {os.path.basename(pdf)} ({page_count}页)").pack(anchor="w", padx=20)
        
        # 显示匹配关系
        match_label = ttk.Label(scrollable_frame, text="\n匹配关系预览：", font=('微软雅黑', 12, 'bold'))
        match_label.pack(anchor="w", padx=10, pady=(10, 5))
        
        main_count = len(self.batch_single_main_pdfs)
        display_count = min(len(self.batch_single_append_pdfs), 20)  # 只显示前20个匹配关系
        
        for i in range(display_count):
            main_index = i % main_count
            main_name = os.path.splitext(os.path.basename(self.batch_single_main_pdfs[main_index]))[0]
            append_name = os.path.splitext(os.path.basename(self.batch_single_append_pdfs[i]))[0]
            ttk.Label(scrollable_frame, text=f"  [{i+1}] {main_name} + {append_name} → {main_name}_{append_name}.pdf").pack(anchor="w", padx=20)
        
        if len(self.batch_single_append_pdfs) > 20:
            remaining = len(self.batch_single_append_pdfs) - 20
            ttk.Label(scrollable_frame, text=f"  ... 还有 {remaining} 个文件（共 {len(self.batch_single_append_pdfs)} 个）").pack(anchor="w", padx=20)
        
        # 显示统计信息
        stats_label = ttk.Label(scrollable_frame, text="\n统计信息：", font=('微软雅黑', 12, 'bold'))
        stats_label.pack(anchor="w", padx=10, pady=(10, 5))
        
        ttk.Label(scrollable_frame, text=f"  主PDF数量：{len(self.batch_single_main_pdfs)}").pack(anchor="w", padx=20)
        ttk.Label(scrollable_frame, text=f"  追加PDF数量：{len(self.batch_single_append_pdfs)}").pack(anchor="w", padx=20)
        ttk.Label(scrollable_frame, text=f"  提取页码：{self.batch_single_page_var.get()}").pack(anchor="w", padx=20)
        ttk.Label(scrollable_frame, text=f"  输出目录：{self.batch_single_output_dir if self.batch_single_output_dir else '未选择'}").pack(anchor="w", padx=20)
        ttk.Label(scrollable_frame, text=f"  将生成：{len(self.batch_single_append_pdfs)} 个文件").pack(anchor="w", padx=20)
    
    def select_student_pdfs(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF文件", "*.pdf")])
        if paths:
            self.student_pdfs = list(paths)
            self.student_pdf_var.set(f"已选择 {len(paths)} 个文件")
    
    def select_class_pdfs(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF文件", "*.pdf")])
        if paths:
            self.class_pdfs = list(paths)
            self.class_pdf_var.set(f"已选择 {len(paths)} 个文件")
    
    def select_student_class_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx;*.xls")])
        if path:
            self.student_class_excel_file = path
            self.student_class_excel_var.set(path)
    
    def select_student_class_name_header(self):
        # 选择姓名表头
        if not self.student_class_excel_file:
            show_error("请先选择Excel文件")
            return
        
        try:
            headers = read_excel_headers_with_merged_cells(self.student_class_excel_file)
            if not headers:
                show_error("无法读取Excel表头")
                return
            
            dialog = tk.Toplevel(self.root)
            dialog.title("选择姓名表头")
            dialog.geometry("300x200")
            dialog.transient(self.root)
            dialog.grab_set()
            
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(frame, text="选择姓名表头:").pack(anchor=tk.W, pady=5)
            header_var = tk.StringVar()
            header_combo = ttk.Combobox(frame, textvariable=header_var, values=headers, width=30)
            header_combo.pack(anchor=tk.W, pady=5)
            header_combo.current(0)  # 默认选中第一个
            
            btn_frame = ttk.Frame(frame)
            btn_frame.pack(fill=tk.X, pady=10)
            
            def select_header():
                header = header_var.get()
                if header:
                    self.student_class_name_header = header
                    self.student_class_name_header_var.set(header)
                    dialog.destroy()
            
            ttk.Button(btn_frame, text="确定", command=select_header).pack(side=tk.RIGHT, padx=5)
            ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.RIGHT)
            
            dialog.wait_window()
        except Exception as e:
            show_error(f"选择表头失败: {str(e)}")
    
    def select_student_class_class_header(self):
        # 选择PDF_B表头
        if not self.student_class_excel_file:
            show_error("请先选择Excel文件")
            return
        
        try:
            headers = read_excel_headers_with_merged_cells(self.student_class_excel_file)
            if not headers:
                show_error("无法读取Excel表头")
                return
            
            dialog = tk.Toplevel(self.root)
            dialog.title("选择PDF_B表头")
            dialog.geometry("300x200")
            dialog.transient(self.root)
            dialog.grab_set()
            
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(frame, text="选择PDF_B表头:").pack(anchor=tk.W, pady=5)
            header_var = tk.StringVar()
            header_combo = ttk.Combobox(frame, textvariable=header_var, values=headers, width=30)
            header_combo.pack(anchor=tk.W, pady=5)
            # 默认选中第二个（如果有），否则选中第一个
            header_combo.current(1 if len(headers) > 1 else 0)
            
            btn_frame = ttk.Frame(frame)
            btn_frame.pack(fill=tk.X, pady=10)
            
            def select_header():
                header = header_var.get()
                if header:
                    self.student_class_class_header = header
                    self.student_class_class_header_var.set(header)
                    dialog.destroy()
            
            ttk.Button(btn_frame, text="确定", command=select_header).pack(side=tk.RIGHT, padx=5)
            ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.RIGHT)
            
            dialog.wait_window()
        except Exception as e:
            show_error(f"选择表头失败: {str(e)}")
    
    def select_student_class_output_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.student_class_output_dir = path
            self.student_class_output_var.set(path)
    
    def select_excel_files(self):
        paths = filedialog.askopenfilenames(filetypes=[("Excel文件", "*.xlsx;*.xls")])
        if paths:
            self.excel_files = list(paths)
            self.excel_files_var.set(f"已选择 {len(paths)} 个文件")
    
    def select_excel_output_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.excel_output_dir = path
            self.excel_output_var.set(path)
    
    def select_word_files(self):
        paths = filedialog.askopenfilenames(filetypes=[("Word文件", "*.docx;*.doc")])
        if paths:
            self.word_files = list(paths)
            self.word_files_var.set(f"已选择 {len(paths)} 个文件")
    
    def select_word_output_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.word_output_dir = path
            self.word_output_var.set(path)
    
    # ==================== 处理方法 ====================
    
    def process_insert(self):
        """处理PDF插入"""
        if not self.insert_pdf:
            show_error("请选择主PDF文件")
            return
        
        interval = int(self.interval_var.get())
        pages = int(self.insert_pages_var.get())
        
        if interval < 1 or pages < 1:
            show_error("插入间隔和每页插入页数必须大于0")
            return
        
        # 获取子PDF路径列表（自动跳过空路径）
        child_pdf_paths = [var.get().strip() for var in self.child_pdf_vars if var.get().strip()]
        
        if not child_pdf_paths:
            show_error("请至少选择一个子PDF文件")
            return
        
        output_dir = filedialog.askdirectory(title="选择输出目录")
        if not output_dir:
            return
        
        try:
            base_name = os.path.splitext(os.path.basename(self.insert_pdf))[0]
            output_path = os.path.join(output_dir, f"{base_name}_插入后.pdf")
            
            # 简单的插入实现
            with open(self.insert_pdf, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                total_pages = len(reader.pages)
                
                writer = PyPDF2.PdfWriter()
                for i in range(total_pages):
                    writer.add_page(reader.pages[i])
                    # 在指定间隔后插入额外页面
                    if (i + 1) % interval == 0:
                        # 使用子PDF路径循环插入
                        for j in range(pages):
                            pdf_index = j % len(child_pdf_paths)
                            insert_path = child_pdf_paths[pdf_index]
                            with open(insert_path, 'rb') as insert_f:
                                insert_reader = PyPDF2.PdfReader(insert_f)
                                for insert_page in insert_reader.pages:
                                    writer.add_page(insert_page)
            
            with open(output_path, 'wb') as f:
                writer.write(f)
            
            show_info(f"PDF插入完成！输出文件：{output_path}")
        except Exception as e:
            show_error(f"插入失败: {str(e)}")
    
    def process_split(self):
        """处理PDF分割"""
        if not self.split_pdf:
            show_error("请选择要分割的PDF文件")
            return
        
        try:
            split_pages = int(self.split_pages_var.get())
        except ValueError:
            show_error("请输入有效的分割页数")
            return
        
        if split_pages < 1:
            show_error("分割页数必须大于0")
            return
        
        output_dir = self.split_output_dir if hasattr(self, 'split_output_dir') else filedialog.askdirectory(title="选择输出目录")
        if not output_dir:
            return
        
        # 获取命名方式
        naming_method = self.naming_var.get()
        names = None
        
        # 如果选择Excel命名，读取Excel数据
        if naming_method == "excel":
            if not hasattr(self, 'excel_file') or not self.excel_file:
                show_error("请选择Excel文件")
                return
            
            names = self.read_excel_names(self.excel_file)
            if not names:
                show_error("无法从Excel文件中读取数据")
                return
        
        # 执行分割
        success, message = PDFHandler.split_pdf(self.split_pdf, split_pages, output_dir, names)
        
        if success:
            show_info(f"PDF分割完成！\n\n{message}\n输出到：{output_dir}")
        else:
            show_error(f"分割失败：{message}")
    
    def read_excel_names(self, excel_file):
        """从Excel文件中读取姓名列表"""
        names = []
        try:
            if excel_file.endswith('.xlsx'):
                from openpyxl import load_workbook
                workbook = load_workbook(excel_file, data_only=True)
                sheet = workbook.active
                
                # 获取合并单元格信息
                merged_cells = sheet.merged_cells
                
                # 扫描前三行找到表头行
                header_row = -1
                headers = []
                
                for row_idx in range(1, min(4, sheet.max_row + 1)):
                    row_headers = []
                    for col_idx in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        
                        # 检查是否是合并单元格
                        is_merged = False
                        for merged_range in merged_cells:
                            if cell.coordinate in merged_range:
                                is_merged = True
                                if cell.coordinate == merged_range.start_cell.coordinate:
                                    row_headers.append(str(cell.value).strip() if cell.value else "")
                                else:
                                    row_headers.append("")
                                break
                        
                        if not is_merged:
                            row_headers.append(str(cell.value).strip() if cell.value else "")
                    
                    # 查找包含"姓名"或"名称"或"成果"的列
                    for i, header in enumerate(row_headers):
                        if header and ('姓名' in header or '名称' in header or '成果' in header):
                            header_row = row_idx
                            name_col = i + 1  # 转换为1-based索引
                            break
                    
                    if header_row != -1:
                        break
                
                # 如果没有找到表头，默认使用第一列
                if header_row == -1:
                    name_col = 1
                
                # 读取姓名列数据
                for row_idx in range(header_row + 1, sheet.max_row + 1):
                    cell = sheet.cell(row=row_idx, column=name_col)
                    name = str(cell.value).strip() if cell.value else ""
                    if name and name != "None":
                        names.append(name)
            else:
                import xlrd
                workbook = xlrd.open_workbook(excel_file)
                sheet = workbook.sheet_by_index(0)
                
                # 扫描前三行找到表头行
                header_row = -1
                name_col = 0
                
                for row_idx in range(min(3, sheet.nrows)):
                    row_headers = [str(sheet.cell(row_idx, col_idx).value).strip() for col_idx in range(sheet.ncols)]
                    for i, header in enumerate(row_headers):
                        if header and ('姓名' in header or '名称' in header or '成果' in header):
                            header_row = row_idx
                            name_col = i
                            break
                    if header_row != -1:
                        break
                
                # 读取姓名列数据
                for row_idx in range(header_row + 1, sheet.nrows):
                    name = str(sheet.cell(row_idx, name_col).value).strip()
                    if name and name != "None":
                        names.append(name)
            
            return names
        except Exception as e:
            show_error(f"读取Excel文件时出错：{str(e)}")
            return []
    
    def process_swap(self):
        """处理PDF页面重排序"""
        if not self.swap_pdfs:
            show_error("请选择要交换页面的PDF文件")
            return
        
        if not hasattr(self, 'swap_page_order') or not self.swap_page_order:
            show_error("请先打开交换方案编辑窗口设置页面顺序")
            return
        
        output_dir = self.swap_output_dir if hasattr(self, 'swap_output_dir') else filedialog.askdirectory(title="选择输出目录")
        if not output_dir:
            return
        
        success_count = 0
        for pdf_path in self.swap_pdfs:
            base_name = os.path.splitext(os.path.basename(pdf_path))[0]
            output_path = os.path.join(output_dir, f"{base_name}_重排后.pdf")
            
            if PDFHandler.reorder_pages(pdf_path, self.swap_page_order, output_path):
                success_count += 1
        
        show_info(f"页面重排完成！成功处理 {success_count}/{len(self.swap_pdfs)} 个文件")
    
    def parse_page_groups(self, range_str):
        """解析页面范围分组（支持如"1-6,7-10"导出多个多页PDF）"""
        groups = []
        if not range_str:
            return groups
        
        parts = range_str.split(',')
        for part in parts:
            part = part.strip()
            if '-' in part:
                try:
                    start, end = part.split('-')
                    start = int(start.strip())
                    end = int(end.strip())
                    if start <= end:
                        groups.append((start, end))
                except ValueError:
                    pass
            else:
                try:
                    page = int(part.strip())
                    groups.append((page, page))
                except ValueError:
                    pass
        
        return groups
    
    def process_export(self):
        """处理指定导出页（批量版）"""
        if not self.export_pdfs:
            show_error("请选择要导出的PDF文件")
            return
        
        page_groups = self.parse_page_groups(self.export_pages_var.get())
        if not page_groups:
            show_error("请输入有效的导出页面范围（如：1-6,7-10）")
            return
        
        output_dir = filedialog.askdirectory(title="选择输出目录")
        if not output_dir:
            return
        
        use_original_name = self.export_use_original_name_var.get()
        success_count = 0
        fail_count = 0
        skipped_count = 0  # 统计页数不足被跳过的文件
        skipped_files = []  # 记录被跳过的文件名
        total_files = len(self.export_pdfs)
        current_file = 0
        
        for pdf_path in self.export_pdfs:
            current_file += 1
            total_pages = PDFHandler.get_page_count(pdf_path)
            base_name = os.path.splitext(os.path.basename(pdf_path))[0]
            file_exported = False  # 标记该文件是否成功导出
            
            for start, end in page_groups:
                # 过滤超出PDF页数的范围
                actual_start = max(start, 1)
                actual_end = min(end, total_pages)
                
                # 如果起始页超过总页数，跳过这个范围组
                if actual_start > total_pages:
                    continue
                
                # 如果实际结束页小于起始页，跳过
                if actual_end < actual_start:
                    continue
                
                # 生成要导出的页面列表
                export_pages = list(range(actual_start, actual_end + 1))
                
                # 生成输出文件名
                if use_original_name:
                    # 使用原文件名，添加序号前缀和页面范围（避免多范围组时覆盖）
                    if start == end:
                        output_filename = f"{current_file:02d}_{base_name}_p{start}.pdf"
                    else:
                        output_filename = f"{current_file:02d}_{base_name}_p{start}-{end}.pdf"
                else:
                    # 不使用原文件名时，添加文件序号和页面范围
                    if start == end:
                        output_filename = f"{current_file:03d}_page_{start}.pdf"
                    else:
                        output_filename = f"{current_file:03d}_pages_{start}-{end}.pdf"
                
                output_path = os.path.join(output_dir, output_filename)
                
                if PDFHandler.extract_pages(pdf_path, export_pages, output_path):
                    success_count += 1
                    file_exported = True
                else:
                    fail_count += 1
            
            # 如果该文件所有范围组都被跳过，记录下来
            if not file_exported and total_pages > 0:
                skipped_count += 1
                skipped_files.append(os.path.basename(pdf_path))
            
            # 更新进度
            progress = (current_file / total_files) * 100
            self.status_var.set(f"处理中... {current_file}/{total_files} ({int(progress)}%)")
            self.root.update()
        
        self.status_var.set("处理完成")
        
        # 显示导出结果
        result_msg = f"导出完成！成功：{success_count}"
        if fail_count > 0:
            result_msg += f"，失败：{fail_count}"
        if skipped_count > 0:
            result_msg += f"，跳过页数不足：{skipped_count}"
        
        if fail_count > 0 or skipped_count > 0:
            show_warning(result_msg)
            if skipped_files:
                show_info(f"被跳过的文件（页数不足）：\n{chr(10).join(skipped_files[:10])}{'...' if len(skipped_files) > 10 else ''}")
        else:
            show_info(result_msg)
    
    def process_rename(self):
        """处理PDF重命名"""
        if not self.rename_pdfs:
            show_error("请选择要重命名的PDF文件")
            return
        
        output_dir = self.rename_output_dir if hasattr(self, 'rename_output_dir') else filedialog.askdirectory(title="选择输出目录")
        if not output_dir:
            return
        
        prefix = self.prefix_var.get()
        suffix = self.suffix_var.get()
        mode = self.rename_mode_var.get()
        
        success_count = 0
        for i, pdf_path in enumerate(self.rename_pdfs):
            base_name = os.path.splitext(os.path.basename(pdf_path))[0]
            
            if mode == "prefix":
                new_name = f"{prefix}{base_name}.pdf"
            elif mode == "suffix":
                new_name = f"{base_name}{suffix}.pdf"
            else:
                # 按班级模式，使用序号
                new_name = f"{base_name}_{i+1:02d}.pdf"
            
            output_path = os.path.join(output_dir, clean_filename(new_name))
            shutil.copy2(pdf_path, output_path)
            success_count += 1
        
        show_info(f"重命名完成！成功处理 {success_count} 个文件")
    
    def process_reorder(self):
        """处理PDF重排序"""
        if not self.reorder_pdfs:
            show_error("请选择要重排序的PDF文件")
            return
        
        output_dir = filedialog.askdirectory(title="选择输出目录")
        if not output_dir:
            return
        
        use_excel = self.reorder_excel_file and self.reorder_name_header and self.reorder_id_header
        
        # 读取Excel数据（如果使用）
        name_to_id = {}
        if use_excel:
            name_id_mapping = read_name_id_mapping(self.reorder_excel_file, self.reorder_name_header, self.reorder_id_header)
            name_to_id = {name: id_value for name, id_value in name_id_mapping}
        
        # 处理PDF文件
        pdf_info = []
        for pdf_path in self.reorder_pdfs:
            original_name = os.path.basename(pdf_path)
            # 从文件名中提取姓名
            match = re.match(r'^([^_\-]+)[_\-].*\.pdf$', original_name)
            extracted_name = match.group(1) if match else original_name.split('.')[0]
            
            id_value = name_to_id.get(extracted_name) if use_excel else None
            pdf_info.append((extracted_name, id_value, original_name, pdf_path))
        
        # 排序
        if use_excel:
            pdf_info.sort(key=lambda x: (0 if x[1] else 1, int(x[1]) if x[1] and str(x[1]).isdigit() else x[1] or x[0]))
        
        # 保存
        success_count = 0
        for i, (name, id_value, original_name, pdf_path) in enumerate(pdf_info, 1):
            if use_excel and id_value:
                try:
                    new_name = f"{name}_{int(id_value):02d}.pdf"
                except:
                    new_name = f"{name}_{id_value}.pdf"
            else:
                new_name = f"{name}_{i:02d}.pdf"
            
            output_path = os.path.join(output_dir, clean_filename(new_name))
            shutil.copy2(pdf_path, output_path)
            success_count += 1
        
        show_info(f"重排序完成！成功处理 {success_count} 个文件")
    
    def process_page(self):
        """处理PDF页面操作"""
        if not self.page_pdfs:
            show_error("请选择PDF文件")
            return
        
        output_dir = self.page_output_dir if hasattr(self, 'page_output_dir') else filedialog.askdirectory(title="选择输出目录")
        if not output_dir:
            return
        
        action = self.page_action_var.get()
        success_count = 0
        
        for pdf_path in self.page_pdfs:
            base_name = os.path.splitext(os.path.basename(pdf_path))[0]
            
            if action == "reverse":
                output_path = os.path.join(output_dir, f"{base_name}_反转.pdf")
                if PDFHandler.reverse_pages(pdf_path, output_path):
                    success_count += 1
            elif action == "swap":
                # 页面交换
                try:
                    page_a = int(self.swap_page_a_var.get())
                    page_b = int(self.swap_page_b_var.get())
                except ValueError:
                    show_error("请输入有效的页码")
                    return
                
                if page_a < 1 or page_b < 1:
                    show_error("页码必须大于0")
                    return
                
                output_path = os.path.join(output_dir, f"{base_name}_交换{page_a}-{page_b}.pdf")
                if PDFHandler.swap_pages(pdf_path, page_a, page_b, output_path):
                    success_count += 1
            else:
                # 旋转
                try:
                    angle = int(self.rotate_angle_var.get())
                except ValueError:
                    show_error("请输入有效的旋转角度")
                    return
                
                rotation_scheme = {}
                # 获取PDF页数
                with open(pdf_path, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    total_pages = len(reader.pages)
                    for i in range(total_pages):
                        rotation_scheme[i + 1] = angle
                
                output_path = os.path.join(output_dir, f"{base_name}_旋转{angle}度.pdf")
                if PDFHandler.rotate_pages(pdf_path, rotation_scheme, output_path):
                    success_count += 1
        
        show_info(f"页面操作完成！成功处理 {success_count}/{len(self.page_pdfs)} 个文件")
    
    def process_reverse(self):
        """处理PDF页面逆序"""
        if not self.reverse_pdfs:
            show_error("请选择PDF文件")
            return
        
        if not self.reverse_output_dir:
            show_error("请选择输出目录")
            return
        
        success_count = 0
        fail_count = 0
        error_messages = []
        
        for pdf_path in self.reverse_pdfs:
            base_name = os.path.splitext(os.path.basename(pdf_path))[0]
            output_path = os.path.join(self.reverse_output_dir, f"{base_name}_reversed.pdf")
            
            if PDFHandler.reverse_pages(pdf_path, output_path):
                success_count += 1
            else:
                fail_count += 1
                error_messages.append(f"  • {base_name}.pdf")
        
        if fail_count > 0:
            error_text = "\n".join(error_messages)
            show_warning(f"页面逆序完成！\n成功：{success_count}，失败：{fail_count}\n\n失败文件：\n{error_text}")
        else:
            show_info(f"页面逆序完成！成功处理 {success_count} 个文件")
    
    def process_batch_extract(self):
        """处理批量提取"""
        if not self.batch_extract_pdfs:
            show_error("请选择要提取的PDF文件")
            return
        
        if not self.batch_extract_output_dir:
            show_error("请选择输出目录")
            return
        
        try:
            page = int(self.batch_extract_page_var.get())
        except ValueError:
            show_error("请输入有效的提取页码")
            return
        
        if page < 1:
            show_error("提取页码必须大于0")
            return
        
        success_count = 0
        fail_count = 0
        
        for pdf_path in self.batch_extract_pdfs:
            base_name = os.path.splitext(os.path.basename(pdf_path))[0]
            output_path = os.path.join(self.batch_extract_output_dir, f"{base_name}-{page}页.pdf")
            
            if PDFHandler.extract_pages(pdf_path, [page], output_path):
                success_count += 1
            else:
                fail_count += 1
        
        if fail_count > 0:
            show_warning(f"批量提取完成！成功：{success_count}，失败：{fail_count}")
        else:
            show_info(f"批量提取完成！成功处理 {success_count} 个文件")
    
    def process_batch_append(self):
        """处理批量追加页"""
        if not self.batch_append_source_pdf:
            show_error("请选择源PDF文件")
            return
        
        if not self.batch_append_pdfs:
            show_error("请选择要追加的PDF文件")
            return
        
        if not self.batch_append_output_dir:
            show_error("请选择输出目录")
            return
        
        success_count = 0
        fail_count = 0
        
        for pdf_path in self.batch_append_pdfs:
            base_name = os.path.splitext(os.path.basename(pdf_path))[0]
            output_path = os.path.join(self.batch_append_output_dir, f"{base_name}-追加页.pdf")
            
            if PDFHandler.merge_pdfs([pdf_path, self.batch_append_source_pdf], output_path):
                success_count += 1
            else:
                fail_count += 1
        
        if fail_count > 0:
            show_warning(f"批量追加完成！成功：{success_count}，失败：{fail_count}")
        else:
            show_info(f"批量追加完成！成功处理 {success_count} 个文件")
    
    def process_merge(self):
        """处理PDF合并"""
        if not self.merge_main_pdf:
            show_error("请选择主PDF文件")
            return
        
        # 获取所有要追加的PDF路径
        append_pdfs = [var.get() for var in self.merge_append_pdf_vars if var.get()]
        if not append_pdfs:
            show_error("请至少选择一个要追加的PDF文件")
            return
        
        if not self.merge_output_dir:
            show_error("请选择输出目录")
            return
        
        # 显示PDF页面信息
        main_page_count = PDFHandler.get_page_count(self.merge_main_pdf)
        info_text = f"主PDF: {os.path.basename(self.merge_main_pdf)} ({main_page_count}页)\n"
        info_text += "要追加的PDF:\n"
        for pdf in append_pdfs:
            page_count = PDFHandler.get_page_count(pdf)
            info_text += f"  • {os.path.basename(pdf)} ({page_count}页)\n"
        
        # 生成输出文件名
        base_name = os.path.splitext(os.path.basename(self.merge_main_pdf))[0]
        output_path = os.path.join(self.merge_output_dir, f"{base_name}-合并.pdf")
        
        # 执行合并
        success, msg = PDFHandler.append_pdfs(self.merge_main_pdf, append_pdfs, output_path)
        if success:
            show_info(f"PDF合并成功！\n{info_text}输出文件：{output_path}")
        else:
            show_error(f"PDF合并失败：{msg}")
    
    def process_batch_single_append(self):
        """处理批量单页追加"""
        import time
        
        if not self.batch_single_main_pdfs:
            show_error("请选择主PDF文件")
            return
        
        if not self.batch_single_append_pdfs:
            show_error("请选择要追加的PDF文件")
            return
        
        if not self.batch_single_output_dir:
            show_error("请选择输出目录")
            return
        
        try:
            page_number = int(self.batch_single_page_var.get())
        except ValueError:
            show_error("请输入有效的页码")
            return
        
        if page_number < 1:
            show_error("页码必须大于0")
            return
        
        # 显示PDF页面信息
        info_text = "主PDF文件:\n"
        for pdf in self.batch_single_main_pdfs:
            page_count = PDFHandler.get_page_count(pdf)
            info_text += f"  • {os.path.basename(pdf)} ({page_count}页)\n"
        
        info_text += f"\n要追加的PDF文件（提取第{page_number}页）:\n"
        for pdf in self.batch_single_append_pdfs:
            page_count = PDFHandler.get_page_count(pdf)
            info_text += f"  • {os.path.basename(pdf)} ({page_count}页)\n"
        
        # 日志打印
        total_start_time = time.time()
        print(f"=== 批量单页追加开始 ===")
        print(f"开始时间: {time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"主PDF数量: {len(self.batch_single_main_pdfs)}")
        print(f"追加PDF数量: {len(self.batch_single_append_pdfs)}")
        print(f"提取页码: {page_number}")
        print(f"输出目录: {self.batch_single_output_dir}")
        print("-" * 60)
        
        success_count = 0
        fail_count = 0
        main_count = len(self.batch_single_main_pdfs)
        error_messages = []
        
        for i, append_pdf in enumerate(self.batch_single_append_pdfs):
            file_start_time = time.time()
            
            # 循环使用主PDF
            main_pdf_index = i % main_count
            main_pdf = self.batch_single_main_pdfs[main_pdf_index]
            
            # 生成输出文件名
            main_name = os.path.splitext(os.path.basename(main_pdf))[0]
            append_name = os.path.splitext(os.path.basename(append_pdf))[0]
            output_path = os.path.join(self.batch_single_output_dir, f"{main_name}_{append_name}.pdf")
            
            # 日志打印 - 开始处理
            print(f"[{i+1}/{len(self.batch_single_append_pdfs)}] 开始处理")
            print(f"  主PDF: {main_name}.pdf")
            print(f"  追加PDF: {append_name}.pdf")
            print(f"  输出: {os.path.basename(output_path)}")
            
            # 执行单页追加
            success, msg = PDFHandler.append_single_page(main_pdf, append_pdf, page_number, output_path)
            
            file_time = time.time() - file_start_time
            
            if success:
                success_count += 1
                print(f"  ✓ 成功 (耗时: {file_time:.2f}秒)")
            else:
                fail_count += 1
                error_messages.append(f"  • {append_name}: {msg}")
                print(f"  ✗ 失败: {msg} (耗时: {file_time:.2f}秒)")
            
            # 每处理10个文件打印一次进度汇总
            if (i + 1) % 10 == 0:
                elapsed_time = time.time() - total_start_time
                avg_time = elapsed_time / (i + 1)
                remaining_files = len(self.batch_single_append_pdfs) - (i + 1)
                estimated_remaining = avg_time * remaining_files
                print(f"\n--- 进度: {i+1}/{len(self.batch_single_append_pdfs)} ---")
                print(f"已耗时: {elapsed_time:.2f}秒")
                print(f"平均耗时: {avg_time:.2f}秒/文件")
                print(f"预计剩余: {estimated_remaining:.2f}秒")
                print("-" * 60)
        
        total_time = time.time() - total_start_time
        
        # 日志打印 - 完成
        print("-" * 60)
        print(f"=== 批量单页追加完成 ===")
        print(f"结束时间: {time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"总耗时: {total_time:.2f}秒")
        print(f"成功: {success_count}, 失败: {fail_count}")
        print(f"平均耗时: {total_time / len(self.batch_single_append_pdfs):.2f}秒/文件")
        
        if fail_count > 0:
            error_text = "\n".join(error_messages)
            show_warning(f"批量单页追加完成！\n成功：{success_count}，失败：{fail_count}\n\n失败详情：\n{error_text}")
        else:
            show_info(f"批量单页追加完成！\n{info_text}成功处理 {success_count} 个文件\n总耗时：{total_time:.2f}秒")
    
    def process_student_class_merge(self):
        """处理PDF_A与PDF_B合并"""
        if not self.student_pdfs:
            show_error("请选择PDF_A文件")
            return
        
        if not self.class_pdfs:
            show_error("请选择PDF_B文件")
            return
        
        if not self.student_class_excel_file:
            show_error("请选择Excel文件")
            return
        
        if not self.student_class_output_dir:
            show_error("请选择输出目录")
            return
        
        # 读取Excel数据
        excel_data = read_excel_data([self.student_class_excel_file], 
                                     self.student_class_name_header, 
                                     self.student_class_class_header)
        
        success_count = 0
        fail_count = 0
        
        # 记录未能匹配的详细信息
        unmatched_students = []  # 在Excel中未找到的PDF_A
        unmatched_classes = []   # 找不到对应PDF_B的PDF_A
        
        for student_pdf in self.student_pdfs:
            student_filename = os.path.basename(student_pdf)
            student_name = os.path.splitext(student_filename)[0]
            if '_' in student_name:
                student_name = student_name.split('_')[0]
            
            # 查找PDF_B
            matched_class = None
            for name, class_name in excel_data.items():
                if name == student_name:
                    matched_class = class_name
                    break
            
            if matched_class:
                # 查找PDF_B
                class_pdf = None
                for pdf in self.class_pdfs:
                    if os.path.splitext(os.path.basename(pdf))[0] == matched_class:
                        class_pdf = pdf
                        break
                
                if class_pdf:
                    output_path = os.path.join(self.student_class_output_dir, 
                                               f"{os.path.splitext(student_filename)[0]}_合并.pdf")
                    
                    if PDFHandler.merge_pdfs([student_pdf, class_pdf], output_path):
                        success_count += 1
                    else:
                        fail_count += 1
                        unmatched_classes.append(f"{student_filename} - 合并失败")
                else:
                    fail_count += 1
                    unmatched_classes.append(f"{student_filename} - 在PDF_B中未找到【{matched_class}】")
            else:
                fail_count += 1
                unmatched_students.append(f"{student_filename} - 在Excel中未找到姓名【{student_name}】")
        
        # 构建结果消息
        result_msg = f"PDF_A与PDF_B合并完成！\n\n成功合并：{success_count}个\n失败：{fail_count}个"
        
        # 添加未匹配详细信息
        if unmatched_students:
            result_msg += f"\n\n【Excel中未找到的PDF_A】\n{chr(10).join(unmatched_students)}"
        
        if unmatched_classes:
            result_msg += f"\n\n【找不到对应PDF_B的PDF_A】\n{chr(10).join(unmatched_classes)}"
        
        # 检查是否有未使用的PDF_B
        used_classes = set()
        for name, class_name in excel_data.items():
            used_classes.add(class_name)
        
        unused_class_pdfs = []
        for pdf in self.class_pdfs:
            class_name = os.path.splitext(os.path.basename(pdf))[0]
            if class_name not in used_classes:
                unused_class_pdfs.append(f"{os.path.basename(pdf)}")
        
        if unused_class_pdfs:
            result_msg += f"\n\n【未使用的PDF_B】\n{chr(10).join(unused_class_pdfs)}"
        
        if fail_count > 0:
            show_warning(result_msg)
        else:
            show_info(result_msg)
    
    def preview_page_pdf(self):
        """预览PDF文件"""
        if not self.page_pdfs:
            show_error("请选择PDF文件")
            return
        
        # 使用第一个PDF文件进行预览
        pdf_path = self.page_pdfs[0]
        pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
        self.show_pdf_preview(pdf_path, f"PDF预览 - {pdf_name}")
    
    def preview_excel_to_pdf(self):
        """预览Excel转PDF效果"""
        if not self.excel_files:
            show_error("请选择Excel文件")
            return
        
        # 获取打印设置参数
        orientation = self.excel_orientation_var.get()
        scale_mode = self.excel_scale_mode_var.get()
        scale = int(self.excel_scale_var.get())
        print_active_sheet = self.excel_print_active_sheet_var.get()
        paper_size = self.excel_paper_size_var.get()
        
        # 解析页码范围
        first_page = int(self.excel_first_page_var.get()) if self.excel_first_page_var.get().strip() else None
        last_page = int(self.excel_last_page_var.get()) if self.excel_last_page_var.get().strip() else None
        
        # 使用第一个Excel文件进行预览
        excel_file = self.excel_files[0]
        
        # 创建临时目录用于预览
        import tempfile
        with tempfile.TemporaryDirectory() as temp_dir:
            # 生成临时PDF路径
            excel_name = os.path.splitext(os.path.basename(excel_file))[0]
            temp_pdf_path = os.path.join(temp_dir, f"{excel_name}_预览.pdf")
            
            try:
                # 转换为PDF
                result = OfficeConverter.excel_to_pdf(
                    excel_file, 
                    temp_pdf_path,
                    orientation=orientation,
                    scale_mode=scale_mode,
                    scale=scale,
                    print_active_sheet=print_active_sheet,
                    paper_size=paper_size,
                    first_page=first_page,
                    last_page=last_page
                )
                
                if result:
                    # 显示预览
                    self.show_pdf_preview(temp_pdf_path, f"Excel转PDF预览 - {excel_name}")
                else:
                    show_error("预览转换失败")
            except Exception as e:
                show_error(f"预览失败: {str(e)}")
    
    def process_excel_to_pdf(self):
        """处理Excel转PDF"""
        if not self.excel_files:
            show_error("请选择Excel文件")
            return
        
        if not self.excel_output_dir:
            show_error("请选择输出目录")
            return
        
        # 获取打印设置参数
        orientation = self.excel_orientation_var.get()
        scale_mode = self.excel_scale_mode_var.get()
        scale = int(self.excel_scale_var.get())
        print_active_sheet = self.excel_print_active_sheet_var.get()
        paper_size = self.excel_paper_size_var.get()
        
        # 解析页码范围
        first_page = int(self.excel_first_page_var.get()) if self.excel_first_page_var.get().strip() else None
        last_page = int(self.excel_last_page_var.get()) if self.excel_last_page_var.get().strip() else None
        
        results = OfficeConverter.batch_excel_to_pdf(
            self.excel_files, 
            self.excel_output_dir,
            progress_callback=lambda p: self.status_var.set(f"处理中... {p:.1f}%"),
            orientation=orientation,
            scale_mode=scale_mode,
            scale=scale,
            print_active_sheet=print_active_sheet,
            paper_size=paper_size,
            first_page=first_page,
            last_page=last_page
        )
        
        if results['failed'] > 0:
            show_warning(f"Excel转PDF完成！成功：{results['success']}，失败：{results['failed']}")
        else:
            show_info(f"Excel转PDF完成！成功处理 {results['success']} 个文件")
    
    def process_word_to_pdf(self):
        """处理Word转PDF"""
        if not self.word_files:
            show_error("请选择Word文件")
            return
        
        if not self.word_output_dir:
            show_error("请选择输出目录")
            return
        
        page_range = None if self.word_export_all_pages_var.get() else self.word_page_range_var.get()
        
        results = OfficeConverter.batch_word_to_pdf(
            self.word_files, 
            self.word_output_dir,
            page_range=page_range,
            progress_callback=lambda p: self.status_var.set(f"处理中... {p:.1f}%")
        )
        
        if results['failed'] > 0:
            show_warning(f"Word转PDF完成！成功：{results['success']}，失败：{results['failed']}")
        else:
            show_info(f"Word转PDF完成！成功处理 {results['success']} 个文件")
    
    def clear_all(self):
        """清空所有设置"""
        # 重置所有变量
        self.init_variables()
        
        # 更新状态
        self.status_var.set("就绪")
        show_info("已清空所有设置")
    
    def save_settings(self):
        """保存设置到文件"""
        settings = {
            'function': self.function_var.get(),
            'export_pages': self.export_pages_var.get(),
            'split_pages': self.split_pages_var.get(),
            'interval': self.interval_var.get(),
            'insert_pages': self.insert_pages_var.get()
        }
        
        try:
            import json
            with open('pdf_editor_settings.json', 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
            show_info("设置保存成功")
        except Exception as e:
            show_error(f"保存设置失败: {str(e)}")
    
    def load_settings(self):
        """从文件加载设置"""
        try:
            import json
            with open('pdf_editor_settings.json', 'r', encoding='utf-8') as f:
                settings = json.load(f)
            
            self.function_var.set(settings.get('function', 'insert'))
            self.export_pages_var.set(settings.get('export_pages', '1-6,1,2,3,4'))
            self.split_pages_var.set(settings.get('split_pages', '10'))
            self.interval_var.set(settings.get('interval', '1'))
            self.insert_pages_var.set(settings.get('insert_pages', '1'))
            
            self.on_function_change()
            show_info("设置加载成功")
        except FileNotFoundError:
            show_warning("未找到设置文件")
        except Exception as e:
            show_error(f"加载设置失败: {str(e)}")
    
    def show_help(self):
        """显示使用说明"""
        help_window = tk.Toplevel(self.root)
        help_window.title("使用说明")
        help_window.geometry("800x600")
        
        text = scrolledtext.ScrolledText(help_window, font=('微软雅黑', 10))
        text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        help_text = """PDF操作工具使用说明

一、功能介绍

1. PDF插入：在PDF文件中按指定间隔插入其他PDF页面
2. PDF分割：将PDF文件按指定页数分割成多个文件
3. PDF页面交换：交换PDF文件中的指定页面
4. 指定导出页：从PDF文件中导出指定页面（支持批量处理）
5. PDF批量重命名：批量修改PDF文件名
6. PDF重排序：按Excel数据对PDF文件进行排序
7. PDF页面：反转页面顺序或旋转页面
8. PDF页面逆序：将PDF文件的页面顺序反转（支持批量处理）
9. 交叉合并：将两个PDF文件按页交叉合并（用于扫描件正反面拼接）
10. 批量提取：从多个PDF文件中提取指定页面
11. 批量追加页：将源PDF页面追加到多个PDF文件
12. PDF合并：将多个PDF文件合并到一个主PDF后面
13. 批量单页追加：从多个PDF中各提取一页并批量追加到主PDF
14. PDF_A与PDF_B合并：根据Excel数据匹配PDF_A和PDF_B
15. Excel批量导出PDF：将Excel文件转换为PDF
16. Word批量导出PDF：将Word文件转换为PDF

二、功能详解

1. PDF插入：选择一个主PDF和多个子PDF，设置插入间隔和每页插入数量
2. PDF分割：选择PDF文件，设置分割页数，支持数字编号或Excel命名
3. PDF页面交换：选择PDF文件，输入要交换的两个页码
4. 指定导出页：选择多个PDF文件，输入要导出的页面范围
5. PDF批量重命名：选择多个PDF文件，选择Excel文件进行匹配命名
6. PDF重排序：选择多个PDF文件，选择Excel文件按指定列排序
7. PDF页面：选择多个PDF文件，选择反转顺序、旋转页面（可设置角度）或交换指定页面
8. PDF页面逆序：选择多个PDF文件，反转页面顺序，输出为原文件名_reversed.pdf
9. 交叉合并：选择两个PDF文件（A正面扫描、B背面扫描已逆序），设置每次各取页数和合并份数
10. 批量提取：选择多个PDF文件，提取指定页码到输出目录
11. 批量追加页：选择多个目标PDF和一个源PDF，将源PDF页面追加到每个目标PDF
12. PDF合并：选择一个主PDF，通过+-按钮添加多个要追加的PDF文件
13. 批量单页追加：选择多个主PDF和多个追加PDF，从每个追加PDF提取一页追加到主PDF（循环匹配）
14. PDF_A与PDF_B合并：选择PDF_A文件、PDF_B文件和Excel匹配文件进行合并
15. Excel批量导出PDF：选择多个Excel文件，支持选择页码范围
16. Word批量导出PDF：选择多个Word文件，支持选择页码范围

三、交叉合并功能说明

适用于扫描件正反面拼接场景：
- PDF_A：正面扫描件（如封面正面、申请表正面）
- PDF_B：背面扫描件，需先使用"PDF页面逆序"处理
- 合并模式：A取N页 → B取M页 → A取N页 → B取M页...

示例（79人档案，每人4页）：
- PDF_A：158页（79人×2页：封面正面、申请表正面）
- PDF_B：158页（已逆序，79人×2页：封面背面、申请表背面）
- 设置：A每次取2页，B每次取2页，共79份
- 输出：封面正面→封面背面→申请表正面→申请表背面（每人4页）

四、使用注意事项

1. Excel文件支持：.xlsx 和 .xls 格式
2. Word文件支持：.docx 和 .doc 格式
3. 批量操作时，Excel文件为可选选项
4. 所有操作都会保留原始文件，输出到指定目录
5. 如果操作失败，会显示详细的错误信息
6. 处理大量文件时，建议分批次处理以避免内存不足
7. 批量单页追加和交叉合并功能会在终端打印详细日志

五、页面范围格式

支持以下格式：
- 单页：1
- 连续页面：1-10
- 多个页面：1,3,5
- 混合格式：1-5,7,9-12

六、提示

- 所有功能都支持批量处理多个文件
- Excel匹配功能可以让输出文件名更规范
- 处理大文件时请耐心等待，状态栏会显示进度
- 交叉合并和批量单页追加功能提供预览按钮
- 建议在处理前使用预览功能确认参数正确
"""
        
        text.insert(tk.END, help_text)
        text.config(state=tk.DISABLED)
    
    def show_pdf_preview(self, pdf_path, title="PDF预览"):
        """显示PDF预览窗口"""
        if not pdf_path or not os.path.exists(pdf_path):
            show_error("PDF文件不存在")
            return
        
        try:
            # 打开PDF文档
            doc = fitz.open(pdf_path)
            total_pages = doc.page_count
            
            if total_pages == 0:
                show_error("PDF文件为空")
                return
            
            # 创建预览窗口
            preview_window = tk.Toplevel(self.root)
            preview_window.title(title)
            preview_window.geometry("900x700")
            preview_window.transient(self.root)
            preview_window.grab_set()
            
            # 创建主框架
            main_frame = ttk.Frame(preview_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # 创建工具栏
            toolbar = ttk.Frame(main_frame)
            toolbar.pack(fill=tk.X, pady=5)
            
            # 翻页按钮
            prev_btn = ttk.Button(toolbar, text="上一页", command=lambda: self._preview_prev_page(doc, canvas, page_label, current_page_var))
            prev_btn.pack(side=tk.LEFT, padx=5)
            
            # 当前页码显示
            current_page_var = tk.IntVar(value=1)
            page_label = ttk.Label(toolbar, textvariable=tk.StringVar(value=f"第 {current_page_var.get()} / {total_pages} 页"))
            page_label.pack(side=tk.LEFT, padx=10)
            
            next_btn = ttk.Button(toolbar, text="下一页", command=lambda: self._preview_next_page(doc, canvas, page_label, current_page_var, total_pages))
            next_btn.pack(side=tk.LEFT, padx=5)
            
            # 页码输入框
            page_entry = ttk.Entry(toolbar, width=5)
            page_entry.pack(side=tk.LEFT, padx=5)
            
            jump_btn = ttk.Button(toolbar, text="跳转", command=lambda: self._preview_jump_to_page(doc, canvas, page_label, current_page_var, total_pages, page_entry.get()))
            jump_btn.pack(side=tk.LEFT, padx=5)
            
            # 缩放控制
            zoom_var = tk.DoubleVar(value=1.0)
            zoom_scale = ttk.Scale(toolbar, from_=0.5, to=2.0, variable=zoom_var, command=lambda val: self._preview_zoom(doc, canvas, current_page_var.get() - 1, float(val)))
            zoom_scale.pack(side=tk.RIGHT, padx=5)
            
            zoom_label = ttk.Label(toolbar, text="缩放")
            zoom_label.pack(side=tk.RIGHT, padx=5)
            
            # 旋转控制
            rotate_left_btn = ttk.Button(toolbar, text="左旋90°", command=lambda: self._preview_rotate(doc, canvas, current_page_var.get() - 1, -90))
            rotate_left_btn.pack(side=tk.RIGHT, padx=5)
            
            rotate_right_btn = ttk.Button(toolbar, text="右旋90°", command=lambda: self._preview_rotate(doc, canvas, current_page_var.get() - 1, 90))
            rotate_right_btn.pack(side=tk.RIGHT, padx=5)
            
            # 创建画布用于显示PDF
            canvas_frame = ttk.Frame(main_frame)
            canvas_frame.pack(fill=tk.BOTH, expand=True)
            
            canvas = tk.Canvas(canvas_frame, bg="white")
            canvas.pack(fill=tk.BOTH, expand=True)
            
            # 绑定键盘事件
            preview_window.bind("<Left>", lambda e: self._preview_prev_page(doc, canvas, page_label, current_page_var))
            preview_window.bind("<Right>", lambda e: self._preview_next_page(doc, canvas, page_label, current_page_var, total_pages))
            preview_window.bind("<Prior>", lambda e: self._preview_prev_page(doc, canvas, page_label, current_page_var))
            preview_window.bind("<Next>", lambda e: self._preview_next_page(doc, canvas, page_label, current_page_var, total_pages))
            
            # 渲染第一页
            self._preview_render_page(doc, canvas, 0, zoom_var.get())
            page_label.config(text=f"第 1 / {total_pages} 页")
            
            # 关闭窗口时清理资源
            preview_window.protocol("WM_DELETE_WINDOW", lambda: self._preview_close(doc, preview_window))
            
        except Exception as e:
            show_error(f"打开PDF失败: {str(e)}")
    
    def _preview_render_page(self, doc, canvas, page_num, zoom=1.0):
        """渲染PDF页面到画布"""
        if page_num < 0 or page_num >= doc.page_count:
            return
        
        page = doc[page_num]
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
        
        # 转换为Tkinter可用的图像格式
        img_data = pix.tobytes("ppm")
        img = tk.PhotoImage(data=img_data, format="ppm")
        
        # 清空画布并显示图像
        canvas.delete("all")
        canvas.image = img  # 保持引用防止被垃圾回收
        canvas.create_image(0, 0, anchor=tk.NW, image=img)
        
        # 设置画布滚动区域
        canvas.config(scrollregion=canvas.bbox("all"))
    
    def _preview_prev_page(self, doc, canvas, page_label, current_page_var):
        """上一页"""
        if current_page_var.get() > 1:
            current_page_var.set(current_page_var.get() - 1)
            self._preview_render_page(doc, canvas, current_page_var.get() - 1)
            page_label.config(text=f"第 {current_page_var.get()} / {doc.page_count} 页")
    
    def _preview_next_page(self, doc, canvas, page_label, current_page_var, total_pages):
        """下一页"""
        if current_page_var.get() < total_pages:
            current_page_var.set(current_page_var.get() + 1)
            self._preview_render_page(doc, canvas, current_page_var.get() - 1)
            page_label.config(text=f"第 {current_page_var.get()} / {total_pages} 页")
    
    def _preview_jump_to_page(self, doc, canvas, page_label, current_page_var, total_pages, page_str):
        """跳转到指定页"""
        try:
            page_num = int(page_str)
            if 1 <= page_num <= total_pages:
                current_page_var.set(page_num)
                self._preview_render_page(doc, canvas, page_num - 1)
                page_label.config(text=f"第 {page_num} / {total_pages} 页")
            else:
                show_warning(f"页码必须在1-{total_pages}之间")
        except ValueError:
            show_error("请输入有效的页码")
    
    def _preview_zoom(self, doc, canvas, page_num, zoom):
        """缩放预览"""
        self._preview_render_page(doc, canvas, page_num, zoom)
    
    def _preview_rotate(self, doc, canvas, page_num, angle):
        """旋转页面"""
        page = doc[page_num]
        page.set_rotation((page.rotation + angle) % 360)
        self._preview_render_page(doc, canvas, page_num)
    
    def _preview_close(self, doc, window):
        """关闭预览窗口"""
        doc.close()
        window.destroy()


def main():
    """主程序入口"""
    root = tk.Tk()
    app = PDFEditorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()